import csv
from io import BytesIO, StringIO
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
from urllib.parse import urlencode
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.core.management.base import CommandError
from django.db import IntegrityError
from django.test import SimpleTestCase, TestCase
from django.urls import reverse
from django.utils import timezone
from decimal import Decimal
from openpyxl import load_workbook
from unittest.mock import patch

from auditoria.models import AuditLog
from finanzas.documentos.services import parse_tax_document
from finanzas.documentos.temp_storage import SESSION_KEY
from finanzas.forms import DocumentoTributarioForm, PaymentForm, TransactionForm
from finanzas.services import asociar_asistencia_a_pago, resumen_financiero_estudiante
from finanzas.services.reconciliacion import reconciliar_integridad_dominio
from finanzas.services.reversas import revertir_pago
from finanzas.services.pagos import (
    confirmar_lote_pagos,
    crear_persona_estudiante_desde_modal,
    enriquecer_pagos_para_listado,
    resumen_consumos_pago,
    texto_copiable_operativo_pago,
)

from asistencias.forms import PersonaRapidaForm
from asistencias.models import Asistencia, ClaseLiberada, Disciplina, SesionClase
from personas.models import Organizacion, Persona, PersonaRol, Rol
from personas.test_factories import crear_usuario_con_rol

from finanzas.models import (
    AttendanceConsumption,
    Category,
    DocumentoTributario,
    Payment,
    PaymentPlan,
    Transaction,
    LotePago,
)


TEST_PASSWORD = "not-a-real-test-password"


class FinanzasServicesCompatibilityTests(SimpleTestCase):
    def test_imports_publicos_antiguos_siguen_disponibles(self):
        from finanzas.services import (
            asignar_consumo_asistencia,
            asociar_asistencia_a_pago,
            imputar_pago_a_deudas,
            resumen_financiero_estudiante,
            resumen_financiero_estudiante_periodo,
        )

        self.assertTrue(callable(asignar_consumo_asistencia))
        self.assertTrue(callable(asociar_asistencia_a_pago))
        self.assertTrue(callable(imputar_pago_a_deudas))
        self.assertTrue(callable(resumen_financiero_estudiante))
        self.assertTrue(callable(resumen_financiero_estudiante_periodo))


class FinanzasAccessTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.org = Organizacion.objects.create(
            nombre="Org Finanzas",
            razon_social="Org Finanzas SPA",
            rut="22.222.222-2",
        )
        self.rol_admin = Rol.objects.create(nombre="Administrador", codigo="ADMINISTRADOR")
        self.rol_estudiante = Rol.objects.create(nombre="Estudiante", codigo="ESTUDIANTE")
        self.rol_finanzas = Rol.objects.create(nombre="Finanzas", codigo="FINANZAS")
        self.rol_profesor = Rol.objects.create(nombre="Profesor", codigo="PROFESOR")
        self.rol_solo_lectura = Rol.objects.create(nombre="Solo lectura", codigo="SOLO_LECTURA")

        self.user_admin = User.objects.create_user("admin_fin", password=TEST_PASSWORD)
        self.persona_admin = Persona.objects.create(
            nombres="Admin",
            apellidos="Fin",
            email="adminfin@example.com",
            user=self.user_admin,
        )
        PersonaRol.objects.create(
            persona=self.persona_admin,
            rol=self.rol_admin,
            organizacion=self.org,
            activo=True,
        )

        self.user_no_admin = User.objects.create_user("noadmin_fin", password=TEST_PASSWORD)
        self.persona_no_admin = Persona.objects.create(
            nombres="No",
            apellidos="Admin",
            email="noadmin@example.com",
            user=self.user_no_admin,
        )
        PersonaRol.objects.create(
            persona=self.persona_no_admin,
            rol=self.rol_estudiante,
            organizacion=self.org,
            activo=True,
        )
        self.user_finanzas = User.objects.create_user("finanzas_user", password=TEST_PASSWORD)
        self.persona_finanzas = Persona.objects.create(
            nombres="Finanzas",
            apellidos="User",
            email="finanzas@example.com",
            user=self.user_finanzas,
        )
        PersonaRol.objects.create(
            persona=self.persona_finanzas,
            rol=self.rol_finanzas,
            organizacion=self.org,
            activo=True,
        )
        self.user_profesor = User.objects.create_user("profesor_fin", password=TEST_PASSWORD)
        self.persona_profesor = Persona.objects.create(
            nombres="Profesor",
            apellidos="Fin",
            email="profesorfin@example.com",
            user=self.user_profesor,
        )
        PersonaRol.objects.create(
            persona=self.persona_profesor,
            rol=self.rol_profesor,
            organizacion=self.org,
            activo=True,
        )
        self.user_solo_lectura = User.objects.create_user("lectura_fin", password=TEST_PASSWORD)
        self.persona_solo_lectura = Persona.objects.create(
            nombres="Lectura",
            apellidos="Fin",
            email="lecturafin@example.com",
            user=self.user_solo_lectura,
        )
        PersonaRol.objects.create(
            persona=self.persona_solo_lectura,
            rol=self.rol_solo_lectura,
            organizacion=self.org,
            activo=True,
        )
        self.user_sin_rol = User.objects.create_user("sinrol_fin", password=TEST_PASSWORD)
        self.persona_sin_rol = Persona.objects.create(
            nombres="Sin",
            apellidos="Rol",
            email="sinrolfin@example.com",
            user=self.user_sin_rol,
        )

    def _url_con_organizacion(self, url, **params):
        return f"{url}?{urlencode({'organizacion': self.org.pk, **params})}"

    def test_finanzas_dashboard_requiere_admin(self):
        self.client.force_login(self.user_no_admin)
        response = self.client.get(reverse("finanzas:dashboard"))
        self.assertEqual(response.status_code, 403)

    def test_finanzas_dashboard_admin_ok(self):
        self.client.force_login(self.user_admin)
        response = self.client.get(reverse("finanzas:dashboard"), {"organizacion": self.org.pk})
        self.assertEqual(response.status_code, 200)

    def test_usuario_finanzas_accede_a_pagos_documentos_y_transacciones(self):
        self.client.force_login(self.user_finanzas)
        for url_name in (
            "finanzas:pagos_list",
            "finanzas:documentos_tributarios_list",
            "finanzas:transacciones_list",
        ):
            response = self.client.get(reverse(url_name), {"organizacion": self.org.pk})
            self.assertEqual(response.status_code, 200)

    def test_listado_pagos_busca_persona_por_fragmentos_sin_tildes(self):
        persona = Persona.objects.create(
            nombres="Bárbara Inés",
            apellidos="Muñoz Cáceres",
            email="barbara.munoz@example.com",
        )
        PersonaRol.objects.create(
            persona=persona,
            rol=self.rol_estudiante,
            organizacion=self.org,
            activo=True,
        )
        Payment.objects.create(
            persona=persona,
            organizacion=self.org,
            fecha_pago="2026-08-10",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=10000,
        )
        self.client.force_login(self.user_finanzas)

        response = self.client.get(
            reverse("finanzas:pagos_list"),
            {
                "organizacion": self.org.pk,
                "periodo_mes": 8,
                "periodo_anio": 2026,
                "q": "barbara munoz",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Bárbara Inés Muñoz Cáceres")

    def test_profesor_no_accede_a_finanzas_completa(self):
        self.client.force_login(self.user_profesor)
        response = self.client.get(reverse("finanzas:dashboard"), {"organizacion": self.org.pk})
        self.assertEqual(response.status_code, 403)

    def test_solo_lectura_no_puede_hacer_post_sensible(self):
        self.client.force_login(self.user_solo_lectura)
        response = self.client.post(reverse("finanzas:pagos_list"), {"organizacion": self.org.pk})
        self.assertEqual(response.status_code, 403)

    def test_usuario_sin_rol_no_accede_a_vistas_sensibles(self):
        self.client.force_login(self.user_sin_rol)
        response = self.client.get(reverse("finanzas:pagos_list"), {"organizacion": self.org.pk})
        self.assertEqual(response.status_code, 403)

    def test_permiso_finanzas_considera_organizacion_filtrada(self):
        otra_org = Organizacion.objects.create(
            nombre="Org Sin Permiso",
            razon_social="Org Sin Permiso SPA",
            rut="99.999.999-9",
        )
        self.client.force_login(self.user_finanzas)
        response = self.client.get(reverse("finanzas:pagos_list"), {"organizacion": otra_org.pk})
        self.assertEqual(response.status_code, 403)

    def test_dashboard_mensual_usa_transacciones_como_fuente_contable(self):
        categoria_ingreso = Category.objects.create(nombre="Ingreso dashboard", tipo=Category.Tipo.INGRESO, activa=True)
        categoria_egreso = Category.objects.create(nombre="Egreso dashboard", tipo=Category.Tipo.EGRESO, activa=True)
        Payment.objects.create(
            persona=self.persona_no_admin,
            organizacion=self.org,
            fecha_pago="2026-02-10",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=100000,
            clases_asignadas=4,
        )
        Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria_ingreso,
            fecha="2026-02-10",
            tipo=Transaction.Tipo.INGRESO,
            monto=25000,
            descripcion="Ingreso contable",
        )
        Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria_egreso,
            fecha="2026-02-11",
            tipo=Transaction.Tipo.EGRESO,
            monto=5000,
            descripcion="Egreso contable",
        )

        self.client.force_login(self.user_finanzas)
        response = self.client.get(
            reverse("finanzas:dashboard"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["ingresos_contables"], 25000)
        self.assertEqual(response.context["egresos_contables"], 5000)
        self.assertEqual(response.context["saldo_contable"], 20000)
        self.assertEqual(response.context["pagos_operacionales_monto"], 100000)
        self.assertEqual(response.context["total_transacciones"], 2)

    def test_dashboard_mensual_respeta_organizacion_y_periodo(self):
        categoria = Category.objects.create(nombre="Ingreso filtro dashboard", tipo=Category.Tipo.INGRESO, activa=True)
        otra_org = Organizacion.objects.create(nombre="Org Dashboard", razon_social="Org Dashboard SPA", rut="98.888.888-8")
        Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-02-10",
            tipo=Transaction.Tipo.INGRESO,
            monto=25000,
            descripcion="Incluida",
        )
        Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-03-10",
            tipo=Transaction.Tipo.INGRESO,
            monto=99999,
            descripcion="Fuera periodo",
        )
        Transaction.objects.create(
            organizacion=otra_org,
            categoria=categoria,
            fecha="2026-02-10",
            tipo=Transaction.Tipo.INGRESO,
            monto=99999,
            descripcion="Fuera organizacion",
        )

        self.client.force_login(self.user_admin)
        response = self.client.get(
            reverse("finanzas:dashboard"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["ingresos_contables"], 25000)
        self.assertEqual(response.context["total_transacciones"], 1)

    def test_dashboard_finanzas_admin_y_finanzas_ven_acciones_operativas(self):
        params = {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk}
        for usuario in (self.user_admin, self.user_finanzas):
            self.client.force_login(usuario)
            response = self.client.get(reverse("finanzas:dashboard"), params)

            self.assertEqual(response.status_code, 200)
            self.assertContains(response, "Agregar pago")
            self.assertContains(response, "Agregar documento")
            self.assertContains(response, "Agregar transacción")
            self.assertContains(response, reverse("finanzas:pagos_list"))
            self.assertContains(response, "open=registrar_pago")
            self.assertContains(response, reverse("finanzas:documento_tributario_importar"))
            self.assertContains(response, reverse("finanzas:transacciones_list"))
            self.assertContains(response, "open=nueva_transaccion")
            self.assertContains(response, f"periodo_mes={params['periodo_mes']}")
            self.assertContains(response, f"periodo_anio={params['periodo_anio']}")
            self.assertContains(response, f"organizacion={params['organizacion']}")

    def test_dashboard_finanzas_solo_lectura_no_ve_acciones_mutables(self):
        self.client.force_login(self.user_solo_lectura)
        response = self.client.get(
            reverse("finanzas:dashboard"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Agregar pago")
        self.assertNotContains(response, "Agregar documento")
        self.assertNotContains(response, "Agregar transacción")

    def test_dashboard_finanzas_profesor_no_accede_a_acciones_financieras(self):
        self.client.force_login(self.user_profesor)
        response = self.client.get(
            reverse("finanzas:dashboard"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 403)

    def test_transacciones_list_abre_modal_desde_dashboard(self):
        self.client.force_login(self.user_finanzas)
        response = self.client.get(
            reverse("finanzas:transacciones_list"),
            {
                "periodo_mes": 2,
                "periodo_anio": 2026,
                "organizacion": self.org.pk,
                "open": "nueva_transaccion",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["open_nueva_transaccion"])

    def test_acciones_dashboard_no_crean_modelos_al_abrir_flujos(self):
        self.client.force_login(self.user_finanzas)
        acciones = [
            (reverse("finanzas:pagos_list"), {"open": "registrar_pago"}),
            (reverse("finanzas:transacciones_list"), {"open": "nueva_transaccion"}),
            (reverse("finanzas:documento_tributario_importar"), {}),
        ]
        for url, extra in acciones:
            params = {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk, **extra}
            response = self.client.get(url, params)
            self.assertEqual(response.status_code, 200)

        self.assertEqual(Payment.objects.count(), 0)
        self.assertEqual(Transaction.objects.count(), 0)
        self.assertEqual(DocumentoTributario.objects.count(), 0)

    def test_libro_caja_csv_exporta_solo_transacciones_ordenadas(self):
        categoria_ingreso = Category.objects.create(nombre="Ingreso libro", tipo=Category.Tipo.INGRESO, activa=True)
        categoria_egreso = Category.objects.create(nombre="Egreso libro", tipo=Category.Tipo.EGRESO, activa=True)
        documento = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_AFECTA,
            folio="LC-1",
            fecha_emision="2026-02-02",
            monto_total=15000,
        )
        transaccion_2 = Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria_egreso,
            fecha="2026-02-03",
            tipo=Transaction.Tipo.EGRESO,
            monto=5000,
            descripcion="Egreso libro",
        )
        transaccion_1 = Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria_ingreso,
            fecha="2026-02-02",
            tipo=Transaction.Tipo.INGRESO,
            monto=15000,
            descripcion="Ingreso libro",
        )
        transaccion_1.documentos_tributarios.add(documento)
        Payment.objects.create(
            persona=self.persona_no_admin,
            organizacion=self.org,
            fecha_pago="2026-02-02",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=99000,
            clases_asignadas=1,
        )

        self.client.force_login(self.user_finanzas)
        response = self.client.get(
            reverse("finanzas:export_libro_caja_csv"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        rows = list(csv.reader(response.content.decode("utf-8-sig").splitlines()))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(rows[0][0], "numero correlativo")
        self.assertIn("Msg", rows[0])
        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[1][0], "1")
        self.assertEqual(rows[2][0], "2")
        self.assertEqual(rows[1][1], "2026-02-02")
        self.assertEqual(rows[1][4], "Ingreso libro")
        self.assertIn("Factura afecta #LC-1", rows[1][7])
        self.assertIn("Ingreso libro", rows[1][8])
        self.assertEqual(rows[2][1], "2026-02-03")
        self.assertNotIn("99000", response.content.decode("utf-8-sig"))

    def test_libro_caja_csv_respeta_organizacion_y_periodo(self):
        categoria = Category.objects.create(nombre="Ingreso libro filtros", tipo=Category.Tipo.INGRESO, activa=True)
        otra_org = Organizacion.objects.create(nombre="Org Libro", razon_social="Org Libro SPA", rut="97.777.777-7")
        Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-02-02",
            tipo=Transaction.Tipo.INGRESO,
            monto=15000,
            descripcion="Incluida",
        )
        Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-03-02",
            tipo=Transaction.Tipo.INGRESO,
            monto=99000,
            descripcion="Fuera periodo",
        )
        Transaction.objects.create(
            organizacion=otra_org,
            categoria=categoria,
            fecha="2026-02-02",
            tipo=Transaction.Tipo.INGRESO,
            monto=88000,
            descripcion="Fuera organizacion",
        )

        self.client.force_login(self.user_admin)
        response = self.client.get(
            reverse("finanzas:export_libro_caja_csv"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        contenido = response.content.decode("utf-8-sig")
        rows = list(csv.reader(contenido.splitlines()))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(rows), 2)
        self.assertIn("Incluida", contenido)
        self.assertNotIn("Fuera periodo", contenido)
        self.assertNotIn("Fuera organizacion", contenido)

    def test_libro_caja_csv_bloquea_periodo_todos(self):
        self.client.force_login(self.user_admin)
        response = self.client.get(
            reverse("finanzas:export_libro_caja_csv"),
            {"periodo_mes": "todos", "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 400)
        self.assertContains(response, "requiere seleccionar un mes y un año", status_code=400)

    def test_libro_caja_csv_permisos(self):
        self.client.force_login(self.user_sin_rol)
        response = self.client.get(
            reverse("finanzas:export_libro_caja_csv"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )
        self.assertEqual(response.status_code, 403)

        self.client.force_login(self.user_profesor)
        response = self.client.get(
            reverse("finanzas:export_libro_caja_csv"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )
        self.assertEqual(response.status_code, 403)

        self.client.force_login(self.user_finanzas)
        response = self.client.get(
            reverse("finanzas:export_libro_caja_csv"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )
        self.assertEqual(response.status_code, 200)

    def _xlsx_rows(self, response):
        workbook = load_workbook(BytesIO(response.content))
        return list(workbook.active.iter_rows(values_only=True))

    def test_export_pagos_alumnos_xlsx_respeta_periodo_organizacion_y_no_es_contable(self):
        otra_org = Organizacion.objects.create(nombre="Org Pagos XLSX", razon_social="Org Pagos XLSX SPA", rut="76.111.111-1")
        Payment.objects.create(
            persona=self.persona_no_admin,
            organizacion=self.org,
            fecha_pago="2026-02-04",
            metodo_pago=Payment.Metodo.TRANSFERENCIA,
            aplica_iva=False,
            monto_referencia=45000,
            clases_asignadas=4,
            observaciones="Pago incluido",
        )
        Payment.objects.create(
            persona=self.persona_no_admin,
            organizacion=self.org,
            fecha_pago="2026-03-04",
            metodo_pago=Payment.Metodo.TRANSFERENCIA,
            aplica_iva=False,
            monto_referencia=99000,
            clases_asignadas=4,
            observaciones="Fuera periodo",
        )
        Payment.objects.create(
            persona=self.persona_no_admin,
            organizacion=otra_org,
            fecha_pago="2026-02-04",
            metodo_pago=Payment.Metodo.TRANSFERENCIA,
            aplica_iva=False,
            monto_referencia=88000,
            clases_asignadas=4,
            observaciones="Fuera organizacion",
        )

        self.client.force_login(self.user_finanzas)
        response = self.client.get(
            reverse("finanzas:export_pagos_alumnos_xlsx"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        rows = self._xlsx_rows(response)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(rows[0][0], "Fecha pago")
        self.assertIn("Estado", rows[0])
        contenido = str(rows)
        self.assertIn("Pago incluido", contenido)
        self.assertNotIn("Fuera periodo", contenido)
        self.assertNotIn("Fuera organizacion", contenido)

    def test_export_transacciones_xlsx_es_contable_y_no_duplica_pagos(self):
        categoria = Category.objects.create(nombre="Ingreso XLSX", tipo=Category.Tipo.INGRESO, activa=True)
        Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-02-05",
            tipo=Transaction.Tipo.INGRESO,
            monto=12000,
            descripcion="Transaccion real",
        )
        Payment.objects.create(
            persona=self.persona_no_admin,
            organizacion=self.org,
            fecha_pago="2026-02-05",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=99000,
            clases_asignadas=1,
            observaciones="Pago operacional no contable",
        )

        self.client.force_login(self.user_admin)
        response = self.client.get(
            reverse("finanzas:export_transacciones_xlsx"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        rows = self._xlsx_rows(response)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(rows[0][0], "Fecha")
        self.assertIn("Msg", rows[0])
        contenido = str(rows)
        self.assertIn("Transaccion real", contenido)
        self.assertNotIn("Pago operacional no contable", contenido)

    def test_export_pagos_profesores_xlsx_usa_fuente_operacional_existente(self):
        disciplina = Disciplina.objects.create(organizacion=self.org, nombre="Danza XLSX")
        sesion = SesionClase.objects.create(
            disciplina=disciplina,
            fecha="2026-02-06",
            estado=SesionClase.Estado.COMPLETADA,
        )
        sesion.profesores.add(self.persona_profesor)
        PersonaRol.objects.filter(persona=self.persona_profesor, rol=self.rol_profesor, organizacion=self.org).update(
            valor_clase=Decimal("10000"),
            retencion_sii=Decimal("10"),
        )
        Asistencia.objects.create(sesion=sesion, persona=self.persona_no_admin, estado=Asistencia.Estado.PRESENTE)

        self.client.force_login(self.user_finanzas)
        response = self.client.get(
            reverse("finanzas:export_pagos_profesores_xlsx"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        rows = self._xlsx_rows(response)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(rows[0][0], "Periodo")
        contenido = str(rows)
        self.assertIn("Profesor Fin", contenido)
        self.assertIn("Estimado operacional", contenido)
        self.assertIn("9000", contenido)

    def test_exports_financieros_xlsx_bloquean_roles_no_autorizados(self):
        for url_name in (
            "finanzas:export_pagos_alumnos_xlsx",
            "finanzas:export_pagos_profesores_xlsx",
            "finanzas:export_transacciones_xlsx",
        ):
            self.client.force_login(self.user_solo_lectura)
            response = self.client.get(reverse(url_name), {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk})
            self.assertEqual(response.status_code, 403)

            self.client.force_login(self.user_profesor)
            response = self.client.get(reverse(url_name), {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk})
            self.assertEqual(response.status_code, 403)

            self.client.force_login(self.user_finanzas)
            response = self.client.get(reverse(url_name), {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk})
            self.assertEqual(response.status_code, 200)

    def test_pagos_list_permita_abrir_modal_con_estudiante_preseleccionado(self):
        self.client.force_login(self.user_finanzas)
        response = self.client.get(
            reverse("finanzas:pagos_list"),
            {
                "periodo_mes": 2,
                "periodo_anio": 2026,
                "organizacion": self.org.pk,
                "persona": self.persona_no_admin.pk,
                "open": "registrar_pago",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["open_registrar_pago"])
        self.assertEqual(str(response.context["form"].initial["persona"]), str(self.persona_no_admin.pk))

    def test_pagos_list_crea_pago_y_auditlog(self):
        self.client.force_login(self.user_admin)

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse("finanzas:pagos_list") + f"?periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}",
                {
                    "guardar_pago": "1",
                    "organizacion": self.org.pk,
                    "persona": self.persona_no_admin.pk,
                    "plan": "",
                    "documento_tributario": "",
                    "fecha_pago": "2026-02-27",
                    "metodo_pago": Payment.Metodo.EFECTIVO,
                    "numero_comprobante": "",
                    "monto_referencia": "15000",
                    "clases_asignadas": "2",
                    "observaciones": "",
                },
            )

        self.assertEqual(response.status_code, 302)
        pago = Payment.objects.get(monto_referencia=15000)
        self.assertTrue(
            AuditLog.objects.filter(
                dominio="finanzas",
                accion=AuditLog.ACCION_CREAR,
                modelo="finanzas.Payment",
                objeto_id=str(pago.pk),
            ).exists()
        )

    def test_pagos_list_exito_redirige_sin_open_registrar_pago(self):
        self.client.force_login(self.user_admin)
        url = (
            reverse("finanzas:pagos_list")
            + f"?periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}"
            "&open=registrar_pago"
        )

        response = self.client.post(
            url,
            {
                "guardar_pago": "1",
                "organizacion": self.org.pk,
                "persona": self.persona_no_admin.pk,
                "plan": "",
                "documento_tributario": "",
                "fecha_pago": "2026-02-27",
                "metodo_pago": Payment.Metodo.EFECTIVO,
                "numero_comprobante": "",
                "monto_referencia": "15000",
                "clases_asignadas": "2",
                "observaciones": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertNotIn("open=registrar_pago", response.url)
        recarga = self.client.get(response.url)
        self.assertFalse(recarga.context["open_registrar_pago"])

    def test_pagos_list_error_validacion_mantiene_modal_abierto(self):
        self.client.force_login(self.user_admin)
        response = self.client.post(
            reverse("finanzas:pagos_list")
            + f"?periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}"
            "&open=registrar_pago",
            {
                "guardar_pago": "1",
                "organizacion": self.org.pk,
                "persona": "",
                "fecha_pago": "2026-02-27",
                "metodo_pago": Payment.Metodo.EFECTIVO,
                "monto_referencia": "",
                "clases_asignadas": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["open_registrar_pago"])
        self.assertTrue(response.context["form"].errors)
        self.assertContains(response, "registrarPagoModal")

    def test_pagos_list_exito_conserva_filtros_al_quitar_open(self):
        self.client.force_login(self.user_admin)
        url = (
            reverse("finanzas:pagos_list")
            + f"?periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}"
            "&q=No&metodo=efectivo&open=registrar_pago"
        )

        response = self.client.post(
            url,
            {
                "guardar_pago": "1",
                "organizacion": self.org.pk,
                "persona": self.persona_no_admin.pk,
                "plan": "",
                "documento_tributario": "",
                "fecha_pago": "2026-02-27",
                "metodo_pago": Payment.Metodo.EFECTIVO,
                "numero_comprobante": "",
                "monto_referencia": "18000",
                "clases_asignadas": "3",
                "observaciones": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn("periodo_mes=2", response.url)
        self.assertIn("periodo_anio=2026", response.url)
        self.assertIn(f"organizacion={self.org.pk}", response.url)
        self.assertIn("q=No", response.url)
        self.assertIn("metodo=efectivo", response.url)
        self.assertNotIn("open=", response.url)

    def test_pago_edit_get_redirige_a_listado_con_modal(self):
        pago = Payment.objects.create(
            persona=self.persona_no_admin,
            organizacion=self.org,
            fecha_pago="2026-02-27",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=1,
        )
        self.client.force_login(self.user_admin)
        query = f"periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}&q=ana&metodo=transferencia"
        response = self.client.get(f"{reverse('finanzas:pago_edit', kwargs={'pk': pago.pk})}?{query}")
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"],
            f"{reverse('finanzas:pagos_list')}?{query}&editar_pago={pago.pk}",
        )

    def test_pagos_list_abre_modal_edicion_cuando_recibe_editar_pago(self):
        pago = Payment.objects.create(
            persona=self.persona_no_admin,
            organizacion=self.org,
            fecha_pago="2026-02-27",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=1,
        )
        self.client.force_login(self.user_admin)
        response = self.client.get(
            reverse("finanzas:pagos_list"),
            {
                "periodo_mes": 2,
                "periodo_anio": 2026,
                "organizacion": self.org.pk,
                "editar_pago": pago.pk,
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["edit_pago"], pago)
        self.assertIsNotNone(response.context["edit_form"])
        self.assertContains(response, 'id="editarPagoModal"', html=False)
        self.assertContains(response, f'action="{reverse("finanzas:pago_edit", kwargs={"pk": pago.pk})}?periodo_mes=2&amp;periodo_anio=2026&amp;organizacion={self.org.pk}"', html=False)
        self.assertContains(response, f'href="{reverse("finanzas:pagos_list")}?periodo_mes=2&amp;periodo_anio=2026&amp;organizacion={self.org.pk}"', html=False)

    def test_pago_edit_post_valido_redirige_sin_editar_pago(self):
        pago = Payment.objects.create(
            persona=self.persona_no_admin,
            organizacion=self.org,
            fecha_pago="2026-02-27",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=1,
        )
        self.client.force_login(self.user_admin)

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                (
                    f"{reverse('finanzas:pago_edit', kwargs={'pk': pago.pk})}"
                    f"?periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}&editar_pago={pago.pk}"
                ),
                {
                    "edit_pago-organizacion": str(self.org.pk),
                    "edit_pago-persona": str(self.persona_no_admin.pk),
                    "edit_pago-plan": "",
                    "edit_pago-documento_tributario": "",
                    "edit_pago-fecha_pago": "2026-02-28",
                    "edit_pago-metodo_pago": Payment.Metodo.EFECTIVO,
                    "edit_pago-numero_comprobante": "",
                    "edit_pago-monto_referencia": "12000",
                    "edit_pago-clases_asignadas": "2",
                    "edit_pago-observaciones": "Pago actualizado",
                },
            )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"],
            f"{reverse('finanzas:pagos_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}",
        )
        log = AuditLog.objects.filter(
            dominio="finanzas",
            accion=AuditLog.ACCION_EDITAR,
            modelo="finanzas.Payment",
            objeto_id=str(pago.pk),
        ).latest("fecha")
        self.assertIn("monto_referencia", log.metadata["cambios"])

    def test_pagos_list_crea_persona_rapida_como_estudiante_en_organizacion_filtrada(self):
        self.client.force_login(self.user_admin)

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse("finanzas:pagos_list") + f"?periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}",
                {
                    "nombres": "Lucia",
                    "apellidos": "Perez",
                    "telefono": "999999",
                    "agregar_persona": "1",
                },
            )

        self.assertEqual(response.status_code, 302)
        persona = Persona.objects.get(nombres="Lucia", apellidos="Perez")
        self.assertTrue(
            PersonaRol.objects.filter(
                persona=persona,
                rol__codigo="ESTUDIANTE",
                organizacion=self.org,
                activo=True,
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                dominio="personas",
                accion=AuditLog.ACCION_CREAR,
                objeto_id=str(persona.pk),
            ).exists()
        )

    def test_pagos_list_nueva_persona_sin_organizacion_se_deniega(self):
        self.client.force_login(self.user_admin)

        response = self.client.post(
            reverse("finanzas:pagos_list") + "?periodo_mes=2&periodo_anio=2026",
            {
                "nombres": "Mario",
                "apellidos": "Lopez",
                "telefono": "888888",
                "agregar_persona": "1",
            },
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(Persona.objects.filter(nombres="Mario", apellidos="Lopez").exists())

    def test_transaccion_detail_muestra_iframe_pdf(self):
        categoria = Category.objects.create(nombre="Arriendo", tipo="egreso", activa=True)
        archivo = SimpleUploadedFile(
            "comprobante.pdf",
            b"%PDF-1.4\n1 0 obj\n<<>>\nendobj\ntrailer\n<<>>\n%%EOF",
            content_type="application/pdf",
        )
        transaccion = Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-02-27",
            tipo=Transaction.Tipo.EGRESO,
            monto=15000,
            descripcion="Pago de arriendo",
            archivo=archivo,
        )

        self.client.force_login(self.user_admin)
        query = f"periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}"
        response = self.client.get(f"{reverse('finanzas:transaccion_detail', kwargs={'pk': transaccion.pk})}?{query}")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["archivo_es_pdf"])
        self.assertContains(response, "<iframe", html=False)
        self.assertContains(response, reverse("finanzas:transaccion_archivo", kwargs={"pk": transaccion.pk}))

    def test_transaccion_detail_muestra_imagen_inline(self):
        categoria = Category.objects.create(nombre="Movilidad", tipo="egreso", activa=True)
        archivo = SimpleUploadedFile(
            "comprobante.jpg",
            b"\xff\xd8\xff\xe0\x00\x10JFIF\x00\x01\x01\x00\x00\x01\x00\x01\x00\x00\xff\xd9",
            content_type="image/jpeg",
        )
        transaccion = Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-02-27",
            tipo=Transaction.Tipo.EGRESO,
            monto=18000,
            descripcion="Taxi",
            archivo=archivo,
        )

        self.client.force_login(self.user_admin)
        response = self.client.get(
            reverse("finanzas:transaccion_detail", kwargs={"pk": transaccion.pk}),
            {"organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["archivo_es_pdf"])
        self.assertTrue(response.context["archivo_es_imagen"])
        self.assertContains(response, "<img", html=False)
        self.assertContains(response, reverse("finanzas:transaccion_archivo", kwargs={"pk": transaccion.pk}))

    def test_transaccion_archivo_permite_iframe_sameorigin(self):
        categoria = Category.objects.create(nombre="Honorarios", tipo="egreso", activa=True)
        archivo = SimpleUploadedFile(
            "respaldo.pdf",
            b"%PDF-1.4\n1 0 obj\n<<>>\nendobj\ntrailer\n<<>>\n%%EOF",
            content_type="application/pdf",
        )
        transaccion = Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-02-27",
            tipo=Transaction.Tipo.EGRESO,
            monto=9900,
            descripcion="Honorarios",
            archivo=archivo,
        )

        self.client.force_login(self.user_admin)
        response = self.client.get(
            reverse("finanzas:transaccion_archivo", kwargs={"pk": transaccion.pk}),
            {"organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["X-Frame-Options"], "SAMEORIGIN")
        self.assertIn("inline;", response["Content-Disposition"])

    def test_archivos_protegidos_no_exponen_ruta_media_y_aislan_organizacion(self):
        otra_org = Organizacion.objects.create(
            nombre="Org archivos ajenos",
            razon_social="Org archivos ajenos SpA",
            rut="22.222.223-0",
        )
        categoria = Category.objects.create(nombre="Respaldo privado", tipo="egreso", activa=True)
        with TemporaryDirectory() as media_root, self.settings(MEDIA_ROOT=media_root):
            documento = DocumentoTributario.objects.create(
                organizacion=otra_org,
                tipo_documento=DocumentoTributario.TipoDocumento.BOLETA_HONORARIOS,
                folio="PRIV-1",
                fecha_emision="2026-02-27",
                nombre_emisor="Emisor privado",
                monto_total=10000,
                archivo_pdf=SimpleUploadedFile(
                    "documento-privado.pdf",
                    b"%PDF-1.4\n%%EOF",
                    content_type="application/pdf",
                ),
            )
            transaccion = Transaction.objects.create(
                organizacion=otra_org,
                categoria=categoria,
                fecha="2026-02-27",
                tipo=Transaction.Tipo.EGRESO,
                monto=10000,
                descripcion="Transacción privada",
                archivo=SimpleUploadedFile(
                    "transaccion-privada.pdf",
                    b"%PDF-1.4\n%%EOF",
                    content_type="application/pdf",
                ),
            )
            url_documento = reverse(
                "finanzas:documento_tributario_archivo",
                kwargs={"pk": documento.pk, "tipo_archivo": "pdf"},
            )
            url_transaccion = reverse(
                "finanzas:transaccion_archivo",
                kwargs={"pk": transaccion.pk},
            )

            self.assertEqual(self.client.get(url_documento).status_code, 302)
            self.client.force_login(self.user_finanzas)
            for url_ajena, url_inexistente in (
                (
                    url_documento,
                    reverse(
                        "finanzas:documento_tributario_archivo",
                        kwargs={"pk": 999991, "tipo_archivo": "pdf"},
                    ),
                ),
                (
                    url_transaccion,
                    reverse("finanzas:transaccion_archivo", kwargs={"pk": 999992}),
                ),
            ):
                with self.subTest(url=url_ajena):
                    respuesta_ajena = self.client.get(url_ajena, {"organizacion": self.org.pk})
                    respuesta_inexistente = self.client.get(url_inexistente, {"organizacion": self.org.pk})
                    self.assertEqual(respuesta_ajena.status_code, 404)
                    self.assertEqual(respuesta_inexistente.status_code, 404)

            self.client.force_login(self.user_admin)
            PersonaRol.objects.create(
                persona=self.persona_admin,
                rol=self.rol_admin,
                organizacion=otra_org,
                activo=True,
            )
            detalle = self.client.get(
                reverse("finanzas:documento_tributario_detail", kwargs={"pk": documento.pk}),
                {"organizacion": otra_org.pk},
            )
            self.assertEqual(detalle.status_code, 200)
            self.assertContains(detalle, url_documento)
            self.assertNotContains(detalle, documento.archivo_pdf.url)

    @patch("finanzas.views.DocumentoTributarioForm.save", side_effect=IntegrityError("conflicto"))
    def test_documento_tributario_edit_muestra_error_legible_si_falla_unicidad(self, _mock_save):
        documento = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.BOLETA_HONORARIOS,
            folio="B-1",
            fecha_emision="2026-02-27",
            rut_emisor="11.111.111-1",
            nombre_emisor="Emisor Original",
            monto_total=10000,
        )
        self.client.force_login(self.user_admin)
        response = self.client.post(
            self._url_con_organizacion(
                reverse("finanzas:documento_tributario_edit", kwargs={"pk": documento.pk})
            ),
            {
                "organizacion": self.org.pk,
                "tipo_documento": DocumentoTributario.TipoDocumento.BOLETA_HONORARIOS,
                "fuente": DocumentoTributario.Fuente.MANUAL,
                "folio": "B-1",
                "fecha_emision": "2026-02-27",
                "nombre_emisor": "Emisor Original",
                "rut_emisor": "11.111.111-1",
                "nombre_receptor": "",
                "rut_receptor": "",
                "monto_neto": "10000",
                "monto_exento": "0",
                "iva_tasa": "0",
                "monto_iva": "0",
                "retencion_tasa": "0",
                "retencion_monto": "0",
                "monto_total": "10000",
                "documento_relacionado": "",
                "enlace_sii": "",
                "metadata_extra": "{}",
                "observaciones": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No se pudo guardar el documento por un conflicto de unicidad")

    def test_reporte_categorias_muestra_grafico_torta(self):
        categoria = Category.objects.create(nombre="Arriendo sala", tipo="egreso", activa=True)
        Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-02-27",
            tipo=Transaction.Tipo.EGRESO,
            monto=25000,
            descripcion="Arriendo febrero",
        )

        self.client.force_login(self.user_admin)
        response = self.client.get(
            reverse("finanzas:reporte_categorias"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'canvas id="categoriasChart"', html=False)
        self.assertContains(response, "Chart(", html=False)
        self.assertContains(response, "Arriendo sala")

    def test_export_pagos_csv_mantiene_headers_y_filas_filtradas(self):
        Payment.objects.create(
            persona=self.persona_no_admin,
            organizacion=self.org,
            fecha_pago="2026-02-25",
            metodo_pago=Payment.Metodo.TRANSFERENCIA,
            numero_comprobante="FEB-1",
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=2,
        )
        Payment.objects.create(
            persona=self.persona_no_admin,
            organizacion=self.org,
            fecha_pago="2026-03-01",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=5000,
            clases_asignadas=1,
        )

        self.client.force_login(self.user_admin)
        response = self.client.get(
            reverse("finanzas:export_pagos_csv"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        rows = list(csv.reader(response.content.decode().splitlines()))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Disposition"], 'attachment; filename="pagos_finanzas.csv"')
        self.assertEqual(rows[0], ["Fecha", "Organizacion", "Persona", "Metodo", "Neto", "IVA", "Total", "Clases"])
        self.assertEqual(len(rows), 2)
        self.assertEqual(
            rows[1],
            [
                "2026-02-25",
                "Org Finanzas",
                "No Admin",
                "Transferencia",
                "10000.00",
                "0.00",
                "10000.00",
                "2",
            ],
        )

    def test_export_transacciones_csv_mantiene_headers_y_filas_filtradas(self):
        categoria_ingreso = Category.objects.create(nombre="Ventas clases", tipo=Category.Tipo.INGRESO, activa=True)
        categoria_egreso = Category.objects.create(nombre="Arriendo sala", tipo=Category.Tipo.EGRESO, activa=True)
        Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria_ingreso,
            fecha="2026-02-25",
            tipo=Transaction.Tipo.INGRESO,
            monto=25000,
            descripcion="Ingreso febrero",
        )
        Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria_egreso,
            fecha="2026-03-01",
            tipo=Transaction.Tipo.EGRESO,
            monto=12000,
            descripcion="Egreso marzo",
        )

        self.client.force_login(self.user_admin)
        response = self.client.get(
            reverse("finanzas:export_transacciones_csv"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        rows = list(csv.reader(response.content.decode().splitlines()))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Disposition"], 'attachment; filename="transacciones_finanzas.csv"')
        self.assertEqual(rows[0], ["Fecha", "Organizacion", "Tipo", "Categoria", "Monto", "Descripcion"])
        self.assertEqual(len(rows), 2)
        self.assertEqual(
            rows[1],
            [
                "2026-02-25",
                "Org Finanzas",
                "Ingreso",
                "Ventas clases",
                "25000.00",
                "Ingreso febrero",
            ],
        )

    def test_payment_plan_primer_plan_queda_por_defecto_y_se_puede_reasignar(self):
        plan_1 = PaymentPlan.objects.create(
            organizacion=self.org,
            nombre="Plan Inicial",
            num_clases=4,
            precio=20000,
            activo=True,
        )
        plan_1.refresh_from_db()
        self.assertTrue(plan_1.es_por_defecto)

        plan_2 = PaymentPlan.objects.create(
            organizacion=self.org,
            nombre="Plan Nuevo",
            num_clases=8,
            precio=35000,
            activo=True,
        )
        plan_2.refresh_from_db()
        self.assertFalse(plan_2.es_por_defecto)

        plan_2.es_por_defecto = True
        plan_2.save()
        plan_1.refresh_from_db()
        plan_2.refresh_from_db()

        self.assertFalse(plan_1.es_por_defecto)
        self.assertTrue(plan_2.es_por_defecto)

        plan_2.delete()
        plan_1.refresh_from_db()
        self.assertTrue(plan_1.es_por_defecto)

    def test_payment_form_precarga_plan_por_defecto_de_la_organizacion(self):
        PaymentPlan.objects.create(
            organizacion=self.org,
            nombre="Plan Base",
            num_clases=4,
            precio=20000,
            activo=True,
        )
        plan_destacado = PaymentPlan.objects.create(
            organizacion=self.org,
            nombre="Plan Destacado",
            num_clases=8,
            precio=30000,
            activo=True,
            es_por_defecto=True,
        )

        form = PaymentForm(initial={"organizacion": self.org.pk})

        self.assertEqual(str(form["plan"].value()), str(plan_destacado.pk))

    def test_payment_form_precarga_aplica_iva_segun_configuracion_de_organizacion(self):
        form_afecta = PaymentForm(initial={"organizacion": self.org.pk})
        self.assertTrue(form_afecta.initial["aplica_iva"])

        org_exenta = Organizacion.objects.create(
            nombre="Org Exenta",
            razon_social="Org Exenta SPA",
            rut="66.666.666-6",
            es_exenta_iva=True,
        )
        form_exenta = PaymentForm(initial={"organizacion": org_exenta.pk})
        self.assertFalse(form_exenta.initial["aplica_iva"])

    def test_plan_edit_renderiza_listado_con_edicion_inline(self):
        plan = PaymentPlan.objects.create(
            organizacion=self.org,
            nombre="Plan Editable",
            num_clases=4,
            precio=20000,
            activo=True,
        )
        self.client.force_login(self.user_admin)

        response = self.client.get(
            reverse("finanzas:plan_edit", kwargs={"pk": plan.pk}),
            {"periodo_mes": 3, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "finanzas/planes_list.html")
        self.assertEqual(response.context["editing_plan_id"], plan.pk)
        self.assertContains(response, "Guardar cambios")
        self.assertContains(response, 'name="es_por_defecto"', html=False)
        self.assertNotContains(response, "Editar plan")

    def test_plan_edit_inline_actualiza_plan_por_defecto(self):
        plan_base = PaymentPlan.objects.create(
            organizacion=self.org,
            nombre="Plan Base",
            num_clases=4,
            precio=20000,
            activo=True,
        )
        plan_otro = PaymentPlan.objects.create(
            organizacion=self.org,
            nombre="Plan Otro",
            num_clases=8,
            precio=30000,
            activo=True,
        )
        self.client.force_login(self.user_admin)

        response = self.client.post(
            f"{reverse('finanzas:plan_edit', kwargs={'pk': plan_otro.pk})}?periodo_mes=3&periodo_anio=2026&organizacion={self.org.pk}",
            {
                "organizacion": self.org.pk,
                "nombre": "Plan Otro",
                "num_clases": 8,
                "precio": 30000,
                "precio_incluye_iva": "",
                "es_por_defecto": "on",
                "fecha_inicio": "",
                "fecha_fin": "",
                "descripcion": "",
                "activo": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        plan_base.refresh_from_db()
        plan_otro.refresh_from_db()
        self.assertFalse(plan_base.es_por_defecto)
        self.assertTrue(plan_otro.es_por_defecto)

    def test_transaction_form_deriva_tipo_desde_categoria(self):
        categoria = Category.objects.create(nombre="Venta", tipo="ingreso", activa=True)
        form = TransactionForm(
            data={
                "organizacion": self.org.pk,
                "categoria": categoria.pk,
                "fecha": "2026-02-27",
                "monto": "25000",
                "descripcion": "Ingreso evento",
                "documentos_tributarios": [],
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        transaccion = form.save(commit=False)
        self.assertEqual(transaccion.tipo, Transaction.Tipo.INGRESO)

    def test_transaccion_edit_get_precarga_fecha_en_formato_html(self):
        categoria = Category.objects.create(nombre="Ingreso edición", tipo=Category.Tipo.INGRESO)
        transaccion = Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-02-27",
            tipo=Transaction.Tipo.INGRESO,
            monto=25000,
            descripcion="Movimiento a editar",
        )
        self.client.force_login(self.user_admin)

        response = self.client.get(
            reverse("finanzas:transaccion_edit", kwargs={"pk": transaccion.pk}),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<input type="date" name="fecha"', html=False)
        self.assertContains(response, 'value="2026-02-27"', html=False)

    def test_transaccion_edit_post_actualiza_sin_regresion(self):
        categoria = Category.objects.create(nombre="Egreso edición", tipo=Category.Tipo.EGRESO)
        transaccion = Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-02-27",
            tipo=Transaction.Tipo.EGRESO,
            monto=25000,
            descripcion="Movimiento original",
        )
        self.client.force_login(self.user_admin)

        response = self.client.post(
            reverse("finanzas:transaccion_edit", kwargs={"pk": transaccion.pk})
            + f"?periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}",
            {
                "organizacion": self.org.pk,
                "categoria": categoria.pk,
                "fecha": "2026-02-28",
                "monto": "30000",
                "descripcion": "Movimiento actualizado",
                "documentos_tributarios": [],
            },
        )

        self.assertEqual(response.status_code, 302)
        transaccion.refresh_from_db()
        self.assertEqual(transaccion.fecha, date(2026, 2, 28))
        self.assertEqual(transaccion.monto, 30000)
        self.assertEqual(transaccion.descripcion, "Movimiento actualizado")

    def test_transaccion_edit_error_conserva_fecha_enviada(self):
        categoria = Category.objects.create(nombre="Ingreso inválido", tipo=Category.Tipo.INGRESO)
        transaccion = Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-02-27",
            tipo=Transaction.Tipo.INGRESO,
            monto=25000,
            descripcion="Movimiento original",
        )
        self.client.force_login(self.user_admin)

        response = self.client.post(
            reverse("finanzas:transaccion_edit", kwargs={"pk": transaccion.pk})
            + f"?periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}",
            {
                "organizacion": self.org.pk,
                "categoria": categoria.pk,
                "fecha": "2026-02-26",
                "monto": "",
                "descripcion": "Movimiento inválido",
                "documentos_tributarios": [],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["form"].errors)
        self.assertContains(response, '<input type="date" name="fecha"', html=False)
        self.assertContains(response, 'value="2026-02-26"', html=False)

    def test_transacciones_list_precarga_organizacion_del_filtro_en_formulario(self):
        self.client.force_login(self.user_admin)

        response = self.client.get(
            reverse("finanzas:transacciones_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["form"].initial["organizacion"], self.org.pk)
        self.assertContains(response, f'<option value="{self.org.pk}" selected>', html=False)

    def test_transaction_form_muestra_extracto_en_opciones_de_documentos(self):
        documento = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.BOLETA_HONORARIOS,
            folio="BH-33",
            fecha_emision="2026-02-27",
            nombre_emisor="Artista",
            nombre_receptor="Org Finanzas",
            monto_total=50000,
            observaciones="Pago honorarios presentacion La Tarea mas Dificil en festival de febrero",
        )

        form = TransactionForm()
        etiqueta = form.fields["documentos_tributarios"].label_from_instance(documento)

        self.assertIn("Boleta de honorarios #BH-33", etiqueta)
        self.assertIn("Pago honorarios presentacion La Tarea mas Dificil", etiqueta)

    def test_documento_tributario_detail_muestra_iframe_pdf_y_asociaciones(self):
        categoria = Category.objects.create(nombre="Honorarios evento", tipo="egreso", activa=True)
        archivo_pdf = SimpleUploadedFile(
            "honorarios.pdf",
            b"%PDF-1.4\n1 0 obj\n<<>>\nendobj\ntrailer\n<<>>\n%%EOF",
            content_type="application/pdf",
        )
        documento = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.BOLETA_HONORARIOS,
            folio="H-100",
            fecha_emision="2026-02-27",
            nombre_emisor="Artista Uno",
            nombre_receptor="Org Finanzas",
            monto_total=125000,
            archivo_pdf=archivo_pdf,
        )
        transaccion = Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-02-27",
            tipo=Transaction.Tipo.EGRESO,
            monto=125000,
            descripcion="Pago honorarios artista",
        )
        transaccion.documentos_tributarios.add(documento)

        self.client.force_login(self.user_admin)
        query = f"periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}"
        response = self.client.get(
            f"{reverse('finanzas:documento_tributario_detail', kwargs={'pk': documento.pk})}?{query}"
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["archivo_es_pdf"])
        self.assertContains(response, "<iframe", html=False)
        self.assertContains(
            response,
            reverse("finanzas:transaccion_detail", kwargs={"pk": transaccion.pk}),
        )

    def test_documentos_tributarios_list_muestra_resumen_del_listado(self):
        categoria = Category.objects.create(nombre="Honorarios evento", tipo="egreso", activa=True)
        documento_1 = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_AFECTA,
            folio="D-1",
            fecha_emision="2026-02-10",
            nombre_emisor=self.org.razon_social,
            rut_emisor=self.org.rut,
            nombre_receptor="Receptor Uno",
            rut_receptor="11.111.111-1",
            monto_neto=10000,
            monto_exento=88100,
            monto_iva=1900,
            monto_total=100000,
        )
        documento_2 = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.BOLETA_HONORARIOS,
            folio="D-2",
            fecha_emision="2026-02-15",
            nombre_emisor="Emisor Dos",
            rut_emisor="33.333.333-3",
            nombre_receptor=self.org.razon_social,
            rut_receptor=self.org.rut,
            monto_neto=50000,
            retencion_monto=15250,
            monto_total=50000,
        )
        Payment.objects.create(
            persona=self.persona_no_admin,
            organizacion=self.org,
            documento_tributario=documento_1,
            fecha_pago="2026-02-20",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=1,
        )
        transaccion = Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-02-21",
            tipo=Transaction.Tipo.EGRESO,
            monto=50000,
            descripcion="Pago honorarios",
        )
        transaccion.documentos_tributarios.add(documento_2)

        self.client.force_login(self.user_admin)
        response = self.client.get(
            reverse("finanzas:documentos_tributarios_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_documentos"], 2)
        self.assertEqual(response.context["monto_total_ingresos_documentales"], 100000)
        self.assertEqual(response.context["monto_total_egresos_documentales"], 50000)
        self.assertEqual(response.context["monto_total_iva"], 1900)
        self.assertEqual(response.context["monto_total_retencion"], 15250)
        self.assertEqual(response.context["total_pagos_asociados"], 1)
        self.assertEqual(response.context["total_transacciones_asociadas"], 1)
        self.assertNotContains(response, "Total documentos")
        self.assertContains(response, "Ingresos")
        self.assertContains(response, "Egresos")
        self.assertContains(response, "IVA")
        self.assertContains(response, "Retencion")
        self.assertContains(response, "<th>Neto</th>", html=False)
        self.assertContains(response, "<th>Exento</th>", html=False)
        self.assertContains(response, "<th>IVA</th>", html=False)
        self.assertContains(response, "<th>Retencion</th>", html=False)
        self.assertNotContains(response, "<th>Organizacion</th>", html=False)
        self.assertContains(response, "$ 100.000")
        self.assertContains(response, "$ 10.000")
        self.assertContains(response, "$ 88.100")
        self.assertContains(response, "$ 1.900")
        self.assertContains(response, "$ 15.250")
        self.assertContains(response, "$ 50.000")

    def test_documentos_tributarios_list_crea_documento_y_auditlog(self):
        self.client.force_login(self.user_admin)

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse("finanzas:documentos_tributarios_list")
                + f"?periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}",
                {
                    "organizacion": self.org.pk,
                    "tipo_documento": DocumentoTributario.TipoDocumento.FACTURA_AFECTA,
                    "fuente": DocumentoTributario.Fuente.MANUAL,
                    "folio": "AUD-1",
                    "fecha_emision": "2026-02-11",
                    "nombre_emisor": "Emisor Audit",
                    "rut_emisor": "11.111.111-1",
                    "nombre_receptor": "Receptor Audit",
                    "rut_receptor": "22.222.222-2",
                    "monto_neto": "10000",
                    "monto_exento": "0",
                    "iva_tasa": "19.00",
                    "monto_iva": "1900",
                    "retencion_tasa": "0",
                    "retencion_monto": "0",
                    "monto_total": "11900",
                    "documento_relacionado": "",
                    "persona_relacionada": "",
                    "organizacion_relacionada": "",
                    "enlace_sii": "",
                    "metadata_extra": "{}",
                    "observaciones": "",
                },
            )

        self.assertEqual(response.status_code, 302)
        documento = DocumentoTributario.objects.get(folio="AUD-1")
        self.assertTrue(
            AuditLog.objects.filter(
                dominio="finanzas",
                accion=AuditLog.ACCION_CREAR,
                modelo="finanzas.DocumentoTributario",
                objeto_id=str(documento.pk),
            ).exists()
        )

    def test_transacciones_list_muestra_resumen_del_listado(self):
        categoria_ingreso = Category.objects.create(nombre="Venta", tipo="ingreso", activa=True)
        categoria_egreso = Category.objects.create(nombre="Honorarios", tipo="egreso", activa=True)
        Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria_ingreso,
            fecha="2026-02-05",
            tipo=Transaction.Tipo.INGRESO,
            monto=120000,
            descripcion="Ingreso evento",
        )
        Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria_egreso,
            fecha="2026-02-06",
            tipo=Transaction.Tipo.EGRESO,
            monto=30000,
            descripcion="Pago artista",
        )
        Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria_egreso,
            fecha="2026-03-06",
            tipo=Transaction.Tipo.EGRESO,
            monto=99999,
            descripcion="Fuera de periodo",
        )

        self.client.force_login(self.user_admin)
        response = self.client.get(
            reverse("finanzas:transacciones_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_transacciones"], 2)
        self.assertEqual(response.context["total_ingresos"], 120000)
        self.assertEqual(response.context["total_egresos"], 30000)
        self.assertEqual(response.context["balance_transacciones"], 90000)
        self.assertContains(response, "Total transacciones")
        self.assertContains(response, "Balance")

    def test_transacciones_list_crea_transaccion_y_auditlog(self):
        categoria = Category.objects.create(nombre="Venta audit", tipo="ingreso", activa=True)
        self.client.force_login(self.user_admin)

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse("finanzas:transacciones_list")
                + f"?periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}",
                {
                    "organizacion": self.org.pk,
                    "categoria": categoria.pk,
                    "fecha": "2026-02-07",
                    "tipo": Transaction.Tipo.INGRESO,
                    "monto": "33000",
                    "descripcion": "Ingreso auditado",
                    "documentos_tributarios": [],
                },
            )

        self.assertEqual(response.status_code, 302)
        transaccion = Transaction.objects.get(descripcion="Ingreso auditado")
        self.assertTrue(
            AuditLog.objects.filter(
                dominio="finanzas",
                accion=AuditLog.ACCION_CREAR,
                modelo="finanzas.Transaction",
                objeto_id=str(transaccion.pk),
            ).exists()
        )

    def test_documento_tributario_importar_parsea_y_muestra_revision_sin_guardar(self):
        xml = SimpleUploadedFile(
            "boleta.xml",
            b"""
            <EnvioDTE>
              <SetDTE>
                <DTE>
                  <Documento>
                    <Encabezado>
                      <IdDoc>
                        <TipoDTE>39</TipoDTE>
                        <Folio>101</Folio>
                        <FchEmis>2026-02-27</FchEmis>
                      </IdDoc>
                      <Emisor>
                        <RUTEmisor>11.111.111-1</RUTEmisor>
                        <RznSoc>Org Finanzas</RznSoc>
                      </Emisor>
                      <Receptor>
                        <RUTRecep>22.222.222-2</RUTRecep>
                        <RznSocRecep>Ana Diaz</RznSocRecep>
                      </Receptor>
                      <Totales>
                        <MntNeto>10000</MntNeto>
                        <IVA>1900</IVA>
                        <MntTotal>11900</MntTotal>
                        <TasaIVA>19</TasaIVA>
                      </Totales>
                    </Encabezado>
                    <Detalle>
                      <NroLinDet>1</NroLinDet>
                      <NmbItem>Plan mensual</NmbItem>
                      <QtyItem>1</QtyItem>
                      <PrcItem>10000</PrcItem>
                      <MontoItem>10000</MontoItem>
                    </Detalle>
                  </Documento>
                </DTE>
              </SetDTE>
            </EnvioDTE>
            """,
            content_type="application/xml",
        )
        self.client.force_login(self.user_admin)

        upload_response = self.client.get(
            reverse("finanzas:documento_tributario_importar"), {"organizacion": self.org.pk}
        )
        self.assertEqual(upload_response.status_code, 200)
        self.assertContains(upload_response, 'name="archivo"', html=False)
        self.assertNotContains(upload_response, 'name="archivo_xml"', html=False)
        self.assertNotContains(upload_response, 'name="archivo_pdf"', html=False)

        response = self.client.post(
            self._url_con_organizacion(reverse("finanzas:documento_tributario_importar")),
            {"accion": "parsear", "archivo": xml},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Revision del documento tributario")
        self.assertContains(response, "101")
        self.assertEqual(response.context["documento_form"].initial["observaciones"], "Plan mensual")
        self.assertEqual(DocumentoTributario.objects.count(), 0)

    def test_documento_tributario_importar_sugiere_persona_relacionada_por_rut_contraparte(self):
        persona_artista = Persona.objects.create(
            nombres="Barbara",
            apellidos="Allendes",
            rut="18.445.523-4",
            email="barbara@example.com",
        )
        pdf_path = Path(__file__).resolve().parent.parent / "public" / "202603_LaTarea+Dificil.12Febrero2026_BarbaraAllendes.pdf"
        pdf = SimpleUploadedFile(
            pdf_path.name,
            pdf_path.read_bytes(),
            content_type="application/pdf",
        )
        self.client.force_login(self.user_admin)

        response = self.client.post(
            f"{reverse('finanzas:documento_tributario_importar')}?periodo_mes=3&periodo_anio=2026&organizacion={self.org.pk}",
            {"accion": "parsear", "archivo": pdf},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["documento_form"].initial["persona_relacionada"], persona_artista.pk)
        self.assertEqual(response.context["documento_form"].initial["organizacion_relacionada"], "")

    def test_documento_tributario_importar_sugiere_organizacion_relacionada_si_no_hay_persona(self):
        contraparte = Organizacion.objects.create(
            nombre="Pereira EIRL",
            razon_social="PEREIRA E.I.R.L.",
            rut="77.752.651-0",
        )
        xml = SimpleUploadedFile(
            "factura.xml",
            b"""
            <EnvioDTE>
              <SetDTE>
                <DTE>
                  <Documento>
                    <Encabezado>
                      <IdDoc>
                        <TipoDTE>34</TipoDTE>
                        <Folio>502</Folio>
                        <FchEmis>2026-03-10</FchEmis>
                      </IdDoc>
                      <Emisor>
                        <RUTEmisor>22.222.222-2</RUTEmisor>
                        <RznSoc>Org Finanzas SPA</RznSoc>
                      </Emisor>
                      <Receptor>
                        <RUTRecep>77.752.651-0</RUTRecep>
                        <RznSocRecep>PEREIRA E.I.R.L.</RznSocRecep>
                      </Receptor>
                      <Totales>
                        <MntExe>500000</MntExe>
                        <MntTotal>500000</MntTotal>
                      </Totales>
                    </Encabezado>
                    <Detalle>
                      <NroLinDet>1</NroLinDet>
                      <NmbItem>Funcion</NmbItem>
                      <QtyItem>1</QtyItem>
                      <PrcItem>500000</PrcItem>
                      <MontoItem>500000</MontoItem>
                    </Detalle>
                  </Documento>
                </DTE>
              </SetDTE>
            </EnvioDTE>
            """,
            content_type="application/xml",
        )
        self.client.force_login(self.user_admin)

        response = self.client.post(
            f"{reverse('finanzas:documento_tributario_importar')}?periodo_mes=3&periodo_anio=2026&organizacion={self.org.pk}",
            {"accion": "parsear", "archivo": xml},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["documento_form"].initial["persona_relacionada"], "")
        self.assertEqual(response.context["documento_form"].initial["organizacion_relacionada"], contraparte.pk)
        self.assertEqual(response.context["review_payload"]["suggestions"]["organizacion_sugerida_id"], contraparte.pk)

    @patch("finanzas.documentos.parsers.PdfFallbackParser._extract_text_with_pdftotext")
    @patch("finanzas.documentos.parsers.PdfFallbackParser._extract_text_with_pypdf", return_value="")
    def test_documento_tributario_importar_precarga_formulario_desde_pdf(
        self,
        _extract_text_with_pypdf,
        extract_text_with_pdftotext,
    ):
        extract_text_with_pdftotext.return_value = """
        R.U.T.: 77.813.508-6
        ESPACIO CULTURAL Y DEPORTIVO ELEMENTOS SPA
        GIRO: REALIZACIÓN DE ACTIVIDADES
        FACTURA NO AFECTA O EXENTA ELECTRONICA
        Nº2
        Fecha Emision: 10 de Marzo del 2026
        SEÑOR(ES):
        PEREIRA E.I.R.L.
        R.U.T.: 77.752.651-0
        GIRO: SERVICIOS DE PRODUCCION DE OBRAS DE TEAT
        DIRECCION: AV. NVA A EINSTEIN 290 PLAZA AMERICA 808
        DESCRIPCION                CANTIDAD     PRECIO       TOTAL
        - Obra Circo Contemporaneo      1       500.000     500.000
        Función La Tarea más difícil - Febrero - 2026
        EXENTO $ 500.000
        TOTAL $ 500.000
        """
        pdf = SimpleUploadedFile(
            "factura.pdf",
            b"%PDF-1.4\n1 0 obj\n<<>>\nendobj\ntrailer\n<<>>\n%%EOF",
            content_type="application/pdf",
        )
        self.client.force_login(self.user_admin)

        response = self.client.post(
            self._url_con_organizacion(reverse("finanzas:documento_tributario_importar")),
            {"accion": "parsear", "archivo": pdf},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Revision del documento tributario")
        self.assertContains(response, 'name="folio"')
        self.assertContains(response, 'value="2"', html=False)
        self.assertContains(response, 'value="2026-03-10"', html=False)
        self.assertContains(response, 'value="ESPACIO CULTURAL Y DEPORTIVO ELEMENTOS SPA"', html=False)
        self.assertContains(response, 'value="77.813.508-6"', html=False)
        self.assertContains(response, 'value="PEREIRA E.I.R.L."', html=False)
        self.assertContains(response, 'value="77.752.651-0"', html=False)
        self.assertContains(response, 'value="500000"', html=False)
        self.assertEqual(
            response.context["documento_form"].initial["observaciones"],
            "Obra Circo Contemporaneo Función La Tarea más difícil - Febrero - 2026",
        )
        self.assertEqual(DocumentoTributario.objects.count(), 0)

        token = next(iter(self.client.session[SESSION_KEY].keys()))
        visor_url = reverse(
            "finanzas:documento_tributario_importacion_archivo",
            kwargs={"token": token, "tipo_archivo": "pdf"},
        )
        self.assertContains(response, visor_url)
        self.assertContains(response, "<iframe", html=False)

        visor_response = self.client.get(self._url_con_organizacion(visor_url))
        self.assertEqual(visor_response.status_code, 200)
        self.assertEqual(visor_response["X-Frame-Options"], "SAMEORIGIN")
        self.assertIn("inline;", visor_response["Content-Disposition"])

    @patch("finanzas.documentos.parsers.PdfFallbackParser._extract_text_with_pdftotext")
    @patch("finanzas.documentos.parsers.PdfFallbackParser._extract_text_with_pypdf", return_value="")
    def test_documento_tributario_importar_precarga_boleta_venta_pdf(
        self,
        _extract_text_with_pypdf,
        extract_text_with_pdftotext,
    ):
        extract_text_with_pdftotext.return_value = """
        ESPACIO CULTURAL Y DEPORTIVO
        ELEMENTOS SPA
        77.813.508-6
        Giro: REALIZACIN DE ACTIVIDADES
        DEPORTIVAS Y CULTURALES PARA LA
        COMUNIDAD
        NUEVA TRES 1020
        Rengo
        BOLETA ELECTRÓNICA NUMERO: 20.035
        REF. VENDEDOR: 17085005-K
        Fecha: 2026-04-15

        Dirección: Santiago

        Medio de pago: Transferencia
        Electrónica
        Taller de Fuerza/Calistenia - Plan
        2 Clases Semanales (Erika Huerta)
                                   $ 36.000

        El IVA incluido en esta boleta es
        de: $ 5.748
        """
        pdf = SimpleUploadedFile(
            "boleta39.pdf",
            b"%PDF-1.4\n1 0 obj\n<<>>\nendobj\ntrailer\n<<>>\n%%EOF",
            content_type="application/pdf",
        )
        self.client.force_login(self.user_admin)

        response = self.client.post(
            self._url_con_organizacion(reverse("finanzas:documento_tributario_importar")),
            {"accion": "parsear", "archivo": pdf},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'value="20035"', html=False)
        self.assertContains(response, 'value="2026-04-15"', html=False)
        self.assertContains(response, 'value="ESPACIO CULTURAL Y DEPORTIVO ELEMENTOS SPA"', html=False)
        self.assertContains(response, 'value="77.813.508-6"', html=False)
        self.assertContains(response, 'value="30252"', html=False)
        self.assertContains(response, 'value="5748"', html=False)
        self.assertContains(response, 'value="36000"', html=False)
        self.assertEqual(
            response.context["documento_form"].initial["observaciones"],
            "Taller de Fuerza/Calistenia - Plan 2 Clases Semanales (Erika Huerta)",
        )
        self.assertEqual(response.context["pago_form"].initial["metodo_pago"], Payment.Metodo.TRANSFERENCIA)
        self.assertEqual(response.context["pago_form"].initial["monto_referencia"], Decimal("36000"))
        self.assertTrue(response.context["pago_form"].initial["aplica_iva"])

    @patch("finanzas.documentos.parsers.PdfFallbackParser._extract_text_with_pdftotext")
    @patch("finanzas.documentos.parsers.PdfFallbackParser._extract_text_with_pypdf", return_value="")
    def test_documento_tributario_importar_precarga_boleta_venta_exenta_pdf(
        self,
        _extract_text_with_pypdf,
        extract_text_with_pdftotext,
    ):
        extract_text_with_pdftotext.return_value = """
        ESPACIO CULTURAL Y DEPORTIVO
        ELEMENTOS SPA
        77.813.508-6
        Giro: REALIZACIN DE ACTIVIDADES
        DEPORTIVAS Y CULTURALES PARA LA
        COMUNIDAD
        NUEVA TRES 1020
        Rengo
        BOLETA EXENTA ELECTRÓNICA NUMERO:
        25.043
        REF. VENDEDOR: 17085005-K
        Fecha: 2026-04-15

        Dirección: Santiago

        Medio de pago: Transferencia
        Electrónica
        Taller de Lyra - Plan 2 Clases Sema
        nales (Josefa Campos)
                                   $ 36.000
        Timbre Electrónico SII
        """
        pdf = SimpleUploadedFile(
            "boleta41.pdf",
            b"%PDF-1.4\n1 0 obj\n<<>>\nendobj\ntrailer\n<<>>\n%%EOF",
            content_type="application/pdf",
        )
        self.client.force_login(self.user_admin)

        response = self.client.post(
            self._url_con_organizacion(reverse("finanzas:documento_tributario_importar")),
            {"accion": "parsear", "archivo": pdf},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'value="25043"', html=False)
        self.assertContains(response, 'value="2026-04-15"', html=False)
        self.assertContains(response, 'value="ESPACIO CULTURAL Y DEPORTIVO ELEMENTOS SPA"', html=False)
        self.assertContains(response, 'value="77.813.508-6"', html=False)
        self.assertContains(response, 'value="36000"', html=False)
        self.assertEqual(
            response.context["documento_form"].initial["observaciones"],
            "Taller de Lyra - Plan 2 Clases Semanales (Josefa Campos)",
        )
        self.assertEqual(response.context["documento_form"].initial["monto_exento"], Decimal("36000"))
        self.assertEqual(response.context["documento_form"].initial["monto_iva"], Decimal("0"))
        self.assertEqual(response.context["documento_form"].initial["monto_neto"], Decimal("0"))
        self.assertEqual(response.context["pago_form"].initial["metodo_pago"], Payment.Metodo.TRANSFERENCIA)
        self.assertEqual(response.context["pago_form"].initial["monto_referencia"], Decimal("36000"))
        self.assertFalse(response.context["pago_form"].initial["aplica_iva"])

    def test_documento_tributario_importar_muestra_xml_subido_en_revision(self):
        xml = SimpleUploadedFile(
            "boleta.xml",
            b"""
            <EnvioDTE>
              <SetDTE>
                <DTE>
                  <Documento>
                    <Encabezado>
                      <IdDoc>
                        <TipoDTE>39</TipoDTE>
                        <Folio>101</Folio>
                        <FchEmis>2026-02-27</FchEmis>
                      </IdDoc>
                    </Encabezado>
                  </Documento>
                </DTE>
              </SetDTE>
            </EnvioDTE>
            """,
            content_type="application/xml",
        )
        self.client.force_login(self.user_admin)

        response = self.client.post(
            self._url_con_organizacion(reverse("finanzas:documento_tributario_importar")),
            {"accion": "parsear", "archivo": xml},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Visor del archivo subido")
        self.assertContains(response, "&lt;EnvioDTE&gt;", html=False)

        token = next(iter(self.client.session[SESSION_KEY].keys()))
        visor_url = reverse(
            "finanzas:documento_tributario_importacion_archivo",
            kwargs={"token": token, "tipo_archivo": "xml"},
        )
        self.assertContains(response, visor_url)

    def test_documento_tributario_importar_confirma_con_metadata_json_serializada(self):
        xml = SimpleUploadedFile(
            "boleta.xml",
            b"""
            <EnvioDTE>
              <SetDTE>
                <DTE>
                  <Documento>
                    <Encabezado>
                      <IdDoc>
                        <TipoDTE>39</TipoDTE>
                        <Folio>101</Folio>
                        <FchEmis>2026-02-27</FchEmis>
                      </IdDoc>
                      <Emisor>
                        <RUTEmisor>11.111.111-1</RUTEmisor>
                        <RznSoc>Org Finanzas</RznSoc>
                      </Emisor>
                      <Receptor>
                        <RUTRecep>22.222.222-2</RUTRecep>
                        <RznSocRecep>Ana Diaz</RznSocRecep>
                      </Receptor>
                      <Totales>
                        <MntNeto>10000</MntNeto>
                        <IVA>1900</IVA>
                        <MntTotal>11900</MntTotal>
                        <TasaIVA>19</TasaIVA>
                      </Totales>
                    </Encabezado>
                  </Documento>
                </DTE>
              </SetDTE>
            </EnvioDTE>
            """,
            content_type="application/xml",
        )
        self.client.force_login(self.user_admin)

        parse_response = self.client.post(
            f"{reverse('finanzas:documento_tributario_importar')}?periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}",
            {"accion": "parsear", "archivo": xml},
        )
        self.assertEqual(parse_response.status_code, 200)

        token = next(iter(self.client.session[SESSION_KEY].keys()))
        documento_initial = dict(parse_response.context["documento_form"].initial)
        post_data = {
            "accion": "confirmar",
            "token_importacion": token,
            **documento_initial,
        }

        confirm_response = self.client.post(
            f"{reverse('finanzas:documento_tributario_importar')}?periodo_mes=2&periodo_anio=2026&organizacion={self.org.pk}",
            post_data,
        )

        self.assertEqual(confirm_response.status_code, 302)
        documento = DocumentoTributario.objects.get(folio="101")
        self.assertIn("importacion_normalizada", documento.metadata_extra)
        self.assertIn("warnings_importacion", documento.metadata_extra)

    def test_documento_tributario_importar_permite_mismo_folio_si_cambia_emisor(self):
        DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.BOLETA_HONORARIOS,
            folio="45",
            fecha_emision="2026-03-12",
            rut_emisor="12.345.678-9",
            nombre_emisor="OTRO EMISOR",
            rut_receptor="77.813.508-6",
            nombre_receptor="ESPACIO CULTURAL Y DEPORTIVO ELEMENTOS SPA",
            monto_total=100000,
        )
        pdf_path = Path(__file__).resolve().parent.parent / "public" / "202603_LaTarea+Dificil.12Febrero2026_BarbaraAllendes.pdf"
        pdf = SimpleUploadedFile(
            pdf_path.name,
            pdf_path.read_bytes(),
            content_type="application/pdf",
        )
        self.client.force_login(self.user_admin)

        parse_response = self.client.post(
            f"{reverse('finanzas:documento_tributario_importar')}?periodo_mes=3&periodo_anio=2026&organizacion={self.org.pk}",
            {"accion": "parsear", "archivo": pdf},
        )

        self.assertEqual(parse_response.status_code, 200)
        self.assertEqual(parse_response.context["review_payload"]["duplicates"], [])

        token = next(iter(self.client.session[SESSION_KEY].keys()))
        post_data = {
            "accion": "confirmar",
            "token_importacion": token,
            **dict(parse_response.context["documento_form"].initial),
        }
        confirm_response = self.client.post(
            f"{reverse('finanzas:documento_tributario_importar')}?periodo_mes=3&periodo_anio=2026&organizacion={self.org.pk}",
            post_data,
        )

        self.assertEqual(confirm_response.status_code, 302)
        self.assertEqual(
            DocumentoTributario.objects.filter(
                organizacion=self.org,
                tipo_documento=DocumentoTributario.TipoDocumento.BOLETA_HONORARIOS,
                folio="45",
            ).count(),
            2,
        )
        self.assertTrue(
            DocumentoTributario.objects.filter(
                organizacion=self.org,
                tipo_documento=DocumentoTributario.TipoDocumento.BOLETA_HONORARIOS,
                folio="45",
                rut_emisor="18.445.523-4",
            ).exists()
        )

    def test_documento_tributario_importar_advierte_duplicado_por_folio_tipo_y_emisor(self):
        DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.BOLETA_HONORARIOS,
            folio="45",
            fecha_emision="2026-03-12",
            rut_emisor="18.445.523-4",
            nombre_emisor="BARBARA BEATRIZ ALLENDES HUERTA",
            rut_receptor="77.813.508-6",
            nombre_receptor="ESPACIO CULTURAL Y DEPORTIVO ELEMENTOS SPA",
            monto_total=100000,
        )
        pdf_path = Path(__file__).resolve().parent.parent / "public" / "202603_LaTarea+Dificil.12Febrero2026_BarbaraAllendes.pdf"
        pdf = SimpleUploadedFile(
            pdf_path.name,
            pdf_path.read_bytes(),
            content_type="application/pdf",
        )
        self.client.force_login(self.user_admin)

        response = self.client.post(
            f"{reverse('finanzas:documento_tributario_importar')}?periodo_mes=3&periodo_anio=2026&organizacion={self.org.pk}",
            {"accion": "parsear", "archivo": pdf},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["review_payload"]["duplicates"]), 1)
        self.assertContains(response, "Documento #")


class FinanzasIntegrationTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.org = Organizacion.objects.create(
            nombre="Org Integracion",
            razon_social="Org Integracion SPA",
            rut="33.333.333-3",
        )
        self.rol_admin = Rol.objects.create(nombre="Administrador", codigo="ADMINISTRADOR")
        self.rol_estudiante = Rol.objects.create(nombre="Estudiante", codigo="ESTUDIANTE")

        self.user_admin = User.objects.create_user("admin_int", password=TEST_PASSWORD)
        self.persona_admin = Persona.objects.create(
            nombres="Admin",
            apellidos="Integracion",
            email="adminint@example.com",
            user=self.user_admin,
        )
        PersonaRol.objects.create(
            persona=self.persona_admin,
            rol=self.rol_admin,
            organizacion=self.org,
            activo=True,
        )

        self.estudiante = Persona.objects.create(
            nombres="Ana",
            apellidos="Diaz",
            email="ana.int@example.com",
        )
        PersonaRol.objects.create(
            persona=self.estudiante,
            rol=self.rol_estudiante,
            organizacion=self.org,
            activo=True,
        )
        self.disciplina = Disciplina.objects.create(
            organizacion=self.org,
            nombre="Yoga",
        )
        self.sesion_1 = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-02-26",
            estado=SesionClase.Estado.PROGRAMADA,
        )
        self.sesion_2 = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-02-27",
            estado=SesionClase.Estado.PROGRAMADA,
        )

    def _crear_pago(self, fecha_pago="2026-02-25", clases_asignadas=1, numero_comprobante="PAGO-1"):
        return Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.org,
            fecha_pago=fecha_pago,
            metodo_pago=Payment.Metodo.TRANSFERENCIA,
            numero_comprobante=numero_comprobante,
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=clases_asignadas,
        )

    def _crear_sesion(self, fecha):
        return SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha=fecha,
            estado=SesionClase.Estado.PROGRAMADA,
        )

    def _crear_asistencia_presente(self, sesion):
        return Asistencia.objects.create(
            sesion=sesion,
            persona=self.estudiante,
            estado=Asistencia.Estado.PRESENTE,
        )

    def test_asistencia_sin_pago_queda_como_deuda(self):
        asistencia = self._crear_asistencia_presente(self.sesion_1)
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)

        self.assertEqual(consumo.persona, self.estudiante)
        self.assertEqual(consumo.clase_fecha, date(2026, 2, 26))
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.DEUDA)
        self.assertIsNone(consumo.pago)

    def test_asistencia_con_pago_disponible_queda_consumida(self):
        pago = self._crear_pago(clases_asignadas=1)

        asistencia = self._crear_asistencia_presente(self.sesion_2)
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)

        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.CONSUMIDO)
        self.assertEqual(consumo.pago, pago)
        self.assertEqual(pago.clases_consumidas, 1)
        self.assertEqual(pago.saldo_clases, 0)

    def test_asistencia_no_consume_pago_de_otro_mes(self):
        pago_enero = self._crear_pago(fecha_pago="2026-01-25", numero_comprobante="ENE-1")

        asistencia = self._crear_asistencia_presente(self.sesion_2)
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)

        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.DEUDA)
        self.assertIsNone(consumo.pago)
        self.assertEqual(pago_enero.saldo_clases, 1)

    def test_pago_nuevo_imputa_deudas_previas(self):
        asistencia_febrero = self._crear_asistencia_presente(self.sesion_1)
        asistencia_marzo = self._crear_asistencia_presente(self._crear_sesion("2026-03-01"))
        consumo_febrero = AttendanceConsumption.objects.get(asistencia=asistencia_febrero)
        consumo_marzo = AttendanceConsumption.objects.get(asistencia=asistencia_marzo)
        self.assertEqual(consumo_febrero.estado, AttendanceConsumption.Estado.DEUDA)
        self.assertEqual(consumo_marzo.estado, AttendanceConsumption.Estado.DEUDA)

        pago = self._crear_pago(fecha_pago="2026-02-28", clases_asignadas=2, numero_comprobante="FEB-1")

        consumo_febrero.refresh_from_db()
        consumo_marzo.refresh_from_db()
        self.assertEqual(consumo_febrero.estado, AttendanceConsumption.Estado.CONSUMIDO)
        self.assertEqual(consumo_febrero.pago, pago)
        self.assertEqual(consumo_marzo.estado, AttendanceConsumption.Estado.DEUDA)
        self.assertIsNone(consumo_marzo.pago)
        self.assertEqual(pago.clases_consumidas, 1)
        self.assertEqual(pago.saldo_clases, 1)

    def test_pago_no_imputa_deuda_de_otro_mes(self):
        asistencia = self._crear_asistencia_presente(self.sesion_1)
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.DEUDA)

        pago_marzo = self._crear_pago(fecha_pago="2026-03-05", numero_comprobante="MAR-1")

        consumo.refresh_from_db()
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.DEUDA)
        self.assertIsNone(consumo.pago)
        self.assertEqual(pago_marzo.saldo_clases, 1)

    def test_asociar_asistencia_a_pago_rechaza_pago_de_otro_mes(self):
        pago_otro_mes = self._crear_pago(fecha_pago="2026-03-05", numero_comprobante="MAR-2")
        asistencia = self._crear_asistencia_presente(self.sesion_1)

        with self.assertRaisesMessage(
            ValueError,
            "Solo se pueden asociar pagos del mismo mes y anio de la asistencia.",
        ):
            asociar_asistencia_a_pago(asistencia, pago_otro_mes)

    def test_resumen_financiero_estudiante_refleja_pagadas_consumidas_deuda_y_saldo(self):
        pago = self._crear_pago(fecha_pago="2026-02-25", clases_asignadas=3)
        asistencia_consumida = self._crear_asistencia_presente(self.sesion_1)
        asistencia_deuda = self._crear_asistencia_presente(self.sesion_2)
        consumo_consumido = AttendanceConsumption.objects.get(asistencia=asistencia_consumida)
        consumo_deuda = AttendanceConsumption.objects.get(asistencia=asistencia_deuda)

        consumo_consumido.pago = pago
        consumo_consumido.estado = AttendanceConsumption.Estado.CONSUMIDO
        consumo_consumido.save(update_fields=["pago", "estado"])
        consumo_deuda.pago = None
        consumo_deuda.estado = AttendanceConsumption.Estado.DEUDA
        consumo_deuda.save(update_fields=["pago", "estado"])

        resumen = resumen_financiero_estudiante(self.estudiante, organizacion=self.org)

        self.assertEqual(resumen["clases_pagadas"], 3)
        self.assertEqual(resumen["clases_consumidas"], 1)
        self.assertEqual(resumen["deuda_pendiente"], 1)
        self.assertEqual(resumen["saldo_clases"], 2)
        self.assertEqual(resumen["fecha_ultimo_pago"], date(2026, 2, 25))

    def test_pago_con_plan_respeta_monto_referencia_editable(self):
        plan = PaymentPlan.objects.create(
            organizacion=self.org,
            nombre="Plan Base",
            num_clases=4,
            precio=20000,
            activo=True,
        )
        pago = Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.org,
            plan=plan,
            fecha_pago="2026-02-28",
            metodo_pago=Payment.Metodo.TRANSFERENCIA,
            aplica_iva=False,
            monto_referencia=15000,
        )
        self.assertEqual(pago.monto_total, 15000)

    def test_form_transferencia_exige_numero_comprobante(self):
        data_base = {
            "organizacion": self.org.pk,
            "persona": self.estudiante.pk,
            "fecha_pago": "2026-02-28",
            "metodo_pago": "transferencia",
            "monto_referencia": "10000",
            "clases_asignadas": "1",
        }
        form = PaymentForm(data=data_base)
        self.assertFalse(form.is_valid())
        self.assertIn("numero_comprobante", form.errors)

        data_efectivo = {
            **data_base,
            "metodo_pago": "efectivo",
            "numero_comprobante": "",
        }
        form_efectivo = PaymentForm(data=data_efectivo)
        self.assertTrue(form_efectivo.is_valid(), form_efectivo.errors)

    def test_form_rechaza_plan_de_otra_organizacion(self):
        otra_org = Organizacion.objects.create(
            nombre="Org Externa",
            razon_social="Org Externa SPA",
            rut="44.444.444-4",
        )
        plan_otro = PaymentPlan.objects.create(
            organizacion=otra_org,
            nombre="Plan Otro",
            num_clases=4,
            precio=22000,
            activo=True,
        )
        form = PaymentForm(
            data={
                "organizacion": self.org.pk,
                "persona": self.estudiante.pk,
                "plan": plan_otro.pk,
                "fecha_pago": "2026-02-28",
                "metodo_pago": "efectivo",
                "monto_referencia": "10000",
                "clases_asignadas": "1",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("plan", form.errors)

    def test_form_rechaza_documento_tributario_de_otra_organizacion(self):
        otra_org = Organizacion.objects.create(
            nombre="Org Docs",
            razon_social="Org Docs SPA",
            rut="55.555.555-5",
        )
        documento_otro = DocumentoTributario.objects.create(
            organizacion=otra_org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_AFECTA,
            folio="F-1",
            fecha_emision="2026-02-28",
            monto_total=10000,
        )
        form = PaymentForm(
            data={
                "organizacion": self.org.pk,
                "persona": self.estudiante.pk,
                "documento_tributario": documento_otro.pk,
                "fecha_pago": "2026-02-28",
                "metodo_pago": "efectivo",
                "monto_referencia": "10000",
                "clases_asignadas": "1",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("documento_tributario", form.errors)

    def test_payment_form_filtra_documentos_por_organizacion_y_periodo(self):
        documento_febrero = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_AFECTA,
            folio="FEB-1",
            fecha_emision="2026-02-10",
            monto_total=10000,
        )
        documento_marzo = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_AFECTA,
            folio="MAR-1",
            fecha_emision="2026-03-10",
            monto_total=10000,
        )
        otra_org = Organizacion.objects.create(nombre="Org Otra", razon_social="Org Otra SPA", rut="66.666.666-6")
        documento_otra_org = DocumentoTributario.objects.create(
            organizacion=otra_org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_AFECTA,
            folio="OTRA-1",
            fecha_emision="2026-02-10",
            monto_total=10000,
        )

        form = PaymentForm(
            initial={"organizacion": self.org.pk},
            periodo_mes=2,
            periodo_anio=2026,
            organizacion=self.org,
        )

        queryset = form.fields["documento_tributario"].queryset
        self.assertIn(documento_febrero, queryset)
        self.assertNotIn(documento_marzo, queryset)
        self.assertNotIn(documento_otra_org, queryset)

    def test_payment_form_edicion_muestra_documento_actual_fuera_de_periodo(self):
        documento_marzo = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_AFECTA,
            folio="MAR-EDIT",
            fecha_emision="2026-03-10",
            monto_total=10000,
        )
        pago = Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.org,
            documento_tributario=documento_marzo,
            fecha_pago="2026-02-28",
            metodo_pago=Payment.Metodo.EFECTIVO,
            monto_referencia=10000,
            clases_asignadas=1,
        )

        form = PaymentForm(instance=pago, periodo_mes=2, periodo_anio=2026, organizacion=self.org)

        self.assertIn(documento_marzo, form.fields["documento_tributario"].queryset)

    def test_payment_form_rechaza_documento_fuera_de_periodo_en_creacion(self):
        documento_marzo = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_AFECTA,
            folio="MAR-POST",
            fecha_emision="2026-03-10",
            monto_total=10000,
        )

        form = PaymentForm(
            data={
                "organizacion": self.org.pk,
                "persona": self.estudiante.pk,
                "documento_tributario": documento_marzo.pk,
                "fecha_pago": "2026-02-28",
                "metodo_pago": "efectivo",
                "monto_referencia": "10000",
                "clases_asignadas": "1",
            },
            periodo_mes=2,
            periodo_anio=2026,
            organizacion=self.org,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("documento_tributario", form.errors)

    def test_transaction_form_filtra_documentos_por_organizacion_y_periodo(self):
        documento_febrero = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_AFECTA,
            folio="TX-FEB",
            fecha_emision="2026-02-10",
            monto_total=10000,
        )
        documento_marzo = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_AFECTA,
            folio="TX-MAR",
            fecha_emision="2026-03-10",
            monto_total=10000,
        )
        otra_org = Organizacion.objects.create(nombre="Org Docs TX", razon_social="Org Docs TX SPA", rut="77.777.777-7")
        documento_otra_org = DocumentoTributario.objects.create(
            organizacion=otra_org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_AFECTA,
            folio="TX-OTRA",
            fecha_emision="2026-02-10",
            monto_total=10000,
        )

        form = TransactionForm(periodo_mes=2, periodo_anio=2026, organizacion=self.org)

        queryset = form.fields["documentos_tributarios"].queryset
        self.assertIn(documento_febrero, queryset)
        self.assertNotIn(documento_marzo, queryset)
        self.assertNotIn(documento_otra_org, queryset)

    def test_transaction_form_edicion_muestra_documento_actual_fuera_de_periodo(self):
        categoria = Category.objects.create(nombre="Ingreso TX", tipo=Category.Tipo.INGRESO)
        documento_marzo = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_AFECTA,
            folio="TX-MAR-EDIT",
            fecha_emision="2026-03-10",
            monto_total=10000,
        )
        transaccion = Transaction.objects.create(
            organizacion=self.org,
            categoria=categoria,
            fecha="2026-02-28",
            tipo=Transaction.Tipo.INGRESO,
            monto=10000,
            descripcion="Movimiento",
        )
        transaccion.documentos_tributarios.add(documento_marzo)

        form = TransactionForm(instance=transaccion, periodo_mes=2, periodo_anio=2026, organizacion=self.org)

        self.assertIn(documento_marzo, form.fields["documentos_tributarios"].queryset)

    def test_transaction_form_rechaza_documento_de_otra_organizacion(self):
        categoria = Category.objects.create(nombre="Ingreso Otra Org", tipo=Category.Tipo.INGRESO)
        otra_org = Organizacion.objects.create(nombre="Org Docs Mal", razon_social="Org Docs Mal SPA", rut="88.888.888-8")
        documento_otra_org = DocumentoTributario.objects.create(
            organizacion=otra_org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_AFECTA,
            folio="TX-MAL",
            fecha_emision="2026-02-10",
            monto_total=10000,
        )

        form = TransactionForm(
            data={
                "organizacion": self.org.pk,
                "categoria": categoria.pk,
                "fecha": "2026-02-28",
                "tipo": "",
                "monto": "10000",
                "descripcion": "Movimiento",
                "documentos_tributarios": [documento_otra_org.pk],
            },
            periodo_mes=2,
            periodo_anio=2026,
            organizacion=self.org,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("documentos_tributarios", form.errors)

    def test_parse_tax_document_dte_xml_normaliza_boleta_venta(self):
        xml = b"""
        <EnvioDTE>
          <SetDTE>
            <DTE>
              <Documento>
                <Encabezado>
                  <IdDoc>
                    <TipoDTE>39</TipoDTE>
                    <Folio>555</Folio>
                    <FchEmis>2026-02-28</FchEmis>
                  </IdDoc>
                  <Emisor>
                    <RUTEmisor>11.111.111-1</RUTEmisor>
                    <RznSoc>Org Integracion</RznSoc>
                  </Emisor>
                  <Receptor>
                    <RUTRecep>22.222.222-2</RUTRecep>
                    <RznSocRecep>Cliente Uno</RznSocRecep>
                  </Receptor>
                  <Totales>
                    <MntNeto>20000</MntNeto>
                    <IVA>3800</IVA>
                    <MntTotal>23800</MntTotal>
                    <TasaIVA>19</TasaIVA>
                  </Totales>
                </Encabezado>
                <Detalle>
                  <NroLinDet>1</NroLinDet>
                  <NmbItem>Taller</NmbItem>
                  <QtyItem>2</QtyItem>
                  <PrcItem>10000</PrcItem>
                  <MontoItem>20000</MontoItem>
                </Detalle>
              </Documento>
            </DTE>
          </SetDTE>
        </EnvioDTE>
        """
        normalized = parse_tax_document(xml_bytes=xml, xml_name="dte.xml", organizacion_id=self.org.pk)

        self.assertEqual(normalized.get_value("encabezado", "categoria_documental"), "sales_receipt")
        self.assertEqual(normalized.get_value("encabezado", "tipo_documento_sugerido"), "boleta_venta_afecta")
        self.assertEqual(normalized.get_value("encabezado", "folio"), "555")
        self.assertEqual(normalized.get_value("montos", "total_bruto"), Decimal("23800"))
        self.assertEqual(len(normalized.lineas), 1)

    def test_parse_tax_document_bhe_xml_normaliza_retencion(self):
        xml = b"""
        <datos>
          <tipodoc>bhe</tipodoc>
          <numeroBoleta>9001</numeroBoleta>
          <fechaBoleta>2026-02-20</fechaBoleta>
          <rutEmisor>12345678</rutEmisor>
          <dvEmisor>9</dvEmisor>
          <rutReceptor>11111111</rutReceptor>
          <dvReceptor>1</dvReceptor>
          <nombreReceptor>Org Integracion</nombreReceptor>
          <domicilioEmisor>Calle Uno</domicilioEmisor>
          <domicilioReceptor>Calle Dos</domicilioReceptor>
          <actividadEconomica>Artista</actividadEconomica>
          <totalHonorarios>100000</totalHonorarios>
          <impuestoHonorarios>15250</impuestoHonorarios>
          <liquidoHonorarios>84750</liquidoHonorarios>
          <porcentajeImpuesto>15.25</porcentajeImpuesto>
          <prestacionServicios>
            <item>Presentacion artistica</item>
          </prestacionServicios>
        </datos>
        """
        normalized = parse_tax_document(xml_bytes=xml, xml_name="bhe.xml", organizacion_id=self.org.pk)

        self.assertEqual(normalized.get_value("encabezado", "categoria_documental"), "fee_receipt")
        self.assertEqual(normalized.get_value("encabezado", "tipo_documento_sugerido"), "boleta_honorarios")
        self.assertEqual(normalized.get_value("montos", "retencion_honorarios"), Decimal("15250"))
        self.assertEqual(normalized.get_value("montos", "porcentaje_retencion"), Decimal("15.25"))
        self.assertEqual(normalized.lineas[0].fields["descripcion"].value, "Presentacion artistica")

    @patch("finanzas.documentos.parsers.PdfFallbackParser._extract_text_with_pdftotext")
    @patch("finanzas.documentos.parsers.PdfFallbackParser._extract_text_with_pypdf", return_value="")
    def test_parse_tax_document_bhe_pdf_extrae_folio_fecha_y_montos(
        self,
        _extract_text_with_pypdf,
        extract_text_with_pdftotext,
    ):
        extract_text_with_pdftotext.return_value = """
                                                                                                            BOLETA DE HONORARIOS
                   ALVARO FRANCISCO VARGAS QUEZADA                                                              ELECTRONICA

                                                                                                                     N ° 125
                            RUT: 17.085.005−K
        GIRO(S): OTRAS ACTIVIDADES PROFESIONALES, CIENTIFICAS Y
                             TECNICAS N.C.P.,
                  SERVICIOS ARTISTICOS Y/O DEPORTIVOS
                             NUEVA 3 ST 51 EL NARANJAL , RENGO
                                   TELEFONO: 956490299

                                                                                                            Fecha: 12 de Marzo de 2026

Señor(es): ESPACIO CULTURAL Y DEPORTIVO ELEMENTOS SPA                                                   Rut: 77.813.508− 6
Domicilio: NUEVA TRES N 1020, EL NARANJAL NORTE, RENGO

Por atención profesional:
FUNCION LA TAREA MAS DIFICIL − FEBRERO − 2026                                                                                  100.000
                                                                                     Total Honorarios: $:                      100.000
                                                                                15.25 % Impto. Retenido:                        15.250
                                                                                                   Total:                       84.750
        """
        normalized = parse_tax_document(
            pdf_bytes=b"%PDF-1.4\n1 0 obj\n<<>>\nendobj\ntrailer\n<<>>\n%%EOF",
            pdf_name="honorarios.pdf",
            organizacion_id=self.org.pk,
        )

        self.assertEqual(normalized.get_value("encabezado", "tipo_documento_sugerido"), "boleta_honorarios")
        self.assertEqual(normalized.get_value("encabezado", "folio"), "125")
        self.assertEqual(normalized.get_value("encabezado", "fecha_emision"), "2026-03-12")
        self.assertEqual(normalized.get_value("emisor", "razon_social"), "ALVARO FRANCISCO VARGAS QUEZADA")
        self.assertEqual(normalized.get_value("emisor", "rut"), "17.085.005-K")
        self.assertEqual(normalized.get_value("receptor", "razon_social"), "ESPACIO CULTURAL Y DEPORTIVO ELEMENTOS SPA")
        self.assertEqual(normalized.get_value("receptor", "rut"), "77.813.508-6")
        self.assertEqual(normalized.get_value("montos", "neto"), Decimal("100000"))
        self.assertEqual(normalized.get_value("montos", "retencion_honorarios"), Decimal("15250"))
        self.assertEqual(normalized.get_value("montos", "porcentaje_retencion"), Decimal("15.25"))
        self.assertEqual(normalized.get_value("montos", "total_liquido"), Decimal("84750"))
        self.assertEqual(normalized.get_value("montos", "total_bruto"), Decimal("100000"))
        self.assertEqual(normalized.lineas[0].fields["descripcion"].value, "FUNCION LA TAREA MAS DIFICIL - FEBRERO - 2026")

    @patch("finanzas.documentos.parsers.PdfFallbackParser._extract_text_with_pdftotext")
    @patch("finanzas.documentos.parsers.PdfFallbackParser._extract_text_with_pypdf", return_value="")
    def test_parse_tax_document_boleta_venta_pdf_extrae_glosa_y_totales(
        self,
        _extract_text_with_pypdf,
        extract_text_with_pdftotext,
    ):
        extract_text_with_pdftotext.return_value = """
        ESPACIO CULTURAL Y DEPORTIVO
        ELEMENTOS SPA
        77.813.508-6
        Giro: REALIZACIN DE ACTIVIDADES
        DEPORTIVAS Y CULTURALES PARA LA
        COMUNIDAD
        NUEVA TRES 1020
        Rengo
        BOLETA ELECTRÓNICA NUMERO: 20.035
        REF. VENDEDOR: 17085005-K
        Fecha: 2026-04-15

        Dirección: Santiago

        Medio de pago: Transferencia
        Electrónica
        Taller de Fuerza/Calistenia - Plan
        2 Clases Semanales (Erika Huerta)
                                   $ 36.000

        El IVA incluido en esta boleta es
        de: $ 5.748
        """
        normalized = parse_tax_document(
            pdf_bytes=b"%PDF-1.4\n1 0 obj\n<<>>\nendobj\ntrailer\n<<>>\n%%EOF",
            pdf_name="boleta39.pdf",
            organizacion_id=self.org.pk,
        )

        self.assertEqual(normalized.get_value("encabezado", "tipo_documento_sugerido"), "boleta_venta_afecta")
        self.assertEqual(normalized.get_value("encabezado", "tipo_tributario"), "39")
        self.assertEqual(normalized.get_value("encabezado", "folio"), "20035")
        self.assertEqual(normalized.get_value("encabezado", "fecha_emision"), "2026-04-15")
        self.assertEqual(normalized.get_value("encabezado", "medio_pago"), "Transferencia Electrónica")
        self.assertEqual(normalized.get_value("emisor", "razon_social"), "ESPACIO CULTURAL Y DEPORTIVO ELEMENTOS SPA")
        self.assertEqual(normalized.get_value("emisor", "rut"), "77.813.508-6")
        self.assertEqual(normalized.get_value("montos", "neto"), Decimal("30252"))
        self.assertEqual(normalized.get_value("montos", "iva"), Decimal("5748"))
        self.assertEqual(normalized.get_value("montos", "tasa_iva"), Decimal("19"))
        self.assertEqual(normalized.get_value("montos", "total_bruto"), Decimal("36000"))
        self.assertEqual(normalized.lineas[0].fields["descripcion"].value, "Taller de Fuerza/Calistenia - Plan 2 Clases Semanales (Erika Huerta)")

    @patch("finanzas.documentos.parsers.PdfFallbackParser._extract_text_with_pdftotext")
    @patch("finanzas.documentos.parsers.PdfFallbackParser._extract_text_with_pypdf", return_value="")
    def test_parse_tax_document_boleta_venta_exenta_pdf_extrae_glosa_y_totales(
        self,
        _extract_text_with_pypdf,
        extract_text_with_pdftotext,
    ):
        extract_text_with_pdftotext.return_value = """
        ESPACIO CULTURAL Y DEPORTIVO
        ELEMENTOS SPA
        77.813.508-6
        Giro: REALIZACIN DE ACTIVIDADES
        DEPORTIVAS Y CULTURALES PARA LA
        COMUNIDAD
        NUEVA TRES 1020
        Rengo
        BOLETA EXENTA ELECTRÓNICA NUMERO:
        25.043
        REF. VENDEDOR: 17085005-K
        Fecha: 2026-04-15

        Dirección: Santiago

        Medio de pago: Transferencia
        Electrónica
        Taller de Lyra - Plan 2 Clases Sema
        nales (Josefa Campos)
                                   $ 36.000
        Timbre Electrónico SII
        """
        normalized = parse_tax_document(
            pdf_bytes=b"%PDF-1.4\n1 0 obj\n<<>>\nendobj\ntrailer\n<<>>\n%%EOF",
            pdf_name="boleta41.pdf",
            organizacion_id=self.org.pk,
        )

        self.assertEqual(normalized.get_value("encabezado", "tipo_documento_sugerido"), "boleta_venta_exenta")
        self.assertEqual(normalized.get_value("encabezado", "tipo_tributario"), "41")
        self.assertEqual(normalized.get_value("encabezado", "folio"), "25043")
        self.assertEqual(normalized.get_value("encabezado", "fecha_emision"), "2026-04-15")
        self.assertEqual(normalized.get_value("encabezado", "medio_pago"), "Transferencia Electrónica")
        self.assertEqual(normalized.get_value("emisor", "razon_social"), "ESPACIO CULTURAL Y DEPORTIVO ELEMENTOS SPA")
        self.assertEqual(normalized.get_value("emisor", "rut"), "77.813.508-6")
        self.assertEqual(normalized.get_value("montos", "neto"), Decimal("0"))
        self.assertEqual(normalized.get_value("montos", "exento"), Decimal("36000"))
        self.assertEqual(normalized.get_value("montos", "iva"), Decimal("0"))
        self.assertEqual(normalized.get_value("montos", "tasa_iva"), Decimal("0"))
        self.assertEqual(normalized.get_value("montos", "total_bruto"), Decimal("36000"))
        self.assertEqual(normalized.lineas[0].fields["descripcion"].value, "Taller de Lyra - Plan 2 Clases Semanales (Josefa Campos)")

    def test_form_edicion_pago_renderiza_fecha_iso_para_input_date(self):
        pago = Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.org,
            fecha_pago="2026-02-28",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=1,
        )
        form = PaymentForm(instance=pago)
        html = form["fecha_pago"].as_widget()
        self.assertIn('value="2026-02-28"', html)

    def test_form_edicion_documento_tributario_renderiza_fecha_iso_para_input_date(self):
        documento = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_EXENTA,
            fuente=DocumentoTributario.Fuente.MANUAL,
            folio="502",
            fecha_emision="2026-03-10",
            nombre_emisor="Emisor",
            rut_emisor="11.111.111-1",
            nombre_receptor="Receptor",
            rut_receptor="22.222.222-2",
            monto_neto=0,
            monto_exento=500000,
            monto_iva=0,
            monto_total=500000,
        )
        form = DocumentoTributarioForm(instance=documento)
        html = form["fecha_emision"].as_widget()
        self.assertIn('value="2026-03-10"', html)

    def test_documento_tributario_form_rechaza_persona_y_organizacion_asociada_al_mismo_tiempo(self):
        persona = Persona.objects.create(nombres="Julia", apellidos="Lopez", rut="12.345.678-5")
        otra_org = Organizacion.objects.create(nombre="Org Tercera", razon_social="Org Tercera SPA", rut="44.444.444-4")
        form = DocumentoTributarioForm(
            data={
                "organizacion": self.org.pk,
                "tipo_documento": DocumentoTributario.TipoDocumento.FACTURA_EXENTA,
                "fuente": DocumentoTributario.Fuente.MANUAL,
                "folio": "501",
                "fecha_emision": "2026-03-10",
                "nombre_emisor": "Emisor",
                "rut_emisor": "11.111.111-1",
                "nombre_receptor": "Receptor",
                "rut_receptor": "22.222.222-2",
                "monto_neto": "0",
                "monto_exento": "500000",
                "iva_tasa": "0",
                "monto_iva": "0",
                "retencion_tasa": "0",
                "retencion_monto": "0",
                "monto_total": "500000",
                "persona_relacionada": persona.pk,
                "organizacion_relacionada": otra_org.pk,
                "metadata_extra": "{}",
                "observaciones": "",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("organizacion_relacionada", form.errors)

    def test_documento_tributario_edit_permite_asociar_persona(self):
        persona = Persona.objects.create(nombres="Julia", apellidos="Lopez", rut="12.345.678-5")
        documento = DocumentoTributario.objects.create(
            organizacion=self.org,
            tipo_documento=DocumentoTributario.TipoDocumento.FACTURA_EXENTA,
            fuente=DocumentoTributario.Fuente.MANUAL,
            folio="777",
            fecha_emision="2026-03-10",
            nombre_emisor="Emisor",
            rut_emisor="11.111.111-1",
            nombre_receptor="Receptor",
            rut_receptor="22.222.222-2",
            monto_exento=500000,
            monto_total=500000,
        )
        self.client.force_login(self.user_admin)

        response = self.client.post(
            f"{reverse('finanzas:documento_tributario_edit', kwargs={'pk': documento.pk})}?organizacion={self.org.pk}",
            {
                "organizacion": self.org.pk,
                "tipo_documento": DocumentoTributario.TipoDocumento.FACTURA_EXENTA,
                "fuente": DocumentoTributario.Fuente.MANUAL,
                "folio": "777",
                "fecha_emision": "2026-03-10",
                "nombre_emisor": "Emisor",
                "rut_emisor": "11.111.111-1",
                "nombre_receptor": "Receptor",
                "rut_receptor": "22.222.222-2",
                "monto_neto": "0",
                "monto_exento": "500000",
                "iva_tasa": "0",
                "monto_iva": "0",
                "retencion_tasa": "0",
                "retencion_monto": "0",
                "monto_total": "500000",
                "documento_relacionado": "",
                "persona_relacionada": persona.pk,
                "organizacion_relacionada": "",
                "enlace_sii": "",
                "metadata_extra": "{}",
                "observaciones": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        documento.refresh_from_db()
        self.assertEqual(documento.persona_relacionada, persona)
        self.assertIsNone(documento.organizacion_relacionada)

    def test_documento_tributario_form_acepta_montos_con_punto_como_miles(self):
        form = DocumentoTributarioForm(
            data={
                "organizacion": self.org.pk,
                "tipo_documento": DocumentoTributario.TipoDocumento.FACTURA_EXENTA,
                "fuente": DocumentoTributario.Fuente.MANUAL,
                "folio": "501",
                "fecha_emision": "2026-03-10",
                "nombre_emisor": "Emisor",
                "rut_emisor": "11.111.111-1",
                "nombre_receptor": "Receptor",
                "rut_receptor": "22.222.222-2",
                "monto_neto": "0",
                "monto_exento": "500.000",
                "iva_tasa": "0",
                "monto_iva": "0",
                "retencion_tasa": "0",
                "retencion_monto": "0",
                "monto_total": "500.000",
                "metadata_extra": "{}",
                "observaciones": "",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["monto_exento"], Decimal("500000"))
        self.assertEqual(form.cleaned_data["monto_total"], Decimal("500000"))

    def test_pagos_list_muestra_resumen_compilado_del_listado(self):
        self.client.force_login(self.user_admin)

        pago_1 = Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.org,
            fecha_pago="2026-02-25",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=2,
        )
        Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.org,
            fecha_pago="2026-02-28",
            metodo_pago=Payment.Metodo.TRANSFERENCIA,
            numero_comprobante="ABC123",
            aplica_iva=False,
            monto_referencia=15000,
            clases_asignadas=3,
        )
        Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.org,
            fecha_pago="2026-03-02",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=5000,
            clases_asignadas=1,
        )

        asistencia = Asistencia.objects.create(sesion=self.sesion_1, persona=self.estudiante)
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.pago, pago_1)

        response = self.client.get(
            reverse("finanzas:pagos_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_pagos_monto"], 25000)
        self.assertEqual(response.context["total_clases_pagadas"], 5)
        self.assertEqual(response.context["total_saldo_clases"], 4)

    def test_pagos_list_muestra_estado_fiscal_y_texto_copiable(self):
        self.client.force_login(self.user_admin)
        plan = PaymentPlan.objects.create(
            organizacion=self.org,
            nombre="Plan Mensual",
            num_clases=4,
            precio=10000,
            activo=True,
        )
        Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.org,
            plan=plan,
            fecha_pago="2026-02-25",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=True,
            monto_incluye_iva=False,
            monto_referencia=10000,
            clases_asignadas=4,
        )
        Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.org,
            fecha_pago="2026-02-26",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=8000,
            clases_asignadas=1,
        )
        disciplina_secundaria = Disciplina.objects.create(
            organizacion=self.org,
            nombre="Pilates",
        )
        sesion_pilates = SesionClase.objects.create(
            disciplina=disciplina_secundaria,
            fecha="2026-02-24",
            estado=SesionClase.Estado.PROGRAMADA,
        )
        Asistencia.objects.create(sesion=self.sesion_1, persona=self.estudiante)
        Asistencia.objects.create(sesion=self.sesion_2, persona=self.estudiante)
        Asistencia.objects.create(sesion=sesion_pilates, persona=self.estudiante)
        Asistencia.objects.filter(sesion=sesion_pilates, persona=self.estudiante).update(
            estado=Asistencia.Estado.AUSENTE
        )

        response = self.client.get(
            reverse("finanzas:pagos_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.org.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Haz clic sobre los montos de neto, IVA o bruto para copiar el valor sin puntos.")
        self.assertContains(response, "<th>IVA</th>", html=False)
        self.assertNotContains(response, "<th>Organizacion</th>", html=False)
        self.assertContains(response, "Afecta")
        self.assertContains(response, "Exenta")
        self.assertContains(response, "11.900")
        self.assertContains(response, "10.000")
        self.assertContains(response, "1.900")
        self.assertContains(response, "8.000")
        self.assertContains(response, "bi-chat-text", html=False)
        self.assertContains(response, 'data-copy-value="10000"', html=False)
        self.assertContains(response, 'data-copy-value="1900"', html=False)
        self.assertContains(response, 'data-copy-value="11900"', html=False)
        self.assertContains(response, 'title="$ 11.900 · clic para copiar 11900"', html=False)
        self.assertContains(
            response,
            'data-copy-text="Taller de Yoga - Plan Mensual (Ana Diaz)"',
            html=False,
        )
        self.assertContains(
            response,
            'title="Taller de Yoga - Plan Mensual (Ana Diaz) · clic para copiar"',
            html=False,
        )

        pago = next(item for item in response.context["pagos"] if item.plan_id)
        self.assertEqual(pago.disciplina_principal_nombre, "Yoga")
        self.assertEqual(pago.texto_copia, "Taller de Yoga - Plan Mensual (Ana Diaz)")

    def test_servicio_pagos_enriquece_filas_para_listado(self):
        plan = PaymentPlan.objects.create(
            organizacion=self.org,
            nombre="Plan Mensual",
            num_clases=4,
            precio=10000,
            activo=True,
        )
        pago = Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.org,
            plan=plan,
            fecha_pago="2026-02-25",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=True,
            monto_referencia=10000,
            clases_asignadas=4,
        )
        pago.disciplina_principal_nombre = "Yoga"

        pagos = enriquecer_pagos_para_listado([pago])

        self.assertEqual(pagos[0].estado_fiscal_label, "Afecta")
        self.assertEqual(pagos[0].estado_fiscal_badge_class, "text-bg-primary")
        self.assertEqual(pagos[0].texto_copia, "Taller de Yoga - Plan Mensual (Ana Diaz)")
        self.assertEqual(pagos[0].monto_neto_copia, "10000")
        self.assertEqual(pagos[0].monto_iva_copia, "1900")
        self.assertEqual(pagos[0].monto_total_copia, "11900")

    def test_servicio_pagos_crea_persona_estudiante_desde_modal(self):
        form = PersonaRapidaForm(
            data={
                "nombres": "Lucia",
                "apellidos": "Perez",
                "telefono": "999999",
            }
        )
        self.assertTrue(form.is_valid(), form.errors)

        persona = crear_persona_estudiante_desde_modal(form=form, organizacion=self.org)

        self.assertEqual(persona.nombres, "Lucia")
        self.assertEqual(persona.apellidos, "Perez")
        self.assertTrue(
            PersonaRol.objects.filter(
                persona=persona,
                rol=self.rol_estudiante,
                organizacion=self.org,
                activo=True,
            ).exists()
        )

    def test_servicio_pagos_resume_consumos_y_saldo(self):
        pago = self._crear_pago(fecha_pago="2026-02-25", clases_asignadas=3)
        asistencia_consumida = self._crear_asistencia_presente(self.sesion_1)
        asistencia_deuda = self._crear_asistencia_presente(self.sesion_2)
        consumo_consumido = AttendanceConsumption.objects.get(asistencia=asistencia_consumida)
        consumo_deuda = AttendanceConsumption.objects.get(asistencia=asistencia_deuda)

        consumo_consumido.pago = pago
        consumo_consumido.estado = AttendanceConsumption.Estado.CONSUMIDO
        consumo_consumido.save(update_fields=["pago", "estado"])
        consumo_deuda.pago = pago
        consumo_deuda.estado = AttendanceConsumption.Estado.DEUDA
        consumo_deuda.save(update_fields=["pago", "estado"])

        resumen = resumen_consumos_pago(pago)

        self.assertEqual(resumen["consumos"], [consumo_deuda, consumo_consumido])
        self.assertEqual(resumen["consumos_consumidos"], 1)
        self.assertEqual(resumen["consumos_pendientes"], 0)
        self.assertEqual(resumen["consumos_deuda"], 1)
        self.assertEqual(resumen["saldo_clases"], 2)

    def test_servicio_pagos_texto_copiable_usa_fallbacks_operativos(self):
        pago = self._crear_pago(fecha_pago="2026-02-25", clases_asignadas=1)
        pago.disciplina_principal_nombre = ""

        self.assertEqual(
            texto_copiable_operativo_pago(pago),
            "Taller de Sin disciplina - Sin plan (Ana Diaz)",
        )


class SprintDosReversaPagosTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.organizacion = Organizacion.objects.create(
            nombre="Org Reversa Sprint 2",
            razon_social="Org Reversa Sprint 2 SpA",
            rut="71.000.000-1",
        )
        self.otra_organizacion = Organizacion.objects.create(
            nombre="Otra Reversa Sprint 2",
            razon_social="Otra Reversa Sprint 2 SpA",
            rut="71.000.000-2",
        )
        self.rol_admin = Rol.objects.create(nombre="Admin Reversa", codigo="ADMIN")
        self.rol_finanzas = Rol.objects.create(nombre="Finanzas Reversa", codigo="FINANZAS")
        self.rol_profesor = Rol.objects.create(nombre="Profesor Reversa", codigo="PROFESOR")
        self.rol_estudiante = Rol.objects.create(nombre="Estudiante Reversa", codigo="ESTUDIANTE")
        self.admin = self._crear_usuario_rol("admin_reversa", self.rol_admin, self.organizacion)
        self.finanzas = self._crear_usuario_rol("finanzas_reversa", self.rol_finanzas, self.organizacion)
        self.profesor = self._crear_usuario_rol("profesor_reversa", self.rol_profesor, self.organizacion)
        self.admin_otra = self._crear_usuario_rol("admin_reversa_otra", self.rol_admin, self.otra_organizacion)
        self.estudiante = Persona.objects.create(nombres="Alumno", apellidos="Reversa")
        PersonaRol.objects.create(
            persona=self.estudiante,
            rol=self.rol_estudiante,
            organizacion=self.organizacion,
            activo=True,
        )
        self.disciplina = Disciplina.objects.create(
            organizacion=self.organizacion,
            nombre="Disciplina Reversa",
        )
        self.sesion = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha=date(2026, 7, 15),
        )

    def _crear_usuario_rol(self, username, rol, organizacion):
        return crear_usuario_con_rol(
            username=username,
            password=TEST_PASSWORD,
            rol=rol,
            organizacion=organizacion,
            apellidos="Sprint",
        )

    def _crear_pago(self, *, clases=1):
        return Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.organizacion,
            fecha_pago=date(2026, 7, 2),
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=20000,
            clases_asignadas=clases,
        )

    def test_revertir_pago_preserva_historia_traza_y_recalcula_consumo(self):
        pago = self._crear_pago()
        asistencia = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.CONSUMIDO)

        with self.captureOnCommitCallbacks(execute=True):
            revertido = revertir_pago(
                pago=pago,
                motivo="Transferencia anulada por banco",
                usuario=self.admin,
            )

        self.assertEqual(revertido.pk, pago.pk)
        self.assertIsNotNone(revertido.revertido_en)
        self.assertEqual(revertido.revertido_por, self.admin)
        self.assertEqual(revertido.motivo_reversa, "Transferencia anulada por banco")
        consumo.refresh_from_db()
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.DEUDA)
        self.assertIsNone(consumo.pago)
        self.assertTrue(Payment.objects.filter(pk=pago.pk).exists())
        self.assertTrue(AuditLog.objects.filter(objeto_id=str(pago.pk), resumen="Pago revertido").exists())

    def test_revertir_pago_reasigna_consumo_a_otro_derecho_valido(self):
        primero = self._crear_pago()
        segundo = self._crear_pago()
        asistencia = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.pago, primero)

        revertir_pago(pago=primero, motivo="Pago duplicado", usuario=self.admin)

        consumo.refresh_from_db()
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.CONSUMIDO)
        self.assertEqual(consumo.pago, segundo)

    def test_revertir_pago_exige_motivo_e_impide_segunda_reversa(self):
        pago = self._crear_pago()
        with self.assertRaisesMessage(ValidationError, "motivo"):
            revertir_pago(pago=pago, motivo="", usuario=self.admin)
        revertir_pago(pago=pago, motivo="Primera reversa", usuario=self.admin)
        with self.assertRaisesMessage(ValidationError, "ya fue revertido"):
            revertir_pago(pago=pago, motivo="Segunda reversa", usuario=self.admin)

    def test_reversa_pago_restringida_a_admin_de_la_organizacion(self):
        pago = self._crear_pago()
        url = (
            reverse("finanzas:pago_revertir", kwargs={"pk": pago.pk})
            + f"?organizacion={self.organizacion.pk}"
        )
        for usuario, esperado in (
            (self.admin, 302),
            (self.finanzas, 403),
            (self.profesor, 403),
            (self.admin_otra, 403),
        ):
            with self.subTest(usuario=usuario.username):
                if pago.revertido_en:
                    pago = self._crear_pago()
                    url = (
                        reverse("finanzas:pago_revertir", kwargs={"pk": pago.pk})
                        + f"?organizacion={self.organizacion.pk}"
                    )
                self.client.force_login(usuario)
                response = self.client.post(url, {"motivo": "Motivo de prueba"})
                self.assertEqual(response.status_code, esperado)
                pago.refresh_from_db()

    def test_accion_reversa_solo_es_visible_para_admin_autorizado(self):
        pago = self._crear_pago()
        url = reverse("finanzas:pagos_list")
        params = {
            "periodo_mes": 7,
            "periodo_anio": 2026,
            "organizacion": self.organizacion.pk,
        }
        self.client.force_login(self.admin)
        response_admin = self.client.get(url, params)
        self.assertContains(
            response_admin,
            reverse("finanzas:pago_revertir", kwargs={"pk": pago.pk}),
        )

        for usuario in (self.finanzas, self.profesor, self.admin_otra):
            with self.subTest(usuario=usuario.username):
                self.client.force_login(usuario)
                response = self.client.get(url, params)
                if usuario == self.finanzas:
                    self.assertEqual(response.status_code, 200)
                    self.assertNotContains(
                        response,
                        reverse("finanzas:pago_revertir", kwargs={"pk": pago.pk}),
                    )
                else:
                    self.assertEqual(response.status_code, 403)

    def test_listado_muestra_revertido_y_excluye_pago_de_resumen(self):
        vigente = self._crear_pago(clases=2)
        revertido = self._crear_pago(clases=3)
        revertir_pago(pago=revertido, motivo="No vigente", usuario=self.admin)
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse("finanzas:pagos_list"),
            {
                "periodo_mes": 7,
                "periodo_anio": 2026,
                "organizacion": self.organizacion.pk,
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Revertido")
        self.assertEqual(response.context["total_clases_pagadas"], vigente.clases_asignadas)
        self.assertEqual(response.context["total_pagos_monto"], vigente.monto_total)


class SprintDosReconciliacionTests(TestCase):
    def setUp(self):
        self.organizacion = Organizacion.objects.create(
            nombre="Org Reconciliación",
            razon_social="Org Reconciliación SpA",
            rut="71.000.000-3",
        )
        self.otra_organizacion = Organizacion.objects.create(
            nombre="Otra Reconciliación",
            razon_social="Otra Reconciliación SpA",
            rut="71.000.000-4",
        )
        self.estudiante = Persona.objects.create(nombres="Alumno", apellidos="Reconciliación")
        self.disciplina = Disciplina.objects.create(
            organizacion=self.organizacion,
            nombre="Reconciliación",
        )
        self.sesion = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha=date(2026, 7, 20),
        )

    def _pago(self, organizacion=None, *, clases=1, fecha_pago=None):
        return Payment.objects.create(
            persona=self.estudiante,
            organizacion=organizacion or self.organizacion,
            fecha_pago=fecha_pago or date(2026, 7, 1),
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=clases,
        )

    def _otra_asistencia(self, *, dia=21):
        sesion = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha=date(2026, 7, dia),
        )
        return Asistencia.objects.create(sesion=sesion, persona=self.estudiante)

    def test_reconciliacion_permite_consumos_compartidos_con_cupo_suficiente(self):
        pago = self._pago(clases=2)
        primera = Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
        )
        segunda = self._otra_asistencia()
        resultado = reconciliar_integridad_dominio()

        self.assertTrue(resultado["ok"])
        self.assertFalse(any(resultado["resumen"].values()))
        self.assertEqual(
            AttendanceConsumption.objects.filter(
                asistencia__in=[primera, segunda],
                pago=pago,
                estado=AttendanceConsumption.Estado.CONSUMIDO,
            ).count(),
            2,
        )

        salida = StringIO()
        call_command("reconciliar_integridad_dominio", stdout=salida)
        self.assertIn("Sin inconsistencias de dominio", salida.getvalue())

    def test_reconciliacion_detecta_sobreconsumo_respecto_de_clases_asignadas(self):
        pago = self._pago(clases=1)
        primera = Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
        )
        segunda = self._otra_asistencia()
        AttendanceConsumption.objects.filter(asistencia=segunda).update(
            estado=AttendanceConsumption.Estado.CONSUMIDO,
            pago=pago,
        )

        resultado = reconciliar_integridad_dominio()

        self.assertEqual(resultado["resumen"]["sobreconsumo_pago"], 1)
        self.assertEqual(
            resultado["detalle"]["sobreconsumo_pago"],
            [
                {
                    "pago_id": pago.pk,
                    "organizacion_id": self.organizacion.pk,
                    "clases_asignadas": 1,
                    "consumos_consumidos": 2,
                }
            ],
        )
        self.assertEqual(
            AttendanceConsumption.objects.filter(
                asistencia__in=[primera, segunda],
                pago=pago,
            ).count(),
            2,
        )

    def test_reconciliacion_detecta_consumo_fuera_periodo_y_sin_pago(self):
        pago_anterior = self._pago(fecha_pago=date(2026, 6, 1))
        fuera_periodo = Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
        )
        huerfano = self._otra_asistencia()
        AttendanceConsumption.objects.filter(asistencia=fuera_periodo).update(
            estado=AttendanceConsumption.Estado.CONSUMIDO,
            pago=pago_anterior,
        )
        AttendanceConsumption.objects.filter(asistencia=huerfano).update(
            estado=AttendanceConsumption.Estado.CONSUMIDO,
            pago=None,
        )

        resultado = reconciliar_integridad_dominio()

        self.assertEqual(resultado["resumen"]["consumo_fuera_periodo"], 1)
        self.assertEqual(resultado["resumen"]["consumo_sin_pago"], 1)

    def test_reconciliacion_detecta_consumo_de_otra_persona_u_organizacion(self):
        pago_ajeno = self._pago(organizacion=self.otra_organizacion)
        asistencia = Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
            estado=Asistencia.Estado.AUSENTE,
        )
        AttendanceConsumption.objects.filter(asistencia=asistencia).update(
            estado=AttendanceConsumption.Estado.CONSUMIDO,
            pago=pago_ajeno,
        )
        resultado = reconciliar_integridad_dominio()
        self.assertFalse(resultado["ok"])
        self.assertEqual(
            resultado["resumen"]["consumo_otra_persona_organizacion"],
            1,
        )

        with self.assertRaises(CommandError):
            call_command("reconciliar_integridad_dominio", stdout=StringIO())

    def test_reconciliacion_detecta_liberada_consumiendo_y_pago_revertido(self):
        user = get_user_model().objects.create_user("audit_reconciliacion", password=TEST_PASSWORD)
        pago = self._pago()
        asistencia = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        ClaseLiberada.objects.create(
            asistencia=asistencia,
            organizacion=self.organizacion,
            motivo="Inconsistencia inducida",
            liberada_por=user,
        )
        Payment.objects.filter(pk=pago.pk).update(
            revertido_en=timezone.now(),
            revertido_por=user,
            motivo_reversa="Inconsistencia inducida",
        )
        resultado = reconciliar_integridad_dominio()
        self.assertEqual(resultado["resumen"]["clase_liberada_consumiendo"], 1)
        self.assertEqual(resultado["resumen"]["pago_revertido_incluido"], 1)

    def test_reconciliacion_detecta_estado_ordinario_pendiente(self):
        asistencia = Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
            estado=Asistencia.Estado.JUSTIFICADA,
        )
        AttendanceConsumption.objects.filter(asistencia=asistencia).update(
            estado=AttendanceConsumption.Estado.PENDIENTE,
            pago=None,
        )

        resultado = reconciliar_integridad_dominio()

        self.assertEqual(resultado["resumen"]["estado_asistencia_incompatible"], 1)


class PagoMasivoDominioTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.org = Organizacion.objects.create(nombre="Org Lotes", razon_social="Org Lotes SPA", rut="76.000.000-1")
        self.otra_org = Organizacion.objects.create(nombre="Otra Org Lotes", razon_social="Otra Org Lotes SPA", rut="76.000.000-2")
        self.rol_admin = Rol.objects.create(nombre="Admin lotes", codigo="ADMIN")
        self.rol_estudiante = Rol.objects.create(nombre="Estudiante lotes", codigo="ESTUDIANTE")
        self.user = User.objects.create_user("admin_lotes", password=TEST_PASSWORD)
        self.admin = Persona.objects.create(nombres="Admin", apellidos="Lotes", user=self.user)
        PersonaRol.objects.create(persona=self.admin, rol=self.rol_admin, organizacion=self.org, activo=True)
        self.plan = PaymentPlan.objects.create(
            organizacion=self.org,
            nombre="Plan lote",
            num_clases=2,
            precio=10000,
            activo=True,
        )
        self.personas = []
        for index in range(20):
            persona = Persona.objects.create(nombres=f"Alumno{index}", apellidos="Lote")
            PersonaRol.objects.create(persona=persona, rol=self.rol_estudiante, organizacion=self.org, activo=True)
            self.personas.append(persona)

    def _filas(self, cantidad):
        return [
            {
                "persona_id": persona.pk,
                "plan_id": self.plan.pk,
                "documento_tributario_id": None,
                "fecha_pago": date(2026, 7, 27),
                "metodo_pago": Payment.Metodo.EFECTIVO,
                "numero_comprobante": "",
                "aplica_iva": False,
                "monto_incluye_iva": False,
                "monto_referencia": Decimal("10000"),
                "clases_asignadas": 0,
                "observaciones": "lote de prueba",
            }
            for persona in self.personas[:cantidad]
        ]

    def test_lote_valido_de_10_crea_pagos_y_auditoria_de_lote(self):
        with self.captureOnCommitCallbacks(execute=True):
            lote, creado = confirmar_lote_pagos(
                usuario=self.user,
                organizacion_id=self.org.pk,
                clave_idempotencia="lote-10",
                filas=self._filas(10),
            )
        self.assertTrue(creado)
        self.assertEqual(lote.cantidad_pagos, 10)
        self.assertEqual(Payment.objects.filter(lote=lote).count(), 10)
        self.assertEqual(AuditLog.objects.filter(objeto_id=str(lote.pk)).count(), 1)

    def test_lote_valido_de_20_conserva_mismos_montos_del_pago_individual(self):
        lote, creado = confirmar_lote_pagos(
            usuario=self.user,
            organizacion_id=self.org.pk,
            clave_idempotencia="lote-20",
            filas=self._filas(20),
        )
        self.assertTrue(creado)
        pagos = list(Payment.objects.filter(lote=lote))
        self.assertEqual(len(pagos), 20)
        self.assertTrue(all(pago.monto_total == Decimal("10000.00") for pago in pagos))
        self.assertTrue(all(pago.clases_asignadas == 2 for pago in pagos))

    def test_fila_invalida_hace_rollback_de_todo_el_lote(self):
        filas = self._filas(10)
        filas[-1]["persona_id"] = self.personas[0].pk
        with self.assertRaises(ValidationError):
            confirmar_lote_pagos(
                usuario=self.user,
                organizacion_id=self.org.pk,
                clave_idempotencia="lote-invalido",
                filas=filas,
            )
        self.assertEqual(Payment.objects.count(), 0)
        self.assertEqual(LotePago.objects.count(), 0)

    def test_misma_clave_idempotente_no_duplica_pagos(self):
        lote, creado = confirmar_lote_pagos(
            usuario=self.user,
            organizacion_id=self.org.pk,
            clave_idempotencia="lote-reintento",
            filas=self._filas(10),
        )
        repetido, creado_repetido = confirmar_lote_pagos(
            usuario=self.user,
            organizacion_id=self.org.pk,
            clave_idempotencia="lote-reintento",
            filas=self._filas(10),
        )
        self.assertTrue(creado)
        self.assertFalse(creado_repetido)
        self.assertEqual(repetido.pk, lote.pk)
        self.assertEqual(Payment.objects.count(), 10)

    def test_persona_de_otra_organizacion_no_puede_formar_parte_del_lote(self):
        persona_ajena = Persona.objects.create(nombres="Ajena", apellidos="Lote")
        PersonaRol.objects.create(persona=persona_ajena, rol=self.rol_estudiante, organizacion=self.otra_org, activo=True)
        filas = self._filas(1)
        filas[0]["persona_id"] = persona_ajena.pk
        with self.assertRaisesMessage(ValidationError, "persona seleccionada"):
            confirmar_lote_pagos(
                usuario=self.user,
                organizacion_id=self.org.pk,
                clave_idempotencia="lote-ajeno",
                filas=filas,
            )
        self.assertEqual(Payment.objects.count(), 0)

    def test_vista_masiva_y_busqueda_estan_limitadas_por_organizacion(self):
        self.client.force_login(self.user)
        url = reverse("finanzas:pago_masivo")
        response = self.client.get(f"{url}?organizacion={self.org.pk}")
        self.assertEqual(response.status_code, 200)
        busqueda = self.client.get(
            reverse("finanzas:pago_masivo_personas"),
            {"organizacion": self.org.pk, "q": "Alumno"},
        )
        self.assertEqual(busqueda.status_code, 200)
        self.assertEqual(len(busqueda.json()["resultados"]), 20)
        persona_con_tildes = Persona.objects.create(
            nombres="Matías Andrés",
            apellidos="Pérez Muñoz",
            email="matias.perez@example.com",
        )
        PersonaRol.objects.create(
            persona=persona_con_tildes,
            rol=self.rol_estudiante,
            organizacion=self.org,
            activo=True,
        )
        busqueda_nombre_completo = self.client.get(
            reverse("finanzas:pago_masivo_personas"),
            {"organizacion": self.org.pk, "q": "matias perez"},
        )
        self.assertEqual(busqueda_nombre_completo.status_code, 200)
        self.assertEqual(
            [item["id"] for item in busqueda_nombre_completo.json()["resultados"]],
            [persona_con_tildes.pk],
        )
        ajena = self.client.get(
            reverse("finanzas:pago_masivo_personas"),
            {"organizacion": self.otra_org.pk, "q": "Alumno"},
        )
        self.assertEqual(ajena.status_code, 404)

    def test_preview_y_confirmacion_crean_un_solo_lote(self):
        self.client.force_login(self.user)
        payload = {
            "organizacion": self.org.pk,
            "fecha_pago": "2026-07-27",
            "plan": self.plan.pk,
            "documento_tributario": "",
            "metodo_pago": Payment.Metodo.EFECTIVO,
            "numero_comprobante": "",
            "aplica_iva": "",
            "monto_incluye_iva": "",
            "monto_referencia": "10000",
            "clases_asignadas": "0",
            "observaciones": "",
            "personas_seleccionadas": ",".join(str(persona.pk) for persona in self.personas[:2]),
            "filas_json": "{}",
            "clave_idempotencia": "vista-lote-1",
            "accion": "preview",
        }
        preview = self.client.post(reverse("finanzas:pago_masivo"), payload)
        self.assertEqual(preview.status_code, 200)
        self.assertContains(preview, "Preview del lote")
        payload["accion"] = "confirmar"
        confirmado = self.client.post(reverse("finanzas:pago_masivo"), payload)
        self.assertEqual(confirmado.status_code, 302)
        self.assertEqual(Payment.objects.count(), 2)
        self.assertEqual(LotePago.objects.count(), 1)
