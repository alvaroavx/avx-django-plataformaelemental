from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from academia.models import Disciplina, SesionClase
from asistencias.models import Asistencia
from cuentas.models import Persona
from organizaciones.models import Organizacion


class Command(BaseCommand):
    help = "Importa asistencias históricas desde planilla Excel."

    def add_arguments(self, parser):
        parser.add_argument(
            "--archivo",
            default="Asistencia Talleres Elementos.xlsx",
            help="Nombre del archivo en /data/ con los registros de asistencia.",
        )

    def handle(self, *args, **options):
        base_dir = Path(settings.BASE_DIR) / "data"
        archivo = base_dir / options["archivo"]
        if not archivo.exists():
            self.stderr.write(f"No se encontró {archivo}.")
            return
        organizacion = Organizacion.objects.first()
        if not organizacion:
            self.stderr.write("No hay organizaciones cargadas.")
            return
        wb = load_workbook(archivo)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        creadas = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, row))
            fecha = data.get("Fecha") or data.get("fecha")
            disciplina_nombre = (data.get("Disciplina") or "").strip()
            estudiante_nombre = (data.get("Estudiante") or data.get("Alumno") or "").strip()
            estado = (data.get("Estado") or "presente").lower()
            if not fecha or not disciplina_nombre or not estudiante_nombre:
                self.stdout.write(f"[Fila {idx}] Incompleta, se omite.")
                continue
            disciplina, _ = Disciplina.objects.get_or_create(
                organizacion=organizacion,
                nombre=disciplina_nombre,
            )
            nombres = estudiante_nombre.split()
            persona, _ = Persona.objects.get_or_create(
                email=f"import-{idx}@placeholder.local",
                defaults={"nombres": nombres[0], "apellidos": " ".join(nombres[1:])},
            )
            sesion, _ = SesionClase.objects.get_or_create(
                disciplina=disciplina,
                fecha=fecha,
                defaults={"cupo_maximo": 20},
            )
            Asistencia.objects.update_or_create(
                sesion=sesion,
                persona=persona,
                defaults={"estado": estado},
            )
            creadas += 1
        self.stdout.write(f"Asistencias importadas/actualizadas: {creadas}")
