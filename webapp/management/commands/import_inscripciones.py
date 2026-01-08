from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from django.utils import timezone
from openpyxl import load_workbook

from cobros.models import Plan, Suscripcion
from cuentas.models import Persona
from organizaciones.models import Organizacion


class Command(BaseCommand):
    help = "Importa inscripciones históricas desde la planilla de Google Forms."

    def add_arguments(self, parser):
        parser.add_argument(
            "--archivo",
            help="Nombre del archivo dentro de /data/. Por defecto usa el formulario estándar.",
            default="Ficha de inscripción_ Espacio Elementos. (Respuestas).xlsx",
        )

    def handle(self, *args, **options):
        base_dir = Path(settings.BASE_DIR) / "data"
        archivo = base_dir / options["archivo"]
        if not archivo.exists():
            self.stderr.write(f"No se encontró el archivo {archivo}. Colócalo dentro de /data/")
            return

        organizacion = Organizacion.objects.first()
        if not organizacion:
            self.stderr.write("No hay organizaciones configuradas. Crea al menos una.")
            return

        wb = load_workbook(archivo)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        creados = 0
        actualizados = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, row))
            nombres = (data.get("Nombres") or data.get("Nombre") or "").strip()
            apellidos = (data.get("Apellidos") or data.get("Apellido") or "").strip()
            email = (data.get("Correo electrónico") or data.get("Email") or "").strip()
            telefono = (data.get("Teléfono") or data.get("Telefono") or "").strip()
            plan_nombre = (data.get("Plan") or data.get("Plan elegido") or "").strip()
            if not nombres:
                self.stdout.write(f"[Fila {idx}] Se ignora por faltar nombres.")
                continue
            if not email:
                email = f"auto-{idx}@placeholder.local"
            persona, created = Persona.objects.get_or_create(
                email=email,
                defaults={"nombres": nombres, "apellidos": apellidos, "telefono": telefono},
            )
            if created:
                creados += 1
            else:
                actualizados += 1
                persona.nombres = persona.nombres or nombres
                persona.apellidos = persona.apellidos or apellidos
                persona.telefono = persona.telefono or telefono
                persona.save()

            if plan_nombre:
                plan = Plan.objects.filter(nombre__iexact=plan_nombre).first()
                if not plan:
                    plan = Plan.objects.create(
                        organizacion=organizacion,
                        nombre=plan_nombre,
                        precio=0,
                        duracion_dias=30,
                        clases_por_semana=1,
                    )
                fecha_inicio = data.get("Fecha de inicio") or timezone.localdate()
                Suscripcion.objects.get_or_create(
                    persona=persona,
                    plan=plan,
                    fecha_inicio=fecha_inicio,
                    defaults={"fecha_fin": fecha_inicio, "estado": Suscripcion.Estado.ACTIVA},
                )
        self.stdout.write(f"Personas creadas: {creados}, actualizadas: {actualizados}")
