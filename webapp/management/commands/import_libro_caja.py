from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from finanzas.models import MovimientoCaja
from organizaciones.models import Organizacion


class Command(BaseCommand):
    help = "Importa movimientos desde el Libro Caja histórico."

    def add_arguments(self, parser):
        parser.add_argument(
            "--archivo",
            default="Libro Caja Espacio Elementos.xlsx",
            help="Nombre del archivo dentro de /data/.",
        )

    def handle(self, *args, **options):
        base_dir = Path(settings.BASE_DIR) / "data"
        archivo = base_dir / options["archivo"]
        if not archivo.exists():
            self.stderr.write(f"No está presente {archivo}.")
            return

        organizacion = Organizacion.objects.first()
        if not organizacion:
            self.stderr.write("Crea una organización antes de importar.")
            return

        wb = load_workbook(archivo)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        creados = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, row))
            fecha = data.get("Fecha") or data.get("fecha")
            tipo = (data.get("Tipo") or "").lower()
            monto = data.get("Monto") or data.get("monto")
            categoria = (data.get("Categoria") or "otros").lower()
            glosa = data.get("Glosa") or ""
            afecta_iva = str(data.get("Afecta IVA") or "").lower() in ("si", "sí", "1", "true")
            if not fecha or not tipo or not monto:
                self.stdout.write(f"[Fila {idx}] incompleta, se omite.")
                continue
            tipo_db = MovimientoCaja.Tipo.INGRESO if "ingreso" in tipo else MovimientoCaja.Tipo.EGRESO
            categoria_db = categoria if categoria in dict(MovimientoCaja.Categoria.choices) else MovimientoCaja.Categoria.OTROS
            MovimientoCaja.objects.create(
                organizacion=organizacion,
                tipo=tipo_db,
                fecha=fecha,
                monto_total=monto,
                afecta_iva=afecta_iva,
                categoria=categoria_db,
                glosa=glosa,
            )
            creados += 1
        self.stdout.write(f"Movimientos importados: {creados}")
