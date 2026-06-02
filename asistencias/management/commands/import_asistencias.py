from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from personas.models import Organizacion, Persona

from asistencias.models import Asistencia, Disciplina, SesionClase


class Command(BaseCommand):
    help = "Importa asistencias históricas desde planilla Excel."

    def add_arguments(self, parser):
        parser.add_argument(
            "--archivo",
            default="Asistencia Talleres Elementos.xlsx",
            help="Nombre del archivo en /data/ con los registros de asistencia.",
        )
        parser.add_argument(
            "--organizacion-id",
            type=int,
            required=True,
            help="ID de la organizacion a la que se asociaran los datos importados.",
        )

    def handle(self, *args, **options):
        organizacion_id = options["organizacion_id"]
        try:
            organizacion = Organizacion.objects.get(pk=organizacion_id)
        except Organizacion.DoesNotExist as exc:
            raise CommandError(f"No existe una organizacion con ID {organizacion_id}.") from exc

        base_dir = Path(settings.BASE_DIR) / "data"
        archivo = base_dir / options["archivo"]
        if not archivo.exists():
            self.stderr.write(f"No se encontró {archivo}.")
            return
        wb = load_workbook(archivo)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        creadas = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, row))
            fecha = data.get("Fecha") or data.get("fecha")
            disciplina_nombre = (data.get("Disciplina") or "").strip()
            estudiante_nombre = (data.get("Estudiante") or data.get("Alumno") or "").strip()
            estado = (data.get("Estado") or "presente").lower()
            if not fecha or not disciplina_nombre or not estudiante_nombre:
                self.stdout.write(f"[Fila {idx}] Incompleta, se omite.")
                continue
            disciplina, _ = Disciplina.objects.get_or_create(
                organizacion=organizacion,
                nombre=disciplina_nombre,
            )
            nombres = estudiante_nombre.split()
            persona, _ = Persona.objects.get_or_create(
                email=f"import-{idx}@placeholder.local",
                defaults={"nombres": nombres[0], "apellidos": " ".join(nombres[1:])},
            )
            sesion, _ = SesionClase.objects.get_or_create(
                disciplina=disciplina,
                fecha=fecha,
                defaults={"cupo_maximo": 20},
            )
            Asistencia.objects.update_or_create(
                sesion=sesion,
                persona=persona,
                defaults={"estado": estado},
            )
            creadas += 1
        self.stdout.write(f"Asistencias importadas/actualizadas: {creadas}")
