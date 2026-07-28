import json
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO
from pathlib import Path
from tempfile import TemporaryDirectory

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.core.exceptions import ValidationError
from django.core.management.base import CommandError
from django.db import close_old_connections, connection
from django.db.models.signals import pre_save
from django.test import RequestFactory, TestCase, TransactionTestCase
from django.test import override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from auditoria.models import AuditLog
from finanzas.models import AttendanceConsumption, Payment, PaymentPlan
from finanzas.services import asignar_consumo_asistencia
from personas.models import Organizacion, Persona, PersonaRol, Rol
from personas.test_factories import asignar_profesora_a_sesion, crear_usuario_con_rol
from plataformaelemental.context import nav_context, organizacion_desde_request, periodo_context

from .models import Asistencia, BloqueHorario, ClaseLiberada, Disciplina, SesionClase
from .services import cambiar_estado_asistencia, liberar_clase, revertir_clase_liberada


TEST_PASSWORD = "not-a-real-test-password"


class ContextoGlobalTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.organizacion = Organizacion.objects.create(
            nombre="Org Contexto",
            razon_social="Org Contexto SPA",
            rut="22.222.222-2",
        )

    def test_periodo_context_usa_fecha_actual_y_todas_las_organizaciones_por_defecto(self):
        request = self.factory.get("/")
        contexto = periodo_context(request)
        hoy = timezone.localdate()

        self.assertEqual(contexto["periodo_mes"], str(hoy.month))
        self.assertEqual(contexto["periodo_anio"], str(hoy.year))
        self.assertEqual(contexto["organizacion_id"], "")
        self.assertIn(("todos", "Todos"), contexto["periodo_meses"])
        self.assertIn(self.organizacion, list(contexto["organizaciones_global"]))

    def test_organizacion_desde_request_respeta_filtro_global(self):
        request = self.factory.get("/", {"organizacion": self.organizacion.pk})

        self.assertEqual(organizacion_desde_request(request), self.organizacion)
        self.assertIsNone(organizacion_desde_request(self.factory.get("/")))
        self.assertIsNone(organizacion_desde_request(self.factory.get("/", {"organizacion": "99999"})))

    def test_nav_context_expone_persona_y_roles_activos(self):
        User = get_user_model()
        user = User.objects.create_user(username="contexto", password=TEST_PASSWORD)
        persona = Persona.objects.create(nombres="Clara", apellidos="Contexto", user=user)
        rol_activo = Rol.objects.create(nombre="Administrador", codigo="ADMIN")
        rol_inactivo = Rol.objects.create(nombre="Profesor", codigo="PROFESOR")
        PersonaRol.objects.create(persona=persona, rol=rol_activo, organizacion=self.organizacion, activo=True)
        PersonaRol.objects.create(persona=persona, rol=rol_inactivo, organizacion=self.organizacion, activo=False)
        request = self.factory.get("/")
        request.user = user

        contexto = nav_context(request)

        self.assertEqual(contexto["persona"], persona)
        self.assertEqual(contexto["roles_usuario"], ["ADMIN"])


class ImportAsistenciasCommandTests(TestCase):
    def setUp(self):
        self.organizacion = Organizacion.objects.create(
            nombre="Org Import",
            razon_social="Org Import SPA",
            rut="33.333.333-3",
        )

    def test_import_asistencias_exige_organizacion_explicita(self):
        with self.assertRaises(CommandError):
            call_command("import_asistencias")

    def test_import_asistencias_con_organizacion_invalida_falla_claro(self):
        with self.assertRaisesMessage(CommandError, "No existe una organizacion con ID 999999."):
            call_command("import_asistencias", organizacion_id=999999)

    def test_import_asistencias_usa_organizacion_explicita(self):
        with TemporaryDirectory() as tmp_dir:
            data_dir = Path(tmp_dir) / "data"
            data_dir.mkdir()
            archivo = data_dir / "asistencias.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Fecha", "Disciplina", "Estudiante", "Estado"])
            sheet.append([date(2026, 5, 4), "Yoga", "Ana Diaz", "presente"])
            workbook.save(archivo)

            with override_settings(BASE_DIR=Path(tmp_dir)):
                salida = StringIO()
                call_command(
                    "import_asistencias",
                    archivo="asistencias.xlsx",
                    organizacion_id=self.organizacion.pk,
                    stdout=salida,
                )

        self.assertIn("Asistencias importadas/actualizadas: 1", salida.getvalue())
        disciplina = Disciplina.objects.get(nombre="Yoga")
        self.assertEqual(disciplina.organizacion, self.organizacion)
        self.assertEqual(Asistencia.objects.count(), 1)


class AsistenciasViewTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username="admin",
            password=TEST_PASSWORD,
            is_superuser=True,
            is_staff=True,
        )
        self.client.force_login(self.user)

        self.organizacion = Organizacion.objects.create(
            nombre="Org Test",
            razon_social="Org Test SPA",
            rut="11.111.111-1",
        )
        self.disciplina = Disciplina.objects.create(
            organizacion=self.organizacion,
            nombre="Flexibilidad",
        )
        self.sesion = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-02-26",
            estado=SesionClase.Estado.PROGRAMADA,
        )
        self.estudiante = Persona.objects.create(
            nombres="Ana",
            apellidos="Diaz",
            email="ana@example.com",
        )
        self.rol_estudiante = Rol.objects.create(nombre="Estudiante", codigo="ESTUDIANTE")
        PersonaRol.objects.create(
            persona=self.estudiante,
            rol=self.rol_estudiante,
            organizacion=self.organizacion,
            activo=True,
        )

    def _xlsx_rows(self, response):
        workbook = load_workbook(BytesIO(response.content))
        return list(workbook.active.iter_rows(values_only=True))

    def _login_admin_organizacion(self, organizacion, username="admin_org"):
        User = get_user_model()
        user = User.objects.create_user(username=username, password=TEST_PASSWORD)
        persona = Persona.objects.create(nombres="Admin", apellidos=username, user=user)
        rol_admin, _ = Rol.objects.get_or_create(nombre="Administrador", codigo="ADMIN")
        PersonaRol.objects.create(
            persona=persona,
            rol=rol_admin,
            organizacion=organizacion,
            activo=True,
        )
        self.client.force_login(user)
        return user

    def test_export_asistencias_xlsx_respeta_periodo_y_organizacion(self):
        otra_org = Organizacion.objects.create(
            nombre="Org Export",
            razon_social="Org Export SPA",
            rut="77.777.777-7",
        )
        otra_disciplina = Disciplina.objects.create(organizacion=otra_org, nombre="Otra disciplina")
        sesion_otro_periodo = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-03-02",
            estado=SesionClase.Estado.COMPLETADA,
        )
        sesion_otra_org = SesionClase.objects.create(
            disciplina=otra_disciplina,
            fecha="2026-02-02",
            estado=SesionClase.Estado.COMPLETADA,
        )
        Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
            estado=Asistencia.Estado.PRESENTE,
            comentario="Incluida",
        )
        Asistencia.objects.create(
            sesion=sesion_otro_periodo,
            persona=self.estudiante,
            estado=Asistencia.Estado.PRESENTE,
            comentario="Fuera periodo",
        )
        Asistencia.objects.create(
            sesion=sesion_otra_org,
            persona=self.estudiante,
            estado=Asistencia.Estado.PRESENTE,
            comentario="Fuera organizacion",
        )

        response = self.client.get(
            reverse("asistencias:export_asistencias_xlsx"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        rows = self._xlsx_rows(response)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(rows[0][0], "Fecha sesion")
        self.assertIn("Periodo", rows[0])
        contenido = str(rows)
        self.assertIn("Incluida", contenido)
        self.assertNotIn("Fuera periodo", contenido)
        self.assertNotIn("Fuera organizacion", contenido)

    def test_export_asistencias_xlsx_bloquea_usuario_sin_permiso(self):
        User = get_user_model()
        usuario = User.objects.create_user("sin_permiso_export", password=TEST_PASSWORD)
        self.client.force_login(usuario)

        response = self.client.get(
            reverse("asistencias:export_asistencias_xlsx"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 403)

    def test_agregar_asistentes_cambia_estado_a_completada(self):
        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                f"{reverse('asistencias:asistencias_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}",
                {
                    "agregar_asistentes": "1",
                    "sesion_id": str(self.sesion.pk),
                    "estudiantes": [str(self.estudiante.pk)],
                },
            )

        self.assertEqual(response.status_code, 302)
        self.sesion.refresh_from_db()
        self.assertEqual(self.sesion.estado, SesionClase.Estado.COMPLETADA)
        log = AuditLog.objects.filter(
            dominio="asistencias",
            accion=AuditLog.ACCION_AGREGAR_ASISTENTES,
            objeto_id=str(self.sesion.pk),
        ).latest("fecha")
        self.assertEqual(log.metadata["asistencias_creadas"], 1)
        self.assertEqual(log.metadata["persona_ids"], [self.estudiante.pk])

    def test_crear_sesion_genera_auditlog(self):
        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                f"{reverse('asistencias:asistencias_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}",
                {
                    "crear_sesion": "1",
                    "disciplina": self.disciplina.pk,
                    "fecha": "2026-02-27",
                    "profesores": [],
                },
            )

        self.assertEqual(response.status_code, 302)
        sesion = SesionClase.objects.get(fecha="2026-02-27", disciplina=self.disciplina)
        self.assertTrue(
            AuditLog.objects.filter(
                dominio="asistencias",
                accion=AuditLog.ACCION_CREAR,
                objeto_id=str(sesion.pk),
            ).exists()
        )

    def test_agregar_asistentes_guardar_y_cerrar_cierra_modal(self):
        response = self.client.post(
            f"{reverse('asistencias:asistencias_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}",
            {
                "agregar_asistentes": "1",
                "accion_guardado_asistencias": "cerrar",
                "sesion_id": str(self.sesion.pk),
                "estudiantes": [str(self.estudiante.pk)],
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            f"{reverse('asistencias:asistencias_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}&sesion_id={self.sesion.pk}",
        )

    def test_agregar_asistentes_guardar_y_cerrar_elimina_flag_open_del_modal(self):
        response = self.client.post(
            (
                f"{reverse('asistencias:asistencias_list')}"
                f"?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}&open=agregar_asistentes"
            ),
            {
                "agregar_asistentes": "1",
                "accion_guardado_asistencias": "cerrar",
                "sesion_id": str(self.sesion.pk),
                "estudiantes": [str(self.estudiante.pk)],
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            f"{reverse('asistencias:asistencias_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}&sesion_id={self.sesion.pk}",
        )

    def test_agregar_asistentes_guardar_y_agregar_otro_reabre_modal(self):
        response = self.client.post(
            f"{reverse('asistencias:asistencias_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}",
            {
                "agregar_asistentes": "1",
                "accion_guardado_asistencias": "continuar",
                "sesion_id": str(self.sesion.pk),
                "estudiantes": [str(self.estudiante.pk)],
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            f"{reverse('asistencias:asistencias_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}&sesion_id={self.sesion.pk}&open=agregar_asistentes",
        )

    def test_agregar_asistentes_lista_solo_estudiantes_de_la_organizacion_de_la_sesion(self):
        otra_organizacion = Organizacion.objects.create(
            nombre="Otra Org",
            razon_social="Otra Org SPA",
            rut="22.222.222-2",
        )
        estudiante_otra_org = Persona.objects.create(
            nombres="Bruno",
            apellidos="Otraorg",
            email="bruno.otraorg@example.com",
        )
        rol_estudiante = Rol.objects.get(codigo="ESTUDIANTE")
        PersonaRol.objects.create(
            persona=estudiante_otra_org,
            rol=rol_estudiante,
            organizacion=otra_organizacion,
            activo=True,
        )

        response = self.client.get(
            reverse("asistencias:asistencias_list"),
            {
                "periodo_mes": 2,
                "periodo_anio": 2026,
                "organizacion": self.organizacion.pk,
                "sesion_id": self.sesion.pk,
                "open": "agregar_asistentes",
            },
        )

        estudiantes = list(response.context["estudiantes"])
        self.assertIn(self.estudiante, estudiantes)
        self.assertNotIn(estudiante_otra_org, estudiantes)
        self.assertContains(response, self.estudiante.nombre_completo)
        self.assertNotContains(response, estudiante_otra_org.nombre_completo)

    def test_agregar_asistentes_preselecciona_sesion_en_modal(self):
        response = self.client.get(
            reverse("asistencias:asistencias_list"),
            {
                "periodo_mes": 2,
                "periodo_anio": 2026,
                "organizacion": self.organizacion.pk,
                "sesion_id": self.sesion.pk,
                "open": "agregar_asistentes",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sesión de asistencia")
        self.assertContains(response, f'value="{self.sesion.pk}" selected', html=False)

    def test_agregar_asistentes_selector_lista_solo_sesiones_de_periodo_y_organizacion(self):
        otra_organizacion = Organizacion.objects.create(
            nombre="Otra Org",
            razon_social="Otra Org SPA",
            rut="22.222.222-2",
        )
        otra_disciplina = Disciplina.objects.create(organizacion=otra_organizacion, nombre="Otra disciplina")
        sesion_otro_periodo = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-03-05",
        )
        sesion_otra_org = SesionClase.objects.create(
            disciplina=otra_disciplina,
            fecha="2026-02-05",
        )

        response = self.client.get(
            reverse("asistencias:asistencias_list"),
            {
                "periodo_mes": 2,
                "periodo_anio": 2026,
                "organizacion": self.organizacion.pk,
                "open": "agregar_asistentes",
            },
        )

        sesiones = list(response.context["sesiones_agregar_asistentes"])
        self.assertIn(self.sesion, sesiones)
        self.assertNotIn(sesion_otro_periodo, sesiones)
        self.assertNotIn(sesion_otra_org, sesiones)

    def test_agregar_asistentes_rechaza_sesion_de_otra_organizacion(self):
        otra_organizacion = Organizacion.objects.create(
            nombre="Otra Org",
            razon_social="Otra Org SPA",
            rut="22.222.222-2",
        )
        otra_disciplina = Disciplina.objects.create(organizacion=otra_organizacion, nombre="Otra disciplina")
        sesion_otra_org = SesionClase.objects.create(
            disciplina=otra_disciplina,
            fecha="2026-02-05",
        )

        response = self.client.post(
            f"{reverse('asistencias:asistencias_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}",
            {
                "agregar_asistentes": "1",
                "sesion_id": str(sesion_otra_org.pk),
                "estudiantes": [str(self.estudiante.pk)],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(
            Asistencia.objects.filter(sesion=sesion_otra_org, persona=self.estudiante).exists()
        )

    def test_agregar_asistentes_usa_sesion_seleccionada(self):
        segunda_sesion = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-02-28",
        )

        response = self.client.post(
            f"{reverse('asistencias:asistencias_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}",
            {
                "agregar_asistentes": "1",
                "sesion_id": str(segunda_sesion.pk),
                "estudiantes": [str(self.estudiante.pk)],
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            Asistencia.objects.filter(sesion=segunda_sesion, persona=self.estudiante).exists()
        )
        self.assertFalse(
            Asistencia.objects.filter(sesion=self.sesion, persona=self.estudiante).exists()
        )

    def test_agregar_asistentes_rechaza_estudiante_de_otra_organizacion(self):
        otra_organizacion = Organizacion.objects.create(
            nombre="Otra Org",
            razon_social="Otra Org SPA",
            rut="22.222.222-2",
        )
        estudiante_otra_org = Persona.objects.create(
            nombres="Bruno",
            apellidos="Otraorg",
            email="bruno.rechazo@example.com",
        )
        rol_estudiante = Rol.objects.get(codigo="ESTUDIANTE")
        PersonaRol.objects.create(
            persona=estudiante_otra_org,
            rol=rol_estudiante,
            organizacion=otra_organizacion,
            activo=True,
        )

        response = self.client.post(
            f"{reverse('asistencias:asistencias_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}",
            {
                "agregar_asistentes": "1",
                "sesion_id": str(self.sesion.pk),
                "estudiantes": [str(estudiante_otra_org.pk)],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(
            Asistencia.objects.filter(sesion=self.sesion, persona=estudiante_otra_org).exists()
        )

    def test_agregar_asistentes_muestra_inactivos_y_los_reactiva_al_agregar(self):
        self.estudiante.activo = False
        self.estudiante.save(update_fields=["activo"])
        persona_rol = PersonaRol.objects.get(
            persona=self.estudiante,
            organizacion=self.organizacion,
            rol__codigo="ESTUDIANTE",
        )
        persona_rol.activo = False
        persona_rol.save(update_fields=["activo"])

        response = self.client.get(
            reverse("asistencias:asistencias_list"),
            {
                "periodo_mes": 2,
                "periodo_anio": 2026,
                "organizacion": self.organizacion.pk,
                "sesion_id": self.sesion.pk,
                "open": "agregar_asistentes",
            },
        )

        self.assertContains(response, self.estudiante.nombre_completo)
        self.assertContains(response, "Inactivo")

        response = self.client.post(
            f"{reverse('asistencias:asistencias_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}",
            {
                "agregar_asistentes": "1",
                "sesion_id": str(self.sesion.pk),
                "estudiantes": [str(self.estudiante.pk)],
            },
        )

        self.assertEqual(response.status_code, 302)
        self.estudiante.refresh_from_db()
        persona_rol.refresh_from_db()
        self.assertTrue(self.estudiante.activo)
        self.assertTrue(persona_rol.activo)
        self.assertTrue(
            Asistencia.objects.filter(sesion=self.sesion, persona=self.estudiante).exists()
        )

    def test_agregar_persona_desde_asistencias_usa_organizacion_filtrada(self):
        response = self.client.post(
            f"{reverse('asistencias:asistencias_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}",
            {
                "agregar_persona": "1",
                "nombres": "Nueva",
                "apellidos": "Persona",
                "telefono": "123456",
            },
        )

        self.assertEqual(response.status_code, 302)
        persona = Persona.objects.get(nombres="Nueva", apellidos="Persona")
        self.assertTrue(
            PersonaRol.objects.filter(
                persona=persona,
                organizacion=self.organizacion,
                rol__codigo="ESTUDIANTE",
            ).exists()
        )

    def test_agregar_persona_desde_asistencias_exige_organizacion_filtrada(self):
        response = self.client.post(
            reverse("asistencias:asistencias_list"),
            {
                "agregar_persona": "1",
                "nombres": "Sin",
                "apellidos": "Organizacion",
                "telefono": "123456",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "Debes seleccionar una organización en el filtro superior antes de crear a la persona.",
        )
        self.assertFalse(Persona.objects.filter(nombres="Sin", apellidos="Organizacion").exists())

    def test_asistencias_cambiar_estado_mantiene_filtros_en_redirect(self):
        url = f"{reverse('asistencias:asistencias_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}"
        response = self.client.post(
            url,
            {
                "cambiar_estado": "1",
                "sesion_id": str(self.sesion.pk),
                "estado": SesionClase.Estado.CANCELADA,
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, url)
        self.sesion.refresh_from_db()
        self.assertEqual(self.sesion.estado, SesionClase.Estado.CANCELADA)

    def test_nueva_sesion_muestra_solo_disciplinas_y_profesores_activos_de_la_organizacion(self):
        otra_organizacion = Organizacion.objects.create(
            nombre="Org Secundaria",
            razon_social="Org Secundaria SPA",
            rut="22.222.222-2",
        )
        disciplina_inactiva = Disciplina.objects.create(
            organizacion=self.organizacion,
            nombre="Pilates",
            activa=False,
        )
        disciplina_otra_org = Disciplina.objects.create(
            organizacion=otra_organizacion,
            nombre="Teatro",
            activa=True,
        )
        profesor_activo = Persona.objects.create(
            nombres="Paula",
            apellidos="Activa",
            email="paula.activa@example.com",
            activo=True,
        )
        profesor_inactivo = Persona.objects.create(
            nombres="Bruno",
            apellidos="Inactivo",
            email="bruno.inactivo@example.com",
            activo=False,
        )
        profesor_otra_org = Persona.objects.create(
            nombres="Marta",
            apellidos="Otraorg",
            email="marta.otraorg@example.com",
            activo=True,
        )
        rol_profesor = Rol.objects.create(nombre="Profesor", codigo="PROFESOR")
        PersonaRol.objects.create(
            persona=profesor_activo,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
        )
        PersonaRol.objects.create(
            persona=profesor_inactivo,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
        )
        PersonaRol.objects.create(
            persona=profesor_otra_org,
            rol=rol_profesor,
            organizacion=otra_organizacion,
            activo=True,
        )

        response = self.client.get(
            reverse("asistencias:asistencias_list"),
            {
                "periodo_mes": 2,
                "periodo_anio": 2026,
                "organizacion": self.organizacion.pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        disciplinas = list(response.context["sesion_form"].fields["disciplina"].queryset)
        profesores = list(response.context["sesion_form"].fields["profesores"].queryset)
        self.assertIn(self.disciplina, disciplinas)
        self.assertNotIn(disciplina_inactiva, disciplinas)
        self.assertNotIn(disciplina_otra_org, disciplinas)
        self.assertIn(profesor_activo, profesores)
        self.assertNotIn(profesor_inactivo, profesores)
        self.assertNotIn(profesor_otra_org, profesores)

    def test_asistencias_list_filtros_locales_muestran_solo_disciplinas_y_profesores_vigentes(self):
        otra_organizacion = Organizacion.objects.create(
            nombre="Org Secundaria",
            razon_social="Org Secundaria SPA",
            rut="22.222.222-2",
        )
        disciplina_inactiva = Disciplina.objects.create(
            organizacion=self.organizacion,
            nombre="Pilates",
            activa=False,
        )
        disciplina_otra_org = Disciplina.objects.create(
            organizacion=otra_organizacion,
            nombre="Teatro",
            activa=True,
        )
        profesor_activo = Persona.objects.create(
            nombres="Paula",
            apellidos="Activa",
            email="paula.filtro@example.com",
            activo=True,
        )
        profesor_inactivo = Persona.objects.create(
            nombres="Bruno",
            apellidos="Inactivo",
            email="bruno.filtro@example.com",
            activo=False,
        )
        profesor_otra_org = Persona.objects.create(
            nombres="Marta",
            apellidos="Otraorg",
            email="marta.filtro@example.com",
            activo=True,
        )
        rol_profesor = Rol.objects.get_or_create(nombre="Profesor", codigo="PROFESOR")[0]
        PersonaRol.objects.create(
            persona=profesor_activo,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
        )
        PersonaRol.objects.create(
            persona=profesor_inactivo,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
        )
        PersonaRol.objects.create(
            persona=profesor_otra_org,
            rol=rol_profesor,
            organizacion=otra_organizacion,
            activo=True,
        )

        response = self.client.get(
            reverse("asistencias:asistencias_list"),
            {
                "periodo_mes": 2,
                "periodo_anio": 2026,
                "organizacion": self.organizacion.pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        disciplinas = list(response.context["disciplinas"])
        profesores = list(response.context["profesores"])
        self.assertIn(self.disciplina, disciplinas)
        self.assertNotIn(disciplina_inactiva, disciplinas)
        self.assertNotIn(disciplina_otra_org, disciplinas)
        self.assertIn(profesor_activo, profesores)
        self.assertNotIn(profesor_inactivo, profesores)
        self.assertNotIn(profesor_otra_org, profesores)

    def test_agregar_asistentes_desde_sesion_detail(self):
        url = reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk})
        response = self.client.post(
            url,
            {
                "agregar_asistentes": "1",
                "estudiantes": [str(self.estudiante.pk)],
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, url)
        self.assertTrue(
            Asistencia.objects.filter(sesion=self.sesion, persona=self.estudiante).exists()
        )
        self.sesion.refresh_from_db()
        self.assertEqual(self.sesion.estado, SesionClase.Estado.COMPLETADA)

    def test_buscar_asistentes_mobile_misma_organizacion_excluye_agregados_y_limita_datos(self):
        self._login_admin_organizacion(self.organizacion)
        Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        for indice in range(12):
            persona = Persona.objects.create(
                nombres=f"Ana {indice}",
                apellidos="Mobile",
                email=f"ana.mobile.{indice}@example.com",
            )
            PersonaRol.objects.create(
                persona=persona,
                rol=self.rol_estudiante,
                organizacion=self.organizacion,
                activo=True,
            )
        otra_organizacion = Organizacion.objects.create(
            nombre="Otra Mobile",
            razon_social="Otra Mobile SPA",
            rut="22.222.222-2",
        )
        persona_otra_org = Persona.objects.create(
            nombres="Ana",
            apellidos="Otra",
            email="ana.otra@example.com",
        )
        PersonaRol.objects.create(
            persona=persona_otra_org,
            rol=self.rol_estudiante,
            organizacion=otra_organizacion,
            activo=True,
        )

        response = self.client.get(
            reverse("asistencias:sesion_asistentes_buscar", kwargs={"pk": self.sesion.pk}),
            {"q": "Ana", "organizacion": otra_organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["ok"])
        self.assertLessEqual(len(data["resultados"]), 10)
        ids = {item["id"] for item in data["resultados"]}
        self.assertNotIn(self.estudiante.pk, ids)
        self.assertNotIn(persona_otra_org.pk, ids)
        self.assertEqual(set(data["resultados"][0].keys()), {"id", "nombre", "inactivo"})

    def test_buscar_asistentes_mobile_busca_por_email_y_rut_sin_exponerlos(self):
        self._login_admin_organizacion(self.organizacion)
        persona = Persona.objects.create(
            nombres="Beatriz",
            apellidos="Rut",
            email="bea.buscar@example.com",
            rut="12.345.678-5",
        )
        PersonaRol.objects.create(
            persona=persona,
            rol=self.rol_estudiante,
            organizacion=self.organizacion,
            activo=True,
        )

        response_email = self.client.get(
            reverse("asistencias:sesion_asistentes_buscar", kwargs={"pk": self.sesion.pk}),
            {"q": "bea.buscar"},
        )
        response_rut = self.client.get(
            reverse("asistencias:sesion_asistentes_buscar", kwargs={"pk": self.sesion.pk}),
            {"q": "12.345"},
        )

        self.assertEqual(response_email.status_code, 200)
        self.assertEqual(response_rut.status_code, 200)
        self.assertEqual(response_email.json()["resultados"][0]["id"], persona.pk)
        self.assertEqual(response_rut.json()["resultados"][0]["id"], persona.pk)
        self.assertNotIn("email", response_email.json()["resultados"][0])
        self.assertNotIn("rut", response_email.json()["resultados"][0])

    def test_agregar_asistente_mobile_crea_asistencia_y_consumo_financiero(self):
        self._login_admin_organizacion(self.organizacion)
        Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.organizacion,
            fecha_pago=date(2026, 2, 1),
            monto_referencia=Decimal("10000"),
            clases_asignadas=1,
        )

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": self.sesion.pk}),
                {"persona_id": self.estudiante.pk},
            )

        self.assertEqual(response.status_code, 201)
        data = response.json()
        self.assertTrue(data["ok"])
        asistencia = Asistencia.objects.get(sesion=self.sesion, persona=self.estudiante)
        self.assertEqual(data["asistencia"]["id"], asistencia.pk)
        self.assertEqual(data["asistencia"]["estado"], Asistencia.Estado.PRESENTE)
        self.assertEqual(data["asistencia"]["estado_label"], asistencia.get_estado_display())
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.CONSUMIDO)
        self.assertIsNotNone(consumo.pago_id)
        self.assertEqual(data["estado_financiero"]["codigo"], "consumido")
        self.assertEqual(data["estado_financiero"]["label"], "Pagada")
        self.sesion.refresh_from_db()
        self.assertEqual(self.sesion.estado, SesionClase.Estado.COMPLETADA)
        self.assertTrue(
            AuditLog.objects.filter(
                dominio="asistencias",
                accion=AuditLog.ACCION_AGREGAR_ASISTENTES,
                metadata__origen="sesion_detail_mobile",
            ).exists()
        )

    def test_agregar_asistente_mobile_acepta_cuerpo_json(self):
        self._login_admin_organizacion(self.organizacion)

        response = self.client.post(
            reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": self.sesion.pk}),
            data=json.dumps({"persona_id": self.estudiante.pk}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertTrue(response.json()["ok"])
        self.assertTrue(Asistencia.objects.filter(sesion=self.sesion, persona=self.estudiante).exists())

    def test_agregar_asistente_mobile_rechaza_persona_otra_organizacion(self):
        self._login_admin_organizacion(self.organizacion)
        otra_organizacion = Organizacion.objects.create(
            nombre="Org Ajena",
            razon_social="Org Ajena SPA",
            rut="23.333.333-3",
        )
        persona_otra_org = Persona.objects.create(
            nombres="Persona",
            apellidos="Ajena",
            email="ajena@example.com",
        )
        PersonaRol.objects.create(
            persona=persona_otra_org,
            rol=self.rol_estudiante,
            organizacion=otra_organizacion,
            activo=True,
        )

        response = self.client.post(
            reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": self.sesion.pk}),
            {"persona_id": persona_otra_org.pk},
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["codigo"], "PERSONA_INVALIDA")
        self.assertFalse(Asistencia.objects.filter(sesion=self.sesion, persona=persona_otra_org).exists())

    def test_agregar_asistente_mobile_rechaza_persona_ya_agregada(self):
        self._login_admin_organizacion(self.organizacion)
        Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)

        response = self.client.post(
            reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": self.sesion.pk}),
            {"persona_id": self.estudiante.pk},
        )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["codigo"], "ASISTENCIA_DUPLICADA")
        self.assertEqual(Asistencia.objects.filter(sesion=self.sesion, persona=self.estudiante).count(), 1)

    def test_agregar_asistente_mobile_doble_post_no_duplica(self):
        self._login_admin_organizacion(self.organizacion)
        url = reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": self.sesion.pk})

        primera = self.client.post(url, {"persona_id": self.estudiante.pk})
        segunda = self.client.post(url, {"persona_id": self.estudiante.pk})

        self.assertEqual(primera.status_code, 201)
        self.assertEqual(segunda.status_code, 409)
        self.assertEqual(Asistencia.objects.filter(sesion=self.sesion, persona=self.estudiante).count(), 1)
        self.assertEqual(
            AttendanceConsumption.objects.filter(asistencia__sesion=self.sesion, persona=self.estudiante).count(),
            1,
        )

    def test_agregar_asistente_mobile_admin_otra_organizacion_recibe_404(self):
        otra_organizacion = Organizacion.objects.create(
            nombre="Org Admin Ajeno",
            razon_social="Org Admin Ajeno SPA",
            rut="24.444.444-4",
        )
        self._login_admin_organizacion(otra_organizacion, username="admin_otra_org")

        response = self.client.post(
            reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": self.sesion.pk}),
            {"persona_id": self.estudiante.pk},
        )

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json()["codigo"], "SESION_NO_ENCONTRADA")
        self.assertFalse(Asistencia.objects.filter(sesion=self.sesion, persona=self.estudiante).exists())

    def test_agregar_asistente_mobile_usuario_sin_permiso_no_puede(self):
        User = get_user_model()
        usuario = User.objects.create_user("sin_permiso_mobile", password=TEST_PASSWORD)
        self.client.force_login(usuario)

        response = self.client.post(
            reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": self.sesion.pk}),
            {"persona_id": self.estudiante.pk},
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()["codigo"], "PERMISO_DENEGADO")
        self.assertFalse(Asistencia.objects.filter(sesion=self.sesion, persona=self.estudiante).exists())

    def test_agregar_asistente_mobile_sesion_inexistente(self):
        self._login_admin_organizacion(self.organizacion)

        response = self.client.post(
            reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": 999999}),
            {"persona_id": self.estudiante.pk},
        )

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json()["codigo"], "SESION_NO_ENCONTRADA")

    def test_agregar_asistente_mobile_persona_invalida(self):
        self._login_admin_organizacion(self.organizacion)
        persona_sin_rol = Persona.objects.create(
            nombres="Sin",
            apellidos="Rol",
            email="sin.rol@example.com",
        )

        response = self.client.post(
            reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": self.sesion.pk}),
            {"persona_id": persona_sin_rol.pk},
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["codigo"], "PERSONA_INVALIDA")
        self.assertFalse(Asistencia.objects.filter(sesion=self.sesion, persona=persona_sin_rol).exists())

    def test_agregar_asistente_mobile_persona_id_malformado(self):
        self._login_admin_organizacion(self.organizacion)

        response = self.client.post(
            reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": self.sesion.pk}),
            {"persona_id": "abc"},
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["codigo"], "PERSONA_INVALIDA")
        self.assertFalse(Asistencia.objects.filter(sesion=self.sesion).exists())

    def test_agregar_asistente_mobile_asistencia_existente_no_toca_consumo(self):
        self._login_admin_organizacion(self.organizacion)
        asistencia = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)

        response = self.client.post(
            reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": self.sesion.pk}),
            {"persona_id": self.estudiante.pk},
        )

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["codigo"], "ASISTENCIA_DUPLICADA")
        self.assertEqual(Asistencia.objects.filter(sesion=self.sesion, persona=self.estudiante).count(), 1)
        self.assertEqual(AttendanceConsumption.objects.get(asistencia=asistencia).pk, consumo.pk)

    def test_buscar_asistentes_mobile_termino_corto_retorna_vacio(self):
        self._login_admin_organizacion(self.organizacion)

        for termino in ["", "a", "A"]:
            with self.subTest(termino=termino):
                response = self.client.get(
                    reverse("asistencias:sesion_asistentes_buscar", kwargs={"pk": self.sesion.pk}),
                    {"q": termino},
                )
                self.assertEqual(response.status_code, 200)
                data = response.json()
                self.assertTrue(data["ok"])
                self.assertEqual(data["resultados"], [])

    def test_buscar_asistentes_mobile_admin_otra_organizacion_recibe_404(self):
        otra_organizacion = Organizacion.objects.create(
            nombre="Org Buscar Ajena",
            razon_social="Org Buscar Ajena SPA",
            rut="25.555.555-5",
        )
        self._login_admin_organizacion(otra_organizacion, username="admin_buscar_ajena")

        response = self.client.get(
            reverse("asistencias:sesion_asistentes_buscar", kwargs={"pk": self.sesion.pk}),
            {"q": "Ana"},
        )

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json()["codigo"], "SESION_NO_ENCONTRADA")

    def test_buscar_asistentes_mobile_usuario_sin_permiso_no_puede(self):
        User = get_user_model()
        usuario = User.objects.create_user("sin_permiso_buscar", password=TEST_PASSWORD)
        self.client.force_login(usuario)

        response = self.client.get(
            reverse("asistencias:sesion_asistentes_buscar", kwargs={"pk": self.sesion.pk}),
            {"q": "Ana"},
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()["codigo"], "PERMISO_DENEGADO")

    def test_agregar_asistente_mobile_crea_deuda_sin_pago_disponible(self):
        self._login_admin_organizacion(self.organizacion)

        response = self.client.post(
            reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": self.sesion.pk}),
            {"persona_id": self.estudiante.pk},
        )

        self.assertEqual(response.status_code, 201)
        data = response.json()
        asistencia = Asistencia.objects.get(sesion=self.sesion, persona=self.estudiante)
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.DEUDA)
        self.assertIsNone(consumo.pago)
        self.assertEqual(data["estado_financiero"]["codigo"], "deuda")
        self.assertEqual(data["estado_financiero"]["label"], "Deuda")

    def test_agregar_asistente_mobile_ausente_sin_pago_queda_deuda(self):
        self._login_admin_organizacion(self.organizacion)
        dispatch_uid = "test_asistencia_mobile_inicia_ausente"

        def iniciar_asistencia_ausente(sender, instance, **kwargs):
            if instance.sesion_id == self.sesion.pk and instance.persona_id == self.estudiante.pk:
                instance.estado = Asistencia.Estado.AUSENTE

        pre_save.connect(
            iniciar_asistencia_ausente,
            sender=Asistencia,
            dispatch_uid=dispatch_uid,
            weak=False,
        )
        try:
            response = self.client.post(
                reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": self.sesion.pk}),
                {"persona_id": self.estudiante.pk},
            )
        finally:
            pre_save.disconnect(sender=Asistencia, dispatch_uid=dispatch_uid)

        self.assertEqual(response.status_code, 201)
        asistencia = Asistencia.objects.get(sesion=self.sesion, persona=self.estudiante)
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(asistencia.estado, Asistencia.Estado.AUSENTE)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.DEUDA)
        self.assertEqual(
            response.json()["estado_financiero"],
            {
                "codigo": "deuda",
                "label": "Deuda",
            },
        )

    def test_sesion_detail_fetch_distingue_error_de_red(self):
        response = self.client.get(reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk}))

        self.assertContains(response, "return fetch(url, options).then")
        self.assertContains(response, ".catch(function () {")
        self.assertContains(response, "No se pudo conectar. Revisa tu conexión e intenta nuevamente.")

    def test_sesion_detail_fetch_distingue_error_http_con_json_valido(self):
        response = self.client.get(reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk}))

        self.assertContains(
            response,
            "return { status: response.status, json: JSON.parse(texto), tipo: 'http_json' };",
        )
        self.assertContains(response, "(data && data.mensaje) || 'Error al agregar. Intenta de nuevo.'")

    def test_sesion_detail_fetch_distingue_error_http_sin_json(self):
        response = self.client.get(reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk}))

        self.assertContains(response, "if (!texto.trim())")
        self.assertContains(
            response,
            "return { status: response.status, json: null, tipo: 'http_sin_json' };",
        )
        self.assertContains(response, "Ocurrió un error del servidor. Intenta nuevamente.")

    def test_sesion_detail_fetch_distingue_json_invalido(self):
        response = self.client.get(reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk}))

        self.assertContains(
            response,
            "return { status: response.status, json: null, tipo: 'json_invalido' };",
        )
        self.assertContains(response, "if (resp.tipo === 'json_invalido')")
        self.assertContains(response, "Respuesta inválida del servidor. Intenta nuevamente.")

    def test_agregar_asistente_mobile_crea_exactamente_un_consumo(self):
        """El post_save signal crea el consumo; no debe haber duplicados."""
        self._login_admin_organizacion(self.organizacion)

        response = self.client.post(
            reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": self.sesion.pk}),
            {"persona_id": self.estudiante.pk},
        )

        self.assertEqual(response.status_code, 201)
        asistencia = Asistencia.objects.get(sesion=self.sesion, persona=self.estudiante)
        self.assertEqual(AttendanceConsumption.objects.filter(asistencia=asistencia).count(), 1)

    def test_agregar_asistente_mobile_sesion_ajena_e_inexistente_son_indistinguibles(self):
        """Un admin de otra org y una sesión inexistente deben recibir el mismo 404."""
        otra_organizacion = Organizacion.objects.create(
            nombre="Org Indistinguible Agregar",
            razon_social="Org Indistinguible Agregar SPA",
            rut="26.666.666-6",
        )
        self._login_admin_organizacion(otra_organizacion, username="admin_indistinguible_agregar")

        response_ajena = self.client.post(
            reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": self.sesion.pk}),
            {"persona_id": self.estudiante.pk},
        )
        response_inexistente = self.client.post(
            reverse("asistencias:sesion_asistente_agregar", kwargs={"pk": 999998}),
            {"persona_id": self.estudiante.pk},
        )

        self.assertEqual(response_ajena.status_code, 404)
        self.assertEqual(response_inexistente.status_code, 404)
        self.assertEqual(response_ajena.json()["codigo"], "SESION_NO_ENCONTRADA")
        self.assertEqual(response_inexistente.json()["codigo"], "SESION_NO_ENCONTRADA")
        self.assertFalse(Asistencia.objects.filter(sesion=self.sesion, persona=self.estudiante).exists())

    def test_buscar_asistentes_mobile_sesion_ajena_e_inexistente_son_indistinguibles(self):
        """Un admin de otra org y una sesión inexistente deben recibir el mismo 404."""
        otra_organizacion = Organizacion.objects.create(
            nombre="Org Indistinguible Buscar",
            razon_social="Org Indistinguible Buscar SPA",
            rut="27.777.777-7",
        )
        self._login_admin_organizacion(otra_organizacion, username="admin_indistinguible_buscar")

        response_ajena = self.client.get(
            reverse("asistencias:sesion_asistentes_buscar", kwargs={"pk": self.sesion.pk}),
            {"q": "Ana"},
        )
        response_inexistente = self.client.get(
            reverse("asistencias:sesion_asistentes_buscar", kwargs={"pk": 999997}),
            {"q": "Ana"},
        )

        self.assertEqual(response_ajena.status_code, 404)
        self.assertEqual(response_inexistente.status_code, 404)
        self.assertEqual(response_ajena.json()["codigo"], "SESION_NO_ENCONTRADA")
        self.assertEqual(response_inexistente.json()["codigo"], "SESION_NO_ENCONTRADA")

    def test_sesion_detail_crea_persona_en_organizacion_de_la_sesion(self):
        otra_organizacion = Organizacion.objects.create(
            nombre="Otra Org",
            razon_social="Otra Org SPA",
            rut="22.222.222-2",
        )
        url = (
            f"{reverse('asistencias:sesion_detail', kwargs={'pk': self.sesion.pk})}"
            f"?periodo_mes=2&periodo_anio=2026&organizacion={otra_organizacion.pk}"
        )

        response = self.client.post(
            url,
            {
                "crear_persona_estudiante": "1",
                "nombres": "Camila",
                "apellidos": "Nueva",
                "telefono": "555",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, url)
        persona = Persona.objects.get(nombres="Camila", apellidos="Nueva")
        self.assertTrue(
            PersonaRol.objects.filter(
                persona=persona,
                organizacion=self.organizacion,
                rol__codigo="ESTUDIANTE",
            ).exists()
        )
        self.assertFalse(
            PersonaRol.objects.filter(
                persona=persona,
                organizacion=otra_organizacion,
                rol__codigo="ESTUDIANTE",
            ).exists()
        )
        self.assertFalse(Asistencia.objects.filter(sesion=self.sesion, persona=persona).exists())

    def test_sesion_detail_crea_persona_y_agrega_a_sesion_con_switch_activo(self):
        url = reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk})

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                url,
                {
                    "crear_persona_estudiante": "1",
                    "nombres": "Lucia",
                    "apellidos": "Asistente",
                    "telefono": "(+56) 9 1111-2222",
                    "agregar_a_sesion": "1",
                },
            )

        self.assertEqual(response.status_code, 302)
        persona = Persona.objects.get(nombres="Lucia", apellidos="Asistente")
        self.assertTrue(Asistencia.objects.filter(sesion=self.sesion, persona=persona).exists())
        self.sesion.refresh_from_db()
        self.assertEqual(self.sesion.estado, SesionClase.Estado.COMPLETADA)
        self.assertTrue(
            AuditLog.objects.filter(
                dominio="personas",
                accion=AuditLog.ACCION_CREAR,
                objeto_id=str(persona.pk),
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                dominio="asistencias",
                accion=AuditLog.ACCION_CREAR,
                metadata__origen="alta_rapida_sesion",
            ).exists()
        )

    def test_sesion_detail_crear_persona_sin_identidad_falla_sin_asistencia(self):
        response = self.client.post(
            reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk}),
            {
                "crear_persona_estudiante": "1",
                "nombres": "Sin",
                "apellidos": "Telefono",
                "telefono": "",
                "agregar_a_sesion": "1",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Debes ingresar al menos un telefono")
        self.assertFalse(Persona.objects.filter(nombres="Sin", apellidos="Telefono").exists())

    def test_sesion_detail_alta_rapida_no_duplica_asistencia(self):
        url = reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk})
        response = self.client.post(
            url,
            {
                "crear_persona_estudiante": "1",
                "nombres": "Duplicada",
                "apellidos": "Asistencia",
                "telefono": "+56922223333",
                "agregar_a_sesion": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        persona = Persona.objects.get(nombres="Duplicada", apellidos="Asistencia")

        response = self.client.post(
            url,
            {
                "agregar_asistentes": "1",
                "estudiantes": [str(persona.pk)],
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(Asistencia.objects.filter(sesion=self.sesion, persona=persona).count(), 1)

    def test_sesion_detail_usuario_sin_permiso_no_crea_persona_ni_asistencia(self):
        User = get_user_model()
        usuario = User.objects.create_user("sin_permiso_sesion", password=TEST_PASSWORD)
        self.client.force_login(usuario)

        response = self.client.post(
            reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk}),
            {
                "crear_persona_estudiante": "1",
                "nombres": "No",
                "apellidos": "Autorizado",
                "telefono": "+56933334444",
                "agregar_a_sesion": "1",
            },
        )

        self.assertEqual(response.status_code, 404)
        self.assertFalse(Persona.objects.filter(nombres="No", apellidos="Autorizado").exists())

    def test_sesion_detail_muestra_modal_para_crear_persona(self):
        response = self.client.get(
            reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk}),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nueva persona")
        self.assertContains(response, 'data-bs-target="#nuevaPersonaSesionModal"', html=False)
        self.assertContains(response, "Agregar a esta sesión")

    def test_estudiantes_list_muestra_metricas_operacionales_y_acciones(self):
        pago = Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.organizacion,
            fecha_pago="2026-02-03",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=30000,
            clases_asignadas=3,
        )
        asistencia = Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
            estado=Asistencia.Estado.PRESENTE,
        )
        consumo = asistencia.consumo_financiero
        consumo.pago = pago
        consumo.estado = AttendanceConsumption.Estado.CONSUMIDO
        consumo.save(update_fields=["pago", "estado"])

        response = self.client.get(
            reverse("asistencias:estudiantes_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        item = next(item for item in response.context["estudiantes"] if item["persona"] == self.estudiante)
        self.assertEqual(item["clases_pagadas"], 3)
        self.assertEqual(item["clases_usadas"], 1)
        self.assertEqual(item["clases_restantes"], 2)
        self.assertEqual(item["total_pagado"], Decimal("30000.00"))
        self.assertEqual(item["asistencias_mes"], 1)
        self.assertEqual(item["estado_financiero"], "OK")
        self.assertContains(response, "Clases pagadas")
        self.assertContains(response, "Registrar pago")
        self.assertContains(response, "open=registrar_pago")

    def test_estudiantes_list_respeta_periodo_y_organizacion_en_metricas(self):
        otra_org = Organizacion.objects.create(
            nombre="Otra Estudiantes",
            razon_social="Otra Estudiantes SPA",
            rut="44.444.444-4",
        )
        Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.organizacion,
            fecha_pago="2026-03-03",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=90000,
            clases_asignadas=9,
        )
        Payment.objects.create(
            persona=self.estudiante,
            organizacion=otra_org,
            fecha_pago="2026-02-03",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=80000,
            clases_asignadas=8,
        )

        response = self.client.get(
            reverse("asistencias:estudiantes_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        item = next(item for item in response.context["estudiantes"] if item["persona"] == self.estudiante)
        self.assertEqual(item["clases_pagadas"], 0)
        self.assertEqual(item["total_pagado"], 0)

    def test_sesion_edit_muestra_solo_disciplinas_y_profesores_vigentes(self):
        disciplina_inactiva = Disciplina.objects.create(
            organizacion=self.organizacion,
            nombre="Pilates",
            activa=False,
        )
        profesor_activo = Persona.objects.create(
            nombres="Paula",
            apellidos="Activa",
            email="paula.edit@example.com",
            activo=True,
        )
        profesor_inactivo = Persona.objects.create(
            nombres="Bruno",
            apellidos="Inactivo",
            email="bruno.edit@example.com",
            activo=False,
        )
        rol_profesor = Rol.objects.get_or_create(nombre="Profesor", codigo="PROFESOR")[0]
        PersonaRol.objects.create(
            persona=profesor_activo,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
        )
        PersonaRol.objects.create(
            persona=profesor_inactivo,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
        )

        response = self.client.get(
            reverse("asistencias:sesion_edit", kwargs={"pk": self.sesion.pk}),
            {
                "periodo_mes": 2,
                "periodo_anio": 2026,
                "organizacion": self.organizacion.pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        disciplinas = list(response.context["form"].fields["disciplina"].queryset)
        profesores = list(response.context["form"].fields["profesores"].queryset)
        self.assertIn(self.disciplina, disciplinas)
        self.assertNotIn(disciplina_inactiva, disciplinas)
        self.assertIn(profesor_activo, profesores)
        self.assertNotIn(profesor_inactivo, profesores)

    def test_sesion_detail_elimina_sesion_y_dependencias(self):
        asistencia = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        self.assertEqual(AttendanceConsumption.objects.filter(asistencia=asistencia).count(), 1)
        url = (
            f"{reverse('asistencias:sesion_detail', kwargs={'pk': self.sesion.pk})}"
            f"?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}"
        )

        response = self.client.post(
            url,
            {
                "eliminar_sesion": "1",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            f"{reverse('asistencias:sesiones_list')}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}",
        )
        self.assertFalse(SesionClase.objects.filter(pk=self.sesion.pk).exists())
        self.assertFalse(Asistencia.objects.filter(pk=asistencia.pk).exists())
        self.assertEqual(AttendanceConsumption.objects.count(), 0)

    def test_sesion_detail_permite_eliminar_asistente_individual(self):
        asistencia = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        self.assertEqual(AttendanceConsumption.objects.filter(asistencia=asistencia).count(), 1)
        url = (
            f"{reverse('asistencias:sesion_detail', kwargs={'pk': self.sesion.pk})}"
            f"?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}"
        )

        response = self.client.post(
            url,
            {
                "eliminar_asistente": "1",
                "asistencia_id": asistencia.pk,
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, url)
        self.assertFalse(Asistencia.objects.filter(pk=asistencia.pk).exists())
        self.assertEqual(AttendanceConsumption.objects.count(), 0)

    def test_asistencias_list_muestra_checkboxes_marcados_para_asistentes_existentes(self):
        Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)

        response = self.client.get(
            reverse("asistencias:asistencias_list"),
            {
                "sesion_id": self.sesion.pk,
                "periodo_mes": 2,
                "periodo_anio": 2026,
                "organizacion": self.organizacion.pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            f'id="estudiante_{self.estudiante.pk}"',
            html=False,
        )
        self.assertContains(
            response,
            f'value="{self.estudiante.pk}"',
            html=False,
        )
        self.assertContains(response, "seleccionar_visibles", html=False)
        self.assertContains(response, "limpiar_seleccion", html=False)
        self.assertContains(response, "checked", html=False)

    def test_asistencias_list_muestra_total_por_disciplina_en_periodo(self):
        otra_disciplina = Disciplina.objects.create(
            organizacion=self.organizacion,
            nombre="Teatro",
        )
        otra_sesion_misma_disciplina = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-02-20",
            estado=SesionClase.Estado.COMPLETADA,
        )
        sesion_otra_disciplina = SesionClase.objects.create(
            disciplina=otra_disciplina,
            fecha="2026-02-21",
            estado=SesionClase.Estado.COMPLETADA,
        )
        segundo_estudiante = Persona.objects.create(
            nombres="Luis",
            apellidos="Rojas",
            email="luis@example.com",
        )
        rol_estudiante = Rol.objects.get(codigo="ESTUDIANTE")
        PersonaRol.objects.create(
            persona=segundo_estudiante,
            rol=rol_estudiante,
            organizacion=self.organizacion,
            activo=True,
        )
        Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        Asistencia.objects.create(sesion=otra_sesion_misma_disciplina, persona=segundo_estudiante)
        Asistencia.objects.create(sesion=sesion_otra_disciplina, persona=self.estudiante)

        response = self.client.get(
            reverse("asistencias:asistencias_list"),
            {
                "sesion_id": self.sesion.pk,
                "periodo_mes": 2,
                "periodo_anio": 2026,
                "organizacion": self.organizacion.pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Total asistencia a Flexibilidad: 2")

    def test_sesion_detail_enlaza_profesor_a_perfil_personas_con_filtros(self):
        profesor = Persona.objects.create(
            nombres="Paula",
            apellidos="Mora",
            email="paula.sesion@example.com",
        )
        rol_profesor = Rol.objects.create(nombre="Profesor", codigo="PROFESOR")
        PersonaRol.objects.create(
            persona=profesor,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
        )
        self.sesion.profesores.set([profesor])

        response = self.client.get(
            reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk}),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        enlace = (
            f'{reverse("personas:persona_detail", kwargs={"pk": profesor.pk})}'
            f"?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}"
        )
        self.assertContains(response, f'href="{enlace}"', html=False)

    def test_ruta_persona_detail_en_asistencias_no_existe(self):
        response = self.client.get(
            f"/asistencias/personas/{self.estudiante.pk}/",
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 404)

    def test_sesion_detail_muestra_estado_de_pago_del_asistente(self):
        pago = Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.organizacion,
            fecha_pago="2026-02-20",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=1,
        )
        asistencia = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.pago, pago)

        response = self.client.get(
            reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk}),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Estado de pago")
        self.assertContains(response, "Pagada")

    def test_sesion_detail_muestra_boton_editar_sesion_con_filtros(self):
        response = self.client.get(
            reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk}),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        enlace = (
            f'{reverse("asistencias:sesion_edit", kwargs={"pk": self.sesion.pk})}'
            f"?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}"
        )
        self.assertContains(response, f'href="{enlace}"', html=False)

    def test_sesion_edit_actualiza_datos_y_redirige_a_detalle(self):
        profesor = Persona.objects.create(
            nombres="Paula",
            apellidos="Edita",
            email="paula.edita@example.com",
        )
        rol_profesor = Rol.objects.create(nombre="Profesor", codigo="PROFESOR")
        PersonaRol.objects.create(
            persona=profesor,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
        )
        otra_disciplina = Disciplina.objects.create(
            organizacion=self.organizacion,
            nombre="Teatro",
        )
        url = (
            f"{reverse('asistencias:sesion_edit', kwargs={'pk': self.sesion.pk})}"
            f"?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}"
        )

        response = self.client.post(
            url,
            {
                "disciplina": otra_disciplina.pk,
                "fecha": "2026-02-27",
                "profesores": [profesor.pk],
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            f"{reverse('asistencias:sesion_detail', kwargs={'pk': self.sesion.pk})}?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}",
        )
        self.sesion.refresh_from_db()
        self.assertEqual(self.sesion.disciplina, otra_disciplina)
        self.assertEqual(str(self.sesion.fecha), "2026-02-27")
        self.assertEqual(list(self.sesion.profesores.all()), [profesor])

    def test_dashboard_estudiantes_activos_cuenta_personas_unicas_con_asistencia(self):
        sesion_extra = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-02-27",
            estado=SesionClase.Estado.PROGRAMADA,
        )
        otro_estudiante = Persona.objects.create(
            nombres="Luis",
            apellidos="Rojas",
            email="luis@example.com",
        )
        rol_estudiante = Rol.objects.get(codigo="ESTUDIANTE")
        PersonaRol.objects.create(
            persona=otro_estudiante,
            rol=rol_estudiante,
            organizacion=self.organizacion,
            activo=True,
        )
        Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        Asistencia.objects.create(sesion=sesion_extra, persona=self.estudiante)
        Asistencia.objects.create(sesion=sesion_extra, persona=otro_estudiante)

        response = self.client.get(
            reverse("asistencias:dashboard"),
            {"periodo_mes": 2, "periodo_anio": 2026},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["estudiantes_activos_mes"], 2)

    def test_menu_superior_permite_cerrar_sesion(self):
        response = self.client.get(reverse("asistencias:dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cerrar sesión")
        self.assertContains(response, 'action="/accounts/logout/"')

        logout_response = self.client.post("/accounts/logout/")

        self.assertEqual(logout_response.status_code, 302)
        self.assertEqual(logout_response.url, "/accounts/login/")

    def test_dashboard_sesiones_realizadas_cuenta_solo_completadas_del_mes(self):
        SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-02-27",
            estado=SesionClase.Estado.COMPLETADA,
        )
        SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-03-01",
            estado=SesionClase.Estado.COMPLETADA,
        )

        response = self.client.get(
            reverse("asistencias:dashboard"),
            {"periodo_mes": 2, "periodo_anio": 2026},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["sesiones_realizadas_mes"], 1)

    def test_dashboard_muestra_columnas_de_deuda_y_mas_asistencia(self):
        otro_estudiante = Persona.objects.create(
            nombres="Luis",
            apellidos="Rojas",
            email="luis.dashboard@example.com",
        )
        tercer_estudiante = Persona.objects.create(
            nombres="Marta",
            apellidos="Lopez",
            email="marta.dashboard@example.com",
        )
        rol_estudiante = Rol.objects.get(codigo="ESTUDIANTE")
        PersonaRol.objects.create(
            persona=otro_estudiante,
            rol=rol_estudiante,
            organizacion=self.organizacion,
            activo=True,
        )
        PersonaRol.objects.create(
            persona=tercer_estudiante,
            rol=rol_estudiante,
            organizacion=self.organizacion,
            activo=True,
        )
        sesion_dos = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-02-27",
            estado=SesionClase.Estado.COMPLETADA,
        )
        asistencia_ana_1 = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        asistencia_ana_2 = Asistencia.objects.create(sesion=sesion_dos, persona=self.estudiante)
        asistencia_luis = Asistencia.objects.create(sesion=self.sesion, persona=otro_estudiante)
        Payment.objects.create(
            persona=tercer_estudiante,
            organizacion=self.organizacion,
            fecha_pago="2026-02-26",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=30000,
            clases_asignadas=3,
        )
        consumo_luis = AttendanceConsumption.objects.get(asistencia=asistencia_luis)
        consumo_luis.estado = AttendanceConsumption.Estado.DEUDA
        consumo_luis.pago = None
        consumo_luis.save(update_fields=["estado", "pago", "actualizado_en"])

        response = self.client.get(
            reverse("asistencias:dashboard"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Estudiantes con deuda")
        self.assertContains(response, "Estudiantes con más asistencia")
        self.assertContains(response, "Alumnos con clases disponibles")
        self.assertNotContains(response, "Estudiantes sin asistencia")
        self.assertContains(response, "Luis Rojas")
        self.assertContains(response, "Ana Diaz")
        self.assertContains(response, "Marta Lopez")
        self.assertContains(response, "3 pagadas")
        self.assertIn(otro_estudiante, list(response.context["estudiantes_con_deuda"]))
        self.assertEqual(response.context["estudiantes_con_deuda"][0], self.estudiante)
        self.assertEqual(response.context["estudiantes_con_deuda"][0].clases_deuda, 2)
        self.assertEqual(response.context["estudiantes_con_deuda"][1], otro_estudiante)
        self.assertEqual(response.context["estudiantes_con_deuda"][1].clases_deuda, 1)
        self.assertEqual(response.context["estudiantes_con_mas_asistencia"][0], self.estudiante)
        self.assertEqual(response.context["estudiantes_con_mas_asistencia"][0].total_asistencias_mes, 2)
        self.assertEqual(response.context["estudiantes_con_clases_restantes"][0]["persona"], tercer_estudiante)
        self.assertEqual(response.context["estudiantes_con_clases_restantes"][0]["saldo_clases"], 3)

    def test_dashboard_mas_asistencia_sin_datos_no_usa_colspan_para_datatables(self):
        response = self.client.get(
            reverse("asistencias:dashboard"),
            {"periodo_mes": 5, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["estudiantes_con_mas_asistencia"]), [])
        self.assertContains(response, 'data-empty-row="true"')
        self.assertContains(response, "Sin asistencias registradas.")
        self.assertNotContains(response, 'colspan="2" class="text-muted">Sin asistencias registradas.</td>')

    def test_sesiones_list_muestra_mensaje_cancelada_en_vez_de_asistentes_cero(self):
        self.sesion.estado = SesionClase.Estado.CANCELADA
        self.sesion.save(update_fields=["estado"])

        response = self.client.get(
            reverse("asistencias:sesiones_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "sesión cancelada")
        self.assertContains(response, 'bi-x-circle-fill', html=False)
        self.assertContains(response, 'title="Cancelada"', html=False)
        self.assertNotContains(response, "asistentes: 0")

    def test_sesiones_list_muestra_iconos_por_estado(self):
        self.disciplina.badge_color = Disciplina.BadgeColor.VERDE
        self.disciplina.save(update_fields=["badge_color"])
        SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-02-27",
            estado=SesionClase.Estado.COMPLETADA,
        )
        SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-02-28",
            estado=SesionClase.Estado.CANCELADA,
        )

        response = self.client.get(
            reverse("asistencias:sesiones_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "disciplina-badge-verde")
        self.assertContains(response, 'd-flex align-items-center gap-2 mb-1', html=False)
        self.assertContains(response, 'bi-calendar-event-fill', html=False)
        self.assertContains(response, 'text-info fs-5 flex-shrink-0', html=False)
        self.assertContains(response, 'title="Programada"', html=False)
        self.assertContains(response, 'bi-check-circle-fill', html=False)
        self.assertContains(response, 'text-success fs-5 flex-shrink-0', html=False)
        self.assertContains(response, 'title="Completada"', html=False)
        self.assertContains(response, 'bi-x-circle-fill', html=False)
        self.assertContains(response, 'text-danger fs-5 flex-shrink-0', html=False)
        self.assertContains(response, 'title="Cancelada"', html=False)

    def test_calendario_usa_url_calendario_y_redirige_url_legacy_sesiones(self):
        self.assertEqual(reverse("asistencias:sesiones_list"), "/asistencias/calendario/")

        response = self.client.get(
            "/asistencias/sesiones/",
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            f"/asistencias/calendario/?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}",
        )

    def test_calendario_crea_sesiones_masivas_en_mes_seleccionado(self):
        response = self.client.post(
            (
                f"{reverse('asistencias:sesiones_list')}"
                f"?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}"
            ),
            {
                "crear_sesiones_masivas": "1",
                "disciplina": str(self.disciplina.pk),
                "dias_semana": ["1", "3"],
                "max_sesiones": "",
                "profesores": [],
            },
        )

        self.assertEqual(response.status_code, 302)
        fechas = set(
            SesionClase.objects.filter(
                disciplina=self.disciplina,
                fecha__year=2026,
                fecha__month=2,
            ).values_list("fecha", flat=True)
        )
        self.assertEqual(len(fechas), 8)
        self.assertIn(date(2026, 2, 3), fechas)
        self.assertIn(date(2026, 2, 26), fechas)

    def test_calendario_creacion_masiva_respeta_maximo_de_sesiones(self):
        disciplina = Disciplina.objects.create(
            organizacion=self.organizacion,
            nombre="Danza",
        )

        response = self.client.post(
            (
                f"{reverse('asistencias:sesiones_list')}"
                f"?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}"
            ),
            {
                "crear_sesiones_masivas": "1",
                "disciplina": str(disciplina.pk),
                "dias_semana": ["1", "3"],
                "max_sesiones": "1",
                "profesores": [],
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(SesionClase.objects.filter(disciplina=disciplina).count(), 1)
        self.assertTrue(SesionClase.objects.filter(disciplina=disciplina, fecha="2026-02-03").exists())

    def test_disciplinas_list_muestra_resumen_operativo(self):
        self.disciplina.badge_color = Disciplina.BadgeColor.CAFE
        self.disciplina.save(update_fields=["badge_color"])
        sesion_realizada = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-02-27",
            estado=SesionClase.Estado.COMPLETADA,
        )
        otro_estudiante = Persona.objects.create(
            nombres="Luis",
            apellidos="Rojas",
            email="luis2@example.com",
        )
        rol_estudiante = Rol.objects.get(codigo="ESTUDIANTE")
        PersonaRol.objects.create(
            persona=otro_estudiante,
            rol=rol_estudiante,
            organizacion=self.organizacion,
            activo=True,
        )
        Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        Asistencia.objects.create(sesion=sesion_realizada, persona=self.estudiante)
        Asistencia.objects.create(sesion=sesion_realizada, persona=otro_estudiante)

        response = self.client.get(
            reverse("asistencias:disciplinas_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        disciplina = response.context["disciplinas"].get(pk=self.disciplina.pk)
        self.assertEqual(disciplina.sesiones_realizadas, 1)
        self.assertEqual(disciplina.sesiones_periodo, 2)
        self.assertEqual(disciplina.asistencias_periodo, 3)
        self.assertEqual(disciplina.estudiantes_unicos, 2)
        self.assertContains(response, "disciplina-badge-cafe")

    def test_disciplinas_list_ordena_activas_primero_y_luego_alfabetico(self):
        disciplina_activa = Disciplina.objects.create(
            organizacion=self.organizacion,
            nombre="Zumba",
            activa=True,
        )
        disciplina_inactiva = Disciplina.objects.create(
            organizacion=self.organizacion,
            nombre="Acrobacia",
            activa=False,
        )

        response = self.client.get(
            reverse("asistencias:disciplinas_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        disciplinas = list(response.context["disciplinas"])
        self.assertEqual(
            [disciplina.nombre for disciplina in disciplinas[:3]],
            ["Flexibilidad", "Zumba", "Acrobacia"],
        )
        self.assertTrue(disciplinas[0].activa)
        self.assertTrue(disciplinas[1].activa)
        self.assertFalse(disciplinas[2].activa)

    def test_disciplina_create_redirige_a_detalle_con_filtros(self):
        query = f"periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}"
        url = f"{reverse('asistencias:disciplina_create')}?{query}"

        response = self.client.post(
            url,
            {
                "organizacion": self.organizacion.pk,
                "nombre": "Contemporaneo",
                "nivel": "Intermedio",
                "badge_color": Disciplina.BadgeColor.MORADO,
                "descripcion": "Taller de danza contemporanea",
                "activa": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        nueva = Disciplina.objects.get(nombre="Contemporaneo")
        self.assertEqual(nueva.badge_color, Disciplina.BadgeColor.MORADO)
        self.assertEqual(
            response.url,
            f"{reverse('asistencias:disciplina_detail', kwargs={'pk': nueva.pk})}?{query}",
        )

    def test_disciplina_edit_actualiza_nombre(self):
        query = f"periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}"
        url = f"{reverse('asistencias:disciplina_edit', kwargs={'pk': self.disciplina.pk})}?{query}"

        response = self.client.post(
            url,
            {
                "organizacion": self.organizacion.pk,
                "nombre": "Flexibilidad avanzada",
                "nivel": "",
                "badge_color": Disciplina.BadgeColor.NARANJO,
                "descripcion": "Actualizada",
                "activa": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.disciplina.refresh_from_db()
        self.assertEqual(self.disciplina.nombre, "Flexibilidad avanzada")
        self.assertEqual(self.disciplina.badge_color, Disciplina.BadgeColor.NARANJO)

    def test_disciplina_form_muestra_ocho_colores_de_badge(self):
        response = self.client.get(reverse("asistencias:disciplina_create"))

        self.assertEqual(response.status_code, 200)
        for color in ["Rojo", "Naranjo", "Azul", "Celeste", "Amarillo", "Verde", "Cafe", "Morado"]:
            self.assertContains(response, color)
        self.assertContains(response, "disciplina-badge-rojo")
        self.assertContains(response, "disciplina-badge-naranjo")
        self.assertContains(response, "disciplina-badge-azul")
        self.assertContains(response, "disciplina-badge-celeste")
        self.assertContains(response, "disciplina-badge-amarillo")
        self.assertContains(response, "disciplina-badge-verde")
        self.assertContains(response, "disciplina-badge-cafe")
        self.assertContains(response, "disciplina-badge-morado")

    def test_disciplina_detail_muestra_profesores_en_descripcion_y_asistentes_en_tabla(self):
        profesor = Persona.objects.create(
            nombres="Paula",
            apellidos="Mora",
            email="paula.disciplina@example.com",
        )
        rol_profesor = Rol.objects.create(nombre="Profesor", codigo="PROFESOR")
        PersonaRol.objects.create(
            persona=profesor,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
        )
        self.sesion.profesores.set([profesor])
        Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)

        response = self.client.get(
            reverse("asistencias:disciplina_detail", kwargs={"pk": self.disciplina.pk}),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Profesores del período")
        self.assertContains(response, "Paula Mora")
        self.assertContains(response, 'id="tabla-sesiones-disciplina"', html=False)
        self.assertContains(response, "<th>Asistentes</th>", html=False)
        self.assertContains(response, "<th>Asistencias</th>", html=False)
        self.assertContains(response, "<th>Estado</th>", html=False)
        self.assertContains(response, "Ana Diaz")
        self.assertNotContains(response, "<th>Profesores</th>", html=False)
        self.assertNotContains(response, "<th>Presentes</th>", html=False)
        self.assertNotContains(response, "<th>Ausentes</th>", html=False)
        self.assertNotContains(response, "<th>Justificadas</th>", html=False)

    def test_profesores_list_boton_ver_perfil_envia_filtros_a_personas(self):
        profesor = Persona.objects.create(
            nombres="Paula",
            apellidos="Mora",
            email="paula@example.com",
        )
        rol_profesor = Rol.objects.create(nombre="Profesor", codigo="PROFESOR")
        PersonaRol.objects.create(
            persona=profesor,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
        )
        self.sesion.profesores.set([profesor])

        response = self.client.get(
            reverse("asistencias:profesores_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        enlace = (
            f'{reverse("personas:persona_detail", kwargs={"pk": profesor.pk})}'
            f"?periodo_mes=2&periodo_anio=2026&organizacion={self.organizacion.pk}"
        )
        self.assertContains(response, f'href="{enlace}"', html=False)
        self.assertNotContains(response, 'id="filtro-organizacion"', html=False)

    def test_profesores_list_oculta_profesores_sin_asistencias_ni_sesiones_activas(self):
        rol_profesor = Rol.objects.create(nombre="Profesor", codigo="PROFESOR")
        profesor_inactivo = Persona.objects.create(
            nombres="Pedro",
            apellidos="Silva",
            email="pedro@example.com",
        )
        profesor_con_sesion = Persona.objects.create(
            nombres="Laura",
            apellidos="Torres",
            email="laura@example.com",
        )
        PersonaRol.objects.create(
            persona=profesor_inactivo,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
        )
        PersonaRol.objects.create(
            persona=profesor_con_sesion,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
        )
        self.sesion.profesores.set([profesor_con_sesion])

        response = self.client.get(
            reverse("asistencias:profesores_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Laura Torres")
        self.assertNotContains(response, "Pedro Silva")

    def test_profesores_list_oculta_profesor_con_solo_sesiones_canceladas(self):
        rol_profesor = Rol.objects.create(nombre="Profesor", codigo="PROFESOR")
        profesor = Persona.objects.create(
            nombres="Mario",
            apellidos="Cancelado",
            email="mario@example.com",
        )
        PersonaRol.objects.create(
            persona=profesor,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
        )
        sesion_cancelada = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-02-27",
            estado=SesionClase.Estado.CANCELADA,
        )
        sesion_cancelada.profesores.set([profesor])

        response = self.client.get(
            reverse("asistencias:profesores_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Mario Cancelado")

    def test_profesores_list_muestra_cards_resumen_del_periodo(self):
        profesor = Persona.objects.create(
            nombres="Paula",
            apellidos="Mora",
            email="paula.cards@example.com",
        )
        rol_profesor = Rol.objects.create(nombre="Profesor", codigo="PROFESOR")
        PersonaRol.objects.create(
            persona=profesor,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
        )
        self.sesion.estado = SesionClase.Estado.COMPLETADA
        self.sesion.save(update_fields=["estado"])
        self.sesion.profesores.set([profesor])
        Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)

        segundo_estudiante = Persona.objects.create(
            nombres="Luis",
            apellidos="Rojas",
            email="luis.cards@example.com",
        )
        PersonaRol.objects.create(
            persona=segundo_estudiante,
            rol=Rol.objects.get(codigo="ESTUDIANTE"),
            organizacion=self.organizacion,
            activo=True,
        )
        sesion_dos = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-02-27",
            estado=SesionClase.Estado.COMPLETADA,
        )
        sesion_dos.profesores.set([profesor])
        Asistencia.objects.create(sesion=sesion_dos, persona=segundo_estudiante)

        response = self.client.get(
            reverse("asistencias:profesores_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context["resumen_profesores"],
            {
                "alumnos_unicos": 2,
                "sesiones_realizadas": 2,
                "asistencias_mes": 2,
                "profesores_activos": 1,
            },
        )
        self.assertContains(response, "Total alumnos únicos")
        self.assertContains(response, "Total sesiones realizadas")
        self.assertContains(response, "Total general de asistencias del mes")
        self.assertContains(response, "Total de profesores activos")

    def test_profesores_list_muestra_pagos_y_organizacion_como_badge_en_nombre(self):
        profesor = Persona.objects.create(
            nombres="Paula",
            apellidos="Mora",
            email="paula.pagos@example.com",
        )
        rol_profesor = Rol.objects.create(nombre="Profesor", codigo="PROFESOR")
        PersonaRol.objects.create(
            persona=profesor,
            rol=rol_profesor,
            organizacion=self.organizacion,
            activo=True,
            valor_clase=Decimal("10000"),
            retencion_sii=Decimal("15.25"),
        )
        self.sesion.estado = SesionClase.Estado.COMPLETADA
        self.sesion.save(update_fields=["estado"])
        self.sesion.profesores.set([profesor])
        Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)

        segundo_estudiante = Persona.objects.create(
            nombres="Luis",
            apellidos="Rojas",
            email="luis.pagos.profesor@example.com",
        )
        PersonaRol.objects.create(
            persona=segundo_estudiante,
            rol=Rol.objects.get(codigo="ESTUDIANTE"),
            organizacion=self.organizacion,
            activo=True,
        )
        sesion_dos = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha="2026-02-27",
            estado=SesionClase.Estado.COMPLETADA,
        )
        sesion_dos.profesores.set([profesor])
        Asistencia.objects.create(sesion=sesion_dos, persona=segundo_estudiante)

        response = self.client.get(
            reverse("asistencias:profesores_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        item_profesor = response.context["profesores"][0]
        self.assertEqual(item_profesor["pago_bruto"], Decimal("20000"))
        self.assertEqual(item_profesor["retencion_sii_monto"], Decimal("3050.00"))
        self.assertEqual(item_profesor["pago_neto"], Decimal("16950.00"))
        self.assertContains(response, "Pago bruto")
        self.assertContains(response, "Retención SII")
        self.assertContains(response, "Pago neto")
        self.assertContains(response, "$ 20.000")
        self.assertContains(response, "$ 3.050")
        self.assertContains(response, "$ 16.950")
        self.assertContains(response, f'<span class="badge text-bg-light ms-1">{self.organizacion.nombre}</span>', html=False)
        self.assertNotContains(response, "<th>Organización</th>", html=False)

    def test_asistencias_list_colorea_asistentes_por_estado_financiero(self):
        rol_estudiante = Rol.objects.get(codigo="ESTUDIANTE")
        estudiante_pagado = Persona.objects.create(
            nombres="Luis",
            apellidos="Pagado",
            email="luis.pagado@example.com",
        )
        estudiante_liberado = Persona.objects.create(
            nombres="Marta",
            apellidos="Liberada",
            email="marta.liberada@example.com",
        )
        PersonaRol.objects.create(
            persona=estudiante_pagado,
            rol=rol_estudiante,
            organizacion=self.organizacion,
            activo=True,
        )
        PersonaRol.objects.create(
            persona=estudiante_liberado,
            rol=rol_estudiante,
            organizacion=self.organizacion,
            activo=True,
        )
        Payment.objects.create(
            persona=estudiante_pagado,
            organizacion=self.organizacion,
            fecha_pago="2026-02-25",
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=1,
        )

        asistencia_deuda = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        asistencia_pagada = Asistencia.objects.create(sesion=self.sesion, persona=estudiante_pagado)
        asistencia_liberada = Asistencia.objects.create(sesion=self.sesion, persona=estudiante_liberado)
        consumo_liberado = AttendanceConsumption.objects.get(asistencia=asistencia_liberada)
        consumo_liberado.estado = AttendanceConsumption.Estado.PENDIENTE
        consumo_liberado.pago = None
        consumo_liberado.save(update_fields=["estado", "pago", "actualizado_en"])

        response = self.client.get(
            reverse("asistencias:asistencias_list"),
            {"periodo_mes": 2, "periodo_anio": 2026, "organizacion": self.organizacion.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            f'class="badge text-bg-warning me-1 text-decoration-none"',
            html=False,
        )
        self.assertContains(response, "Ana Diaz")
        self.assertContains(
            response,
            f'class="badge text-bg-success me-1 text-decoration-none"',
            html=False,
        )
        self.assertContains(response, "Luis Pagado")
        self.assertContains(
            response,
            f'class="badge text-bg-primary me-1 text-decoration-none"',
            html=False,
        )
        self.assertContains(response, "Marta Liberada")


class SprintDosDominioAsistenciasTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.organizacion = Organizacion.objects.create(
            nombre="Org Sprint 2",
            razon_social="Org Sprint 2 SpA",
            rut="70.000.000-1",
        )
        self.otra_organizacion = Organizacion.objects.create(
            nombre="Otra Org Sprint 2",
            razon_social="Otra Org Sprint 2 SpA",
            rut="70.000.000-2",
        )
        self.rol_admin = Rol.objects.create(nombre="Admin Sprint 2", codigo="ADMIN")
        self.rol_profesor = Rol.objects.create(nombre="Profesor Sprint 2", codigo="PROFESOR")
        self.rol_estudiante = Rol.objects.create(nombre="Estudiante Sprint 2", codigo="ESTUDIANTE")
        self.admin = User.objects.create_user("admin_sprint2", password=TEST_PASSWORD)
        self.persona_admin = Persona.objects.create(
            nombres="Admin",
            apellidos="Sprint",
            user=self.admin,
        )
        PersonaRol.objects.create(
            persona=self.persona_admin,
            rol=self.rol_admin,
            organizacion=self.organizacion,
            activo=True,
        )
        self.profesor_asignado = self._crear_profesor(
            "profesor_asignado",
            self.organizacion,
        )
        self.profesor_no_asignado = self._crear_profesor(
            "profesor_no_asignado",
            self.organizacion,
        )
        self.profesor_otra_org = self._crear_profesor(
            "profesor_otra_org",
            self.otra_organizacion,
        )
        self.admin_otra_org = User.objects.create_user(
            "admin_otra_org_sprint2",
            password=TEST_PASSWORD,
        )
        persona_admin_otra = Persona.objects.create(
            nombres="Admin",
            apellidos="Otra org",
            user=self.admin_otra_org,
        )
        PersonaRol.objects.create(
            persona=persona_admin_otra,
            rol=self.rol_admin,
            organizacion=self.otra_organizacion,
            activo=True,
        )
        self.estudiante = Persona.objects.create(
            nombres="Estudiante",
            apellidos="Sprint",
        )
        PersonaRol.objects.create(
            persona=self.estudiante,
            rol=self.rol_estudiante,
            organizacion=self.organizacion,
            activo=True,
        )
        self.disciplina = Disciplina.objects.create(
            organizacion=self.organizacion,
            nombre="Yoga Sprint 2",
        )
        self.sesion = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha=date(2026, 7, 15),
        )
        asignar_profesora_a_sesion(user=self.profesor_asignado, sesion=self.sesion)

    def _crear_profesor(self, username, organizacion):
        return crear_usuario_con_rol(
            username=username,
            password=TEST_PASSWORD,
            rol=self.rol_profesor,
            organizacion=organizacion,
            apellidos="Sprint",
        )

    def _crear_pago(self, *, clases=10, plan=None, organizacion=None):
        return Payment.objects.create(
            persona=self.estudiante,
            organizacion=organizacion or self.organizacion,
            plan=plan,
            fecha_pago=date(2026, 7, 2),
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=clases,
        )

    def test_matriz_transiciones_recalcula_consumo_sin_duplicados(self):
        casos = (
            (Asistencia.Estado.PRESENTE, Asistencia.Estado.AUSENTE, AttendanceConsumption.Estado.CONSUMIDO),
            (Asistencia.Estado.PRESENTE, Asistencia.Estado.JUSTIFICADA, AttendanceConsumption.Estado.CONSUMIDO),
            (Asistencia.Estado.AUSENTE, Asistencia.Estado.PRESENTE, AttendanceConsumption.Estado.CONSUMIDO),
            (Asistencia.Estado.JUSTIFICADA, Asistencia.Estado.PRESENTE, AttendanceConsumption.Estado.CONSUMIDO),
            (Asistencia.Estado.PRESENTE, Asistencia.Estado.PRESENTE, AttendanceConsumption.Estado.CONSUMIDO),
            (Asistencia.Estado.AUSENTE, Asistencia.Estado.AUSENTE, AttendanceConsumption.Estado.CONSUMIDO),
            (Asistencia.Estado.JUSTIFICADA, Asistencia.Estado.JUSTIFICADA, AttendanceConsumption.Estado.CONSUMIDO),
        )
        for indice, (origen, destino, esperado) in enumerate(casos):
            persona = Persona.objects.create(
                nombres=f"Transición {indice}",
                apellidos="Sprint",
            )
            PersonaRol.objects.create(
                persona=persona,
                rol=self.rol_estudiante,
                organizacion=self.organizacion,
                activo=True,
            )
            Payment.objects.create(
                persona=persona,
                organizacion=self.organizacion,
                fecha_pago=date(2026, 7, 2),
                metodo_pago=Payment.Metodo.EFECTIVO,
                aplica_iva=False,
                monto_referencia=10000,
                clases_asignadas=1,
            )
            asistencia = Asistencia.objects.create(
                sesion=self.sesion,
                persona=persona,
                estado=origen,
            )
            with self.subTest(origen=origen, destino=destino):
                asistencia, consumo = cambiar_estado_asistencia(
                    asistencia=asistencia,
                    estado=destino,
                    usuario=self.admin,
                )
                segundo = asignar_consumo_asistencia(asistencia)
                self.assertEqual(asistencia.estado, destino)
                self.assertEqual(consumo.estado, esperado)
                self.assertEqual(segundo.pk, consumo.pk)
                self.assertEqual(
                    AttendanceConsumption.objects.filter(asistencia=asistencia).count(),
                    1,
                )
                if esperado == AttendanceConsumption.Estado.CONSUMIDO:
                    self.assertIsNotNone(consumo.pago_id)
                else:
                    self.assertIsNone(consumo.pago_id)

    def test_correccion_entre_estados_ordinarios_no_recupera_clase(self):
        pago = self._crear_pago(clases=1)
        asistencia = Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
            estado=Asistencia.Estado.PRESENTE,
        )
        self.assertEqual(pago.saldo_clases, 0)

        cambiar_estado_asistencia(
            asistencia=asistencia,
            estado=Asistencia.Estado.JUSTIFICADA,
            usuario=self.admin,
        )
        pago.refresh_from_db()
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(pago.saldo_clases, 0)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.CONSUMIDO)
        self.assertEqual(consumo.pago, pago)

        cambiar_estado_asistencia(
            asistencia=asistencia,
            estado=Asistencia.Estado.PRESENTE,
            usuario=self.admin,
        )
        pago.refresh_from_db()
        self.assertEqual(pago.saldo_clases, 0)
        self.assertEqual(
            AttendanceConsumption.objects.filter(asistencia=asistencia).count(),
            1,
        )

    def test_ausencia_con_pago_valido_queda_consumida(self):
        pago = self._crear_pago(clases=1)
        asistencia = Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
            estado=Asistencia.Estado.AUSENTE,
        )
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.CONSUMIDO)
        self.assertEqual(consumo.pago, pago)

    def test_ausencia_sin_pago_valido_queda_deuda(self):
        asistencia = Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
            estado=Asistencia.Estado.AUSENTE,
        )
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.DEUDA)
        self.assertIsNone(consumo.pago)

    def test_justificacion_con_pago_valido_queda_consumida(self):
        pago = self._crear_pago(clases=1)
        asistencia = Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
            estado=Asistencia.Estado.JUSTIFICADA,
        )
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.CONSUMIDO)
        self.assertEqual(consumo.pago, pago)

    def test_plan_o_pago_mes_anterior_no_se_arrastra(self):
        plan = PaymentPlan.objects.create(
            organizacion=self.organizacion,
            nombre="Plan junio",
            num_clases=2,
            precio=20000,
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 30),
        )
        Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.organizacion,
            plan=plan,
            fecha_pago=date(2026, 6, 2),
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=20000,
            clases_asignadas=2,
        )
        asistencia = Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
            estado=Asistencia.Estado.AUSENTE,
        )
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.DEUDA)
        self.assertIsNone(consumo.pago)

    def test_presente_sin_pago_o_derecho_queda_deuda(self):
        asistencia = Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
            estado=Asistencia.Estado.PRESENTE,
        )
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.DEUDA)
        self.assertIsNone(consumo.pago)

    def test_plan_vencido_no_otorga_derecho(self):
        plan = PaymentPlan.objects.create(
            organizacion=self.organizacion,
            nombre="Plan vencido",
            num_clases=2,
            precio=20000,
            fecha_fin=date(2026, 6, 30),
        )
        self._crear_pago(clases=2, plan=plan)
        asistencia = Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
        )
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.DEUDA)
        self.assertIsNone(consumo.pago)

    def test_plan_sin_saldo_no_consume_clase_adicional(self):
        pago = self._crear_pago(clases=1)
        primera = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        otra_sesion = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha=date(2026, 7, 16),
        )
        segunda = Asistencia.objects.create(sesion=otra_sesion, persona=self.estudiante)
        self.assertEqual(
            AttendanceConsumption.objects.get(asistencia=primera).estado,
            AttendanceConsumption.Estado.CONSUMIDO,
        )
        self.assertEqual(
            AttendanceConsumption.objects.get(asistencia=segunda).estado,
            AttendanceConsumption.Estado.DEUDA,
        )
        self.assertEqual(pago.consumos.filter(estado=AttendanceConsumption.Estado.CONSUMIDO).count(), 1)

    def test_pago_de_otra_organizacion_no_otorga_derecho(self):
        self._crear_pago(clases=3, organizacion=self.otra_organizacion)
        asistencia = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.DEUDA)
        self.assertIsNone(consumo.pago)

    def test_eliminar_asistencia_elimina_consumo_y_recupera_saldo(self):
        pago = self._crear_pago(clases=1)
        asistencia = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        self.assertEqual(pago.saldo_clases, 0)
        asistencia.delete()
        pago.refresh_from_db()
        self.assertEqual(pago.saldo_clases, 1)
        self.assertFalse(AttendanceConsumption.objects.filter(asistencia_id=asistencia.pk).exists())

    def test_clase_liberada_es_explicita_auditable_y_reversible(self):
        pago = self._crear_pago(clases=1)
        asistencia = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        with self.captureOnCommitCallbacks(execute=True):
            liberacion, consumo = liberar_clase(
                asistencia=asistencia,
                motivo="Invitación institucional",
                usuario=self.admin,
            )
        self.assertTrue(liberacion.activa)
        self.assertEqual(liberacion.organizacion, self.organizacion)
        self.assertEqual(liberacion.liberada_por, self.admin)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.PENDIENTE)
        self.assertIsNone(consumo.pago)
        pago.refresh_from_db()
        self.assertEqual(pago.saldo_clases, 1)
        self.assertTrue(AuditLog.objects.filter(objeto_id=str(liberacion.pk), resumen="Clase liberada").exists())

        with self.captureOnCommitCallbacks(execute=True):
            liberacion, consumo = revertir_clase_liberada(
                asistencia=asistencia,
                usuario=self.admin,
            )
        self.assertFalse(liberacion.activa)
        self.assertEqual(liberacion.revertida_por, self.admin)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.CONSUMIDO)
        self.assertEqual(consumo.pago, pago)
        self.assertTrue(
            AuditLog.objects.filter(
                objeto_id=str(liberacion.pk),
                resumen="Clase liberada revertida",
            ).exists()
        )

    def test_reversa_clase_liberada_sin_derecho_genera_deuda(self):
        asistencia = Asistencia.objects.create(
            sesion=self.sesion,
            persona=self.estudiante,
            estado=Asistencia.Estado.JUSTIFICADA,
        )
        liberacion, consumo = liberar_clase(
            asistencia=asistencia,
            motivo="Excepción sin pago",
            usuario=self.admin,
        )
        self.assertTrue(liberacion.activa)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.PENDIENTE)
        self.assertIsNone(consumo.pago)

        _, consumo = revertir_clase_liberada(
            asistencia=asistencia,
            usuario=self.admin,
        )
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.DEUDA)
        self.assertIsNone(consumo.pago)

    def test_liberar_clase_exige_motivo_y_es_idempotente_controlado(self):
        asistencia = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        with self.assertRaisesMessage(ValidationError, "motivo"):
            liberar_clase(asistencia=asistencia, motivo="", usuario=self.admin)
        liberar_clase(asistencia=asistencia, motivo="Primera liberación", usuario=self.admin)
        with self.assertRaisesMessage(ValidationError, "ya tiene"):
            liberar_clase(asistencia=asistencia, motivo="Duplicada", usuario=self.admin)
        self.assertEqual(ClaseLiberada.objects.filter(asistencia=asistencia).count(), 1)

    def test_matriz_permisos_sesion_por_asignacion_y_organizacion(self):
        url = reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk})
        casos_get = (
            (self.admin, 200),
            (self.profesor_asignado, 200),
            (self.profesor_no_asignado, 404),
            (self.profesor_otra_org, 404),
            (self.admin_otra_org, 404),
        )
        for usuario, esperado in casos_get:
            with self.subTest(usuario=usuario.username):
                self.client.force_login(usuario)
                response = self.client.get(url, {"organizacion": self.organizacion.pk})
                self.assertEqual(response.status_code, esperado)

        self.client.force_login(self.profesor_asignado)
        response = self.client.post(
            url + f"?organizacion={self.organizacion.pk}",
            {
                "cambiar_estado_asistencia": "1",
                "asistencia_id": Asistencia.objects.create(
                    sesion=self.sesion,
                    persona=self.estudiante,
                    estado=Asistencia.Estado.AUSENTE,
                ).pk,
                "estado_asistencia": Asistencia.Estado.PRESENTE,
            },
        )
        self.assertEqual(response.status_code, 302)

    def test_profesora_no_puede_liberar_clase(self):
        asistencia = Asistencia.objects.create(sesion=self.sesion, persona=self.estudiante)
        self.client.force_login(self.profesor_asignado)
        response = self.client.post(
            reverse("asistencias:sesion_detail", kwargs={"pk": self.sesion.pk})
            + f"?organizacion={self.organizacion.pk}",
            {
                "liberar_clase": "1",
                "asistencia_id": asistencia.pk,
                "motivo_liberacion": "No autorizado",
            },
        )
        self.assertEqual(response.status_code, 403)
        self.assertFalse(ClaseLiberada.objects.filter(asistencia=asistencia).exists())


class SprintTresJornadaMovilTests(TestCase):
    def setUp(self):
        self.hoy = timezone.localdate()
        self.organizacion = Organizacion.objects.create(
            nombre="Org Jornada",
            razon_social="Org Jornada SpA",
            rut="72.000.000-1",
        )
        self.otra_organizacion = Organizacion.objects.create(
            nombre="Otra Jornada",
            razon_social="Otra Jornada SpA",
            rut="72.000.000-2",
        )
        self.rol_profesor = Rol.objects.create(
            nombre="Profesora Jornada",
            codigo="PROFESOR",
        )
        self.rol_estudiante = Rol.objects.create(
            nombre="Estudiante Jornada",
            codigo="ESTUDIANTE",
        )
        self.profesora_asignada = crear_usuario_con_rol(
            username="profesora_jornada",
            password=TEST_PASSWORD,
            rol=self.rol_profesor,
            organizacion=self.organizacion,
            nombres="Paula",
            apellidos="Asignada",
        )
        self.profesora_no_asignada = crear_usuario_con_rol(
            username="profesora_no_asignada_jornada",
            password=TEST_PASSWORD,
            rol=self.rol_profesor,
            organizacion=self.organizacion,
            nombres="Nora",
            apellidos="No asignada",
        )
        self.profesora_otra_organizacion = crear_usuario_con_rol(
            username="profesora_otra_jornada",
            password=TEST_PASSWORD,
            rol=self.rol_profesor,
            organizacion=self.otra_organizacion,
            nombres="Olga",
            apellidos="Otra",
        )
        self.estudiante = Persona.objects.create(
            nombres="Elena",
            apellidos="Elegible",
            email="elena.jornada@example.com",
        )
        PersonaRol.objects.create(
            persona=self.estudiante,
            rol=self.rol_estudiante,
            organizacion=self.organizacion,
            activo=True,
        )
        self.estudiante_otra_org = Persona.objects.create(
            nombres="Alicia",
            apellidos="Ajena",
            email="alicia.ajena@example.com",
        )
        PersonaRol.objects.create(
            persona=self.estudiante_otra_org,
            rol=self.rol_estudiante,
            organizacion=self.otra_organizacion,
            activo=True,
        )
        self.disciplina = Disciplina.objects.create(
            organizacion=self.organizacion,
            nombre="Yoga Jornada",
        )
        self.disciplina_otra = Disciplina.objects.create(
            organizacion=self.otra_organizacion,
            nombre="Pilates Ajeno",
        )
        self.bloque_temprano = BloqueHorario.objects.create(
            organizacion=self.organizacion,
            nombre="Temprano",
            dia_semana=self.hoy.weekday(),
            hora_inicio="09:00",
            hora_fin="10:00",
            disciplina=self.disciplina,
        )
        self.bloque_tarde = BloqueHorario.objects.create(
            organizacion=self.organizacion,
            nombre="Tarde",
            dia_semana=self.hoy.weekday(),
            hora_inicio="18:00",
            hora_fin="19:00",
            disciplina=self.disciplina,
        )
        self.sesion_tarde = SesionClase.objects.create(
            disciplina=self.disciplina,
            bloque=self.bloque_tarde,
            fecha=self.hoy,
        )
        self.sesion_temprano = SesionClase.objects.create(
            disciplina=self.disciplina,
            bloque=self.bloque_temprano,
            fecha=self.hoy,
        )
        self.sesion_no_asignada = SesionClase.objects.create(
            disciplina=self.disciplina,
            fecha=self.hoy,
        )
        self.sesion_otra_org = SesionClase.objects.create(
            disciplina=self.disciplina_otra,
            fecha=self.hoy,
        )
        asignar_profesora_a_sesion(
            user=self.profesora_asignada,
            sesion=self.sesion_temprano,
        )
        asignar_profesora_a_sesion(
            user=self.profesora_asignada,
            sesion=self.sesion_tarde,
        )
        asignar_profesora_a_sesion(
            user=self.profesora_otra_organizacion,
            sesion=self.sesion_otra_org,
        )

    def _login_asignada(self):
        self.client.force_login(self.profesora_asignada)

    def test_integracion_jornada_usa_postgresql(self):
        self.assertEqual(connection.vendor, "postgresql")
        self.assertEqual(connection.settings_dict["ENGINE"], "django.db.backends.postgresql")

    def test_hoy_muestra_solo_sesiones_asignadas_en_orden_cronologico(self):
        self._login_asignada()
        response = self.client.get(reverse("asistencias:sesiones_hoy"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [sesion.pk for sesion in response.context["sesiones"]],
            [self.sesion_temprano.pk, self.sesion_tarde.pk],
        )
        self.assertContains(response, "Yoga Jornada")
        self.assertContains(response, "Org Jornada")
        self.assertContains(response, "09:00")
        self.assertContains(response, "18:00")
        self.assertContains(response, "Paula Asignada")
        self.assertNotContains(response, "Pilates Ajeno")
        self.assertNotContains(
            response,
            reverse(
                "asistencias:sesion_detail",
                kwargs={"pk": self.sesion_no_asignada.pk},
            ),
        )

    def test_hoy_muestra_estado_vacio_comprensible(self):
        self.client.force_login(self.profesora_no_asignada)
        response = self.client.get(reverse("asistencias:sesiones_hoy"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No tienes clases asignadas hoy")
        self.assertEqual(list(response.context["sesiones"]), [])

    def test_detalle_restringe_profesora_no_asignada_y_otra_organizacion(self):
        url = reverse(
            "asistencias:sesion_detail",
            kwargs={"pk": self.sesion_temprano.pk},
        )
        for usuaria, esperado in (
            (self.profesora_asignada, 200),
            (self.profesora_no_asignada, 404),
            (self.profesora_otra_organizacion, 404),
        ):
            with self.subTest(usuaria=usuaria.username):
                self.client.force_login(usuaria)
                self.assertEqual(self.client.get(url).status_code, esperado)

    def test_desactivar_user_corta_lectura_y_escritura_con_sesion_abierta(self):
        self._login_asignada()
        self.assertEqual(self.client.get(reverse("asistencias:sesiones_hoy")).status_code, 200)
        self.profesora_asignada.is_active = False
        self.profesora_asignada.save(update_fields=["is_active"])

        lectura = self.client.get(reverse("asistencias:sesiones_hoy"))
        escritura = self.client.post(
            reverse(
                "asistencias:sesion_asistente_agregar",
                kwargs={"pk": self.sesion_temprano.pk},
            ),
            {"persona_id": self.estudiante.pk},
        )

        self.assertEqual(lectura.status_code, 302)
        self.assertIn(reverse("login"), lectura.url)
        self.assertEqual(escritura.status_code, 403)
        self.assertEqual(escritura.json()["codigo"], "PERMISO_DENEGADO")
        self.assertFalse(
            Asistencia.objects.filter(
                sesion=self.sesion_temprano,
                persona=self.estudiante,
            ).exists()
        )

    def test_desactivar_rol_corta_lectura_y_escritura_con_sesion_abierta(self):
        self._login_asignada()
        rol = PersonaRol.objects.get(
            persona__user=self.profesora_asignada,
            organizacion=self.organizacion,
            rol=self.rol_profesor,
        )
        rol.activo = False
        rol.save(update_fields=["activo"])

        lectura = self.client.get(
            reverse(
                "asistencias:sesion_detail",
                kwargs={"pk": self.sesion_temprano.pk},
            )
        )
        escritura = self.client.post(
            reverse(
                "asistencias:sesion_asistente_agregar",
                kwargs={"pk": self.sesion_temprano.pk},
            ),
            {"persona_id": self.estudiante.pk},
        )

        self.assertEqual(lectura.status_code, 404)
        self.assertEqual(escritura.status_code, 403)
        self.assertEqual(escritura.json()["codigo"], "PERMISO_DENEGADO")
        self.assertFalse(
            Asistencia.objects.filter(
                sesion=self.sesion_temprano,
                persona=self.estudiante,
            ).exists()
        )

    def test_quitar_asignacion_corta_lectura_y_escritura_con_sesion_abierta(self):
        self._login_asignada()
        self.sesion_temprano.profesores.remove(self.profesora_asignada.persona)

        lectura = self.client.get(
            reverse(
                "asistencias:sesion_detail",
                kwargs={"pk": self.sesion_temprano.pk},
            )
        )
        escritura = self.client.post(
            reverse(
                "asistencias:sesion_asistente_agregar",
                kwargs={"pk": self.sesion_temprano.pk},
            ),
            {"persona_id": self.estudiante.pk},
        )

        self.assertEqual(lectura.status_code, 404)
        self.assertEqual(escritura.status_code, 404)
        self.assertEqual(escritura.json()["codigo"], "SESION_NO_ENCONTRADA")
        self.assertFalse(
            Asistencia.objects.filter(
                sesion=self.sesion_temprano,
                persona=self.estudiante,
            ).exists()
        )

    def test_matriz_gate2_html_busqueda_json_y_post_por_actor(self):
        User = get_user_model()
        sin_rol = User.objects.create_user("jornada_sin_rol", password=TEST_PASSWORD)
        Persona.objects.create(
            nombres="Usuario",
            apellidos="Sin rol",
            email="jornada.sin.rol@example.com",
            user=sin_rol,
        )
        rol_inactivo = crear_usuario_con_rol(
            username="profesora_rol_inactivo",
            password=TEST_PASSWORD,
            rol=self.rol_profesor,
            organizacion=self.organizacion,
            nombres="Profesora",
            apellidos="Rol inactivo",
        )
        PersonaRol.objects.filter(
            persona__user=rol_inactivo,
            organizacion=self.organizacion,
            rol=self.rol_profesor,
        ).update(activo=False)
        rol_admin = Rol.objects.create(nombre="Administración jornada", codigo="ADMINISTRADOR")
        administracion = crear_usuario_con_rol(
            username="administracion_jornada",
            password=TEST_PASSWORD,
            rol=rol_admin,
            organizacion=self.organizacion,
            nombres="Administración",
            apellidos="Jornada",
        )
        emergencia = User.objects.create_superuser(
            "emergencia_jornada",
            email="emergencia.jornada@example.com",
            password=TEST_PASSWORD,
        )
        detalle_url = reverse(
            "asistencias:sesion_detail",
            kwargs={"pk": self.sesion_temprano.pk},
        )
        busqueda_url = reverse(
            "asistencias:sesion_asistentes_buscar",
            kwargs={"pk": self.sesion_temprano.pk},
        )
        agregar_url = reverse(
            "asistencias:sesion_asistente_agregar",
            kwargs={"pk": self.sesion_temprano.pk},
        )
        casos = (
            ("anonimo", None, 302, 403, 403),
            ("sin_rol", sin_rol, 404, 403, 403),
            ("rol_inactivo", rol_inactivo, 404, 403, 403),
            ("profesora_asignada", self.profesora_asignada, 200, 200, 400),
            ("profesora_no_asignada", self.profesora_no_asignada, 404, 404, 404),
            ("otra_organizacion", self.profesora_otra_organizacion, 404, 404, 404),
            ("administracion", administracion, 200, 200, 400),
            ("superusuario_emergencia", emergencia, 200, 200, 400),
        )

        for actor, usuario, html_status, busqueda_status, post_status in casos:
            with self.subTest(actor=actor):
                self.client.logout()
                if usuario:
                    self.client.force_login(usuario)
                query = {"organizacion": self.organizacion.pk} if actor == "administracion" else {}
                detalle = self.client.get(detalle_url, query)
                busqueda = self.client.get(busqueda_url, {"q": "Elena"})
                agregar = self.client.post(agregar_url, {"persona_id": 999999})
                self.assertEqual(detalle.status_code, html_status)
                self.assertEqual(busqueda.status_code, busqueda_status)
                self.assertEqual(agregar.status_code, post_status)
                if post_status == 400:
                    self.assertEqual(agregar.json()["codigo"], "PERSONA_INVALIDA")

    def test_busqueda_profesora_limita_resultados_a_organizacion_y_sesion(self):
        Asistencia.objects.create(
            sesion=self.sesion_temprano,
            persona=self.estudiante,
        )
        elegible = Persona.objects.create(
            nombres="Elena",
            apellidos="Disponible",
            email="elena.disponible@example.com",
        )
        PersonaRol.objects.create(
            persona=elegible,
            rol=self.rol_estudiante,
            organizacion=self.organizacion,
            activo=True,
        )
        self._login_asignada()
        response = self.client.get(
            reverse(
                "asistencias:sesion_asistentes_buscar",
                kwargs={"pk": self.sesion_temprano.pk},
            ),
            {"q": "Elena"},
        )

        self.assertEqual(response.status_code, 200)
        resultados = response.json()["resultados"]
        self.assertEqual([item["id"] for item in resultados], [elegible.pk])
        self.assertNotIn(self.estudiante_otra_org.pk, [item["id"] for item in resultados])
        self.assertEqual(set(resultados[0]), {"id", "nombre", "inactivo"})

    def test_busqueda_directa_sesion_no_asignada_es_indistinguible_de_inexistente(self):
        self._login_asignada()
        responses = (
            self.client.get(
                reverse(
                    "asistencias:sesion_asistentes_buscar",
                    kwargs={"pk": self.sesion_no_asignada.pk},
                ),
                {"q": "Elena"},
            ),
            self.client.get(
                reverse(
                    "asistencias:sesion_asistentes_buscar",
                    kwargs={"pk": 999991},
                ),
                {"q": "Elena"},
            ),
        )
        for response in responses:
            self.assertEqual(response.status_code, 404)
            self.assertEqual(response.json()["codigo"], "SESION_NO_ENCONTRADA")

    def test_agregado_profesora_crea_consumo_mensual_sin_exponer_finanzas(self):
        pago = Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.organizacion,
            fecha_pago=self.hoy,
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=1,
        )
        self._login_asignada()
        response = self.client.post(
            reverse(
                "asistencias:sesion_asistente_agregar",
                kwargs={"pk": self.sesion_temprano.pk},
            ),
            {"persona_id": self.estudiante.pk},
        )

        self.assertEqual(response.status_code, 201)
        asistencia = Asistencia.objects.get(
            sesion=self.sesion_temprano,
            persona=self.estudiante,
        )
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.CONSUMIDO)
        self.assertEqual(consumo.pago, pago)
        self.assertIsNone(response.json()["estado_financiero"])
        self.assertNotIn("persona_url", response.json()["asistencia"])

    def test_agregado_profesora_rechaza_persona_ajena_y_reintento(self):
        self._login_asignada()
        url = reverse(
            "asistencias:sesion_asistente_agregar",
            kwargs={"pk": self.sesion_temprano.pk},
        )
        ajena = self.client.post(
            url,
            {"persona_id": self.estudiante_otra_org.pk},
        )
        primera = self.client.post(url, {"persona_id": self.estudiante.pk})
        reintento = self.client.post(url, {"persona_id": self.estudiante.pk})

        self.assertEqual(ajena.status_code, 400)
        self.assertEqual(ajena.json()["codigo"], "PERSONA_INVALIDA")
        self.assertEqual(primera.status_code, 201)
        self.assertEqual(reintento.status_code, 409)
        self.assertEqual(reintento.json()["codigo"], "ASISTENCIA_DUPLICADA")
        self.assertEqual(
            Asistencia.objects.filter(
                sesion=self.sesion_temprano,
                persona=self.estudiante,
            ).count(),
            1,
        )

    def test_edicion_rapida_es_idempotente_y_no_recupera_cupo(self):
        pago = Payment.objects.create(
            persona=self.estudiante,
            organizacion=self.organizacion,
            fecha_pago=self.hoy,
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=1,
        )
        asistencia = Asistencia.objects.create(
            sesion=self.sesion_temprano,
            persona=self.estudiante,
        )
        self._login_asignada()
        url = reverse(
            "asistencias:sesion_asistencia_estado",
            kwargs={
                "pk": self.sesion_temprano.pk,
                "asistencia_pk": asistencia.pk,
            },
        )

        ausencia = self.client.post(url, {"estado": Asistencia.Estado.AUSENTE})
        repeticion = self.client.post(url, {"estado": Asistencia.Estado.AUSENTE})
        justificada = self.client.post(
            url,
            {"estado": Asistencia.Estado.JUSTIFICADA},
        )

        self.assertEqual(ausencia.status_code, 200)
        self.assertEqual(repeticion.status_code, 200)
        self.assertEqual(justificada.status_code, 200)
        self.assertIsNone(justificada.json()["estado_financiero"])
        self.assertNotIn("estado_financiero", justificada.json()["asistencia"])
        asistencia.refresh_from_db()
        consumo = AttendanceConsumption.objects.get(asistencia=asistencia)
        self.assertEqual(asistencia.estado, Asistencia.Estado.JUSTIFICADA)
        self.assertEqual(consumo.estado, AttendanceConsumption.Estado.CONSUMIDO)
        self.assertEqual(consumo.pago, pago)
        self.assertEqual(
            AttendanceConsumption.objects.filter(asistencia=asistencia).count(),
            1,
        )
        pago.refresh_from_db()
        self.assertEqual(pago.saldo_clases, 0)

    def test_edicion_rapida_rechaza_identificador_ajeno_y_estado_invalido(self):
        asistencia_ajena = Asistencia.objects.create(
            sesion=self.sesion_otra_org,
            persona=self.estudiante_otra_org,
        )
        self._login_asignada()
        url_ajena = reverse(
            "asistencias:sesion_asistencia_estado",
            kwargs={
                "pk": self.sesion_temprano.pk,
                "asistencia_pk": asistencia_ajena.pk,
            },
        )
        url_valida = reverse(
            "asistencias:sesion_asistencia_estado",
            kwargs={
                "pk": self.sesion_temprano.pk,
                "asistencia_pk": 999992,
            },
        )

        ajena = self.client.post(
            url_ajena,
            {"estado": Asistencia.Estado.AUSENTE},
        )
        inexistente = self.client.post(
            url_valida,
            {"estado": Asistencia.Estado.AUSENTE},
        )
        invalido = self.client.post(
            reverse(
                "asistencias:sesion_asistencia_estado",
                kwargs={
                    "pk": self.sesion_temprano.pk,
                    "asistencia_pk": Asistencia.objects.create(
                        sesion=self.sesion_temprano,
                        persona=self.estudiante,
                    ).pk,
                },
            ),
            {"estado": "desconocido"},
        )

        self.assertEqual(ajena.status_code, 404)
        self.assertEqual(ajena.json()["codigo"], "ASISTENCIA_NO_ENCONTRADA")
        self.assertEqual(inexistente.status_code, 404)
        self.assertEqual(inexistente.json()["codigo"], "ASISTENCIA_NO_ENCONTRADA")
        self.assertEqual(invalido.status_code, 400)
        self.assertEqual(invalido.json()["codigo"], "ESTADO_INVALIDO")

        self.client.force_login(self.profesora_no_asignada)
        no_asignada = self.client.post(
            reverse(
                "asistencias:sesion_asistencia_estado",
                kwargs={
                    "pk": self.sesion_temprano.pk,
                    "asistencia_pk": Asistencia.objects.get(
                        sesion=self.sesion_temprano,
                        persona=self.estudiante,
                    ).pk,
                },
            ),
            {"estado": Asistencia.Estado.AUSENTE},
        )
        self.assertEqual(no_asignada.status_code, 404)
        self.assertEqual(no_asignada.json()["codigo"], "SESION_NO_ENCONTRADA")

    def test_profesora_no_puede_quitar_liberar_ni_ver_detalles_financieros(self):
        asistencia = Asistencia.objects.create(
            sesion=self.sesion_temprano,
            persona=self.estudiante,
        )
        liberar_clase(
            asistencia=asistencia,
            motivo="Motivo administrativo reservado",
            usuario=self.profesora_asignada,
        )
        self._login_asignada()
        url = reverse(
            "asistencias:sesion_detail",
            kwargs={"pk": self.sesion_temprano.pk},
        )
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["puede_quitar_asistente"])
        self.assertFalse(response.context["puede_liberar_clase"])
        self.assertFalse(response.context["puede_administrar_sesion"])
        self.assertNotContains(response, "Estado de pago")
        self.assertNotContains(response, reverse("personas:persona_detail", args=[self.estudiante.pk]))
        self.assertContains(response, "Clase liberada")
        self.assertNotContains(response, "Motivo administrativo reservado")
        self.assertContains(response, "data-estado-control")
        self.assertContains(response, 'aria-live="polite"', html=False)
        self.assertContains(response, 'aria-pressed="true"', html=False)

        eliminar = self.client.post(
            url,
            {
                "eliminar_asistente": "1",
                "asistencia_id": asistencia.pk,
            },
        )
        self.assertEqual(eliminar.status_code, 403)
        self.assertTrue(Asistencia.objects.filter(pk=asistencia.pk).exists())

    def test_navegacion_profesora_expone_hoy_sin_panel_administrativo(self):
        self._login_asignada()
        response = self.client.get(reverse("asistencias:sesiones_hoy"))

        self.assertContains(response, reverse("asistencias:sesiones_hoy"))
        self.assertNotContains(
            response,
            f'href="{reverse("asistencias:dashboard")}"',
            html=False,
        )
        self.assertNotContains(
            response,
            f'href="{reverse("finanzas:dashboard")}"',
            html=False,
        )
        self.assertContains(response, 'aria-label="Abrir menú"', html=False)
        self.assertContains(response, 'min-height: 44px', html=False)


class SprintDosConcurrenciaConsumosTests(TransactionTestCase):
    reset_sequences = True

    def test_dos_asistencias_compiten_por_un_cupo_sin_sobreconsumo(self):
        organizacion = Organizacion.objects.create(
            nombre="Org Concurrencia Sprint 2",
            razon_social="Org Concurrencia Sprint 2 SpA",
            rut="70.000.000-3",
        )
        disciplina = Disciplina.objects.create(organizacion=organizacion, nombre="Concurrencia")
        estudiante = Persona.objects.create(nombres="Estudiante", apellidos="Concurrente")
        pago = Payment.objects.create(
            persona=estudiante,
            organizacion=organizacion,
            fecha_pago=date(2026, 7, 1),
            metodo_pago=Payment.Metodo.EFECTIVO,
            aplica_iva=False,
            monto_referencia=10000,
            clases_asignadas=0,
        )
        asistencias = []
        for dia in (10, 11):
            sesion = SesionClase.objects.create(
                disciplina=disciplina,
                fecha=date(2026, 7, dia),
            )
            asistencias.append(
                Asistencia.objects.create(
                    sesion=sesion,
                    persona=estudiante,
                    estado=Asistencia.Estado.AUSENTE,
                )
            )
        Payment.objects.filter(pk=pago.pk).update(clases_asignadas=1)

        def recalcular_consumo(asistencia_id):
            close_old_connections()
            try:
                asistencia = Asistencia.objects.get(pk=asistencia_id)
                asignar_consumo_asistencia(asistencia)
            finally:
                close_old_connections()

        with ThreadPoolExecutor(max_workers=2) as executor:
            list(executor.map(recalcular_consumo, [item.pk for item in asistencias]))

        estados = list(
            AttendanceConsumption.objects.filter(asistencia__in=asistencias)
            .order_by("id")
            .values_list("estado", flat=True)
        )
        self.assertCountEqual(
            estados,
            [
                AttendanceConsumption.Estado.CONSUMIDO,
                AttendanceConsumption.Estado.DEUDA,
            ],
        )
        self.assertEqual(
            pago.consumos.filter(estado=AttendanceConsumption.Estado.CONSUMIDO).count(),
            1,
        )
