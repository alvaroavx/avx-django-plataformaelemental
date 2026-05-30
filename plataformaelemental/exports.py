from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def periodo_sufijo_archivo(periodo):
    mes = periodo.get("mes")
    anio = periodo.get("anio")
    if mes is None and anio is None:
        return "todos"
    if mes is None:
        return f"{anio}_todos"
    if anio is None:
        return f"todos_{mes:02d}"
    return f"{anio}_{mes:02d}"


def xlsx_response(*, filename, sheet_title, headers, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title[:31]

    worksheet.append(headers)
    header_fill = PatternFill("solid", fgColor="E9ECEF")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in rows:
        worksheet.append(row)

    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 45)

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
